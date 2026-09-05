import os
import zipfile
from pathlib import Path
import sqlite3
from typing import Optional
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

from modules.common.dbutil import get_conn, table_exists, get_table_columns, find_existing_column
from modules.common.fmt import safe_float as _safe_float, safe_int as _safe_int, fmt_krw as _currency

BASE_DIR = Path(__file__).resolve().parents[2]

STATEMENT_COMPANY_NAME = "㈜ 데일리시가"
STATEMENT_BANK_NAME = "신한은행"
STATEMENT_BANK_ACCOUNT = "140-015-512046"
STATEMENT_ACCOUNT_HOLDER = "㈜데일리시가"
STATEMENT_DOC_PREFIX = "TS"
STATEMENT_PAYMENT_DUE_DAYS = 3
STATEMENT_TEMPLATE_DIR = BASE_DIR / "templates"
STATEMENT_TEMPLATE_PATH = Path(os.getenv("DAILYCIGAR_STATEMENT_TEMPLATE_PATH", str(STATEMENT_TEMPLATE_DIR / "거래명세서_template.xlsx")))
# -----------------------------
# DB helpers
# -----------------------------


def set_merged_cell_value(ws, cell_ref: str, value):
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                ws[merged_range.start_cell.coordinate] = value
                return
    ws[cell_ref] = value

def _apply_alignment_safe(ws, cell_ref: str, alignment):
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell_ref in merged_range:
                ws[merged_range.start_cell.coordinate].alignment = alignment
                return
    else:
        cell.alignment = alignment


def ensure_required_tables(conn):
    required_tables = ["partner_mst", "wholesale_sales", "product_mst"]
    return [t for t in required_tables if not table_exists(conn, t)]


def ensure_wholesale_sales_columns(conn):
    existing = set(get_table_columns(conn, "wholesale_sales"))
    required = {
        "supply_price": "REAL",
        "supply_amount": "REAL",
        "vat_amount": "REAL",
        "total_amount_vat": "REAL",
        "grade_code_applied": "TEXT",
        "discount_rate_applied": "REAL",
        "normal_supply_amount": "REAL",
        "updated_at": "TEXT",
    }
    cur = conn.cursor()
    changed = False
    for col, col_type in required.items():
        if col not in existing:
            cur.execute(f"ALTER TABLE wholesale_sales ADD COLUMN {col} {col_type}")
            changed = True
    if changed:
        conn.commit()


def ensure_statement_tables(conn):
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS statement_doc_sequence (
            doc_type TEXT NOT NULL,
            doc_date TEXT NOT NULL,
            last_seq INTEGER NOT NULL DEFAULT 0,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (doc_type, doc_date)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS statement_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_no TEXT NOT NULL UNIQUE,
            sale_date TEXT NOT NULL,
            due_date TEXT,
            partner_id INTEGER NOT NULL,
            partner_name TEXT,
            file_name TEXT,
            notes TEXT,
            supply_amount REAL DEFAULT 0,
            vat_amount REAL DEFAULT 0,
            total_amount REAL DEFAULT 0,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )

    conn.commit()

def load_partner_purchase_summary_from_grade(conn) -> pd.DataFrame:
    """
    거래처별 "현재 등급을 달성한 시점 이후" 누적 구매 요약.

    과거에는 partner_mst.grade_acquired_date(수동 편집 폼에서만 갱신되는 필드,
    실제 등급업 처리 화면인 '거래처등급관리'에서는 갱신되지 않아 사실상 가입일에
    고정되어 있었음)를 기준으로 계산해서 부정확했다. 지금은 도매 가격 계산에
    실제로 쓰이는 것과 동일한 compute_grade_reset_cumulative() 엔진을 재사용해,
    각 거래처의 "현재 등급 시작일"과 그 이후 누적 공급가를 정확히 구한다.
    """
    ensure_wholesale_sales_columns(conn)

    result_cols = [
        "partner_id",
        "purchase_item_names",
        "purchase_supply_amount_sum",
        "grade_start_date",
        "next_grade_code",
        "next_grade_remaining",
    ]

    if not table_exists(conn, "partner_mst") or not table_exists(conn, "wholesale_sales"):
        return pd.DataFrame(columns=result_cols)

    partners_df = load_partners(conn)
    if partners_df.empty:
        return pd.DataFrame(columns=result_cols)

    thresholds = load_partner_grade_thresholds(conn)
    today_str = pd.Timestamp.today().strftime("%Y-%m-%d")

    sales_df = load_wholesale_sales_for_grid(conn)
    if sales_df.empty:
        sales_df = pd.DataFrame(columns=["partner_id", "sale_date", "supply_amount", "product_name"])
    sales_df["supply_amount"] = pd.to_numeric(sales_df.get("supply_amount", 0), errors="coerce").fillna(0)
    sales_df["product_name"] = sales_df.get("product_name", "").fillna("").astype(str).str.strip()
    sales_df["sale_date"] = sales_df.get("sale_date", "").astype(str)

    tier_rows = thresholds.sort_values("min_purchase_amount").reset_index(drop=True) if not thresholds.empty else thresholds
    grade_codes_list = tier_rows["grade_code"].tolist() if not tier_rows.empty else []
    grade_thresholds_list = tier_rows["min_purchase_amount"].tolist() if not tier_rows.empty else []

    def _join_unique_names(series):
        vals = []
        seen = set()
        for x in series:
            x = str(x or "").strip()
            if x and x not in seen:
                vals.append(x)
                seen.add(x)
        return ", ".join(vals)

    rows = []
    for _, prow in partners_df.iterrows():
        partner_id = int(prow["id"])
        join_date = str(prow.get("join_date") or today_str)

        if thresholds.empty:
            grade_start_date = join_date
            cumulative = 0.0
            next_grade_code = None
            next_remaining = None
        else:
            progress = compute_grade_reset_cumulative(conn, partner_id, today_str, join_date, thresholds)
            grade_start_date = progress["grade_start_date"]
            cumulative = progress["cumulative_since_start"]
            cur_idx = grade_codes_list.index(progress["grade_code"]) if progress["grade_code"] in grade_codes_list else -1
            if cur_idx >= 0 and cur_idx + 1 < len(grade_codes_list):
                next_grade_code = grade_codes_list[cur_idx + 1]
                next_remaining = max(grade_thresholds_list[cur_idx + 1] - cumulative, 0)
            else:
                next_grade_code = None
                next_remaining = None

        p_sales = sales_df[(sales_df["partner_id"] == partner_id) & (sales_df["sale_date"] >= grade_start_date)]
        item_names = _join_unique_names(p_sales["product_name"])

        rows.append({
            "partner_id": partner_id,
            "purchase_item_names": item_names,
            "purchase_supply_amount_sum": cumulative,
            "grade_start_date": grade_start_date,
            "next_grade_code": next_grade_code,
            "next_grade_remaining": next_remaining,
        })

    return pd.DataFrame(rows, columns=result_cols)

def issue_statement_document_no(conn, sale_date_str: str) -> str:
    return issue_document_no(conn, "STATEMENT", STATEMENT_DOC_PREFIX, sale_date_str)

def save_statement_history(
    conn,
    header: dict,
    totals: dict,
    file_name: str,
):
    ensure_statement_tables(conn)

    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO statement_history (
            document_no,
            sale_date,
            due_date,
            partner_id,
            partner_name,
            file_name,
            notes,
            supply_amount,
            vat_amount,
            total_amount
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            header.get("document_no", ""),
            header.get("sale_date", ""),
            header.get("due_date", ""),
            _safe_int(header.get("partner_id", 0), 0),
            header.get("partner_name", ""),
            file_name,
            header.get("notes", ""),
            _safe_float(totals.get("supply_amount", 0), 0),
            _safe_float(totals.get("vat_amount", 0), 0),
            _safe_float(totals.get("total_amount", 0), 0),
        ),
    )
    conn.commit()


# -----------------------------
# Loaders
# -----------------------------
def load_partners(conn) -> pd.DataFrame:
    cols = get_table_columns(conn, "partner_mst")
    select_cols = [
        "id",
        "partner_name",
        "partner_type",
        "business_no",
        "owner_name",
        "contact_name",
        "phone",
        "email",
        "address",
        "current_grade_code",
        "status",
        "join_date",
        "grade_acquired_date",
        "notes",
        "program_exempt_yn",
    ]
    final_select = []
    for col in select_cols:
        if col in cols:
            final_select.append(col)
        else:
            final_select.append(f"NULL AS {col}")

    sql = f"""
        SELECT {', '.join(final_select)}
        FROM partner_mst
        ORDER BY partner_name
    """
    return pd.read_sql(sql, conn)

def load_partner_last_purchase(conn) -> pd.DataFrame:
    """거래처별 마지막 도매 구매일 / 구매일수(거래 횟수)."""
    empty = pd.DataFrame(columns=["partner_id", "last_purchase_date", "purchase_count"])
    if not table_exists(conn, "wholesale_sales"):
        return empty
    sql = """
        SELECT
            partner_id,
            MAX(sale_date) AS last_purchase_date,
            COUNT(DISTINCT sale_date) AS purchase_count
        FROM wholesale_sales
        WHERE partner_id IS NOT NULL AND COALESCE(sale_date, '') <> ''
        GROUP BY partner_id
    """
    try:
        df = pd.read_sql(sql, conn)
    except Exception:
        return empty
    if df.empty:
        return empty
    df["partner_id"] = pd.to_numeric(df["partner_id"], errors="coerce").astype("Int64")
    df["purchase_count"] = pd.to_numeric(df["purchase_count"], errors="coerce").fillna(0).astype(int)
    return df


PARTNER_SORT_RECENT = "최근구매순"
PARTNER_SORT_NAME = "거래처명순"
PARTNER_SORT_AMOUNT = "구매금액순"
PARTNER_SORT_OPTIONS = (PARTNER_SORT_RECENT, PARTNER_SORT_NAME, PARTNER_SORT_AMOUNT)


def sort_partners(df: pd.DataFrame, sort_key: str = PARTNER_SORT_RECENT) -> pd.DataFrame:
    """
    거래처 목록 정렬 (순수 함수).
    - 최근구매순: 마지막 구매일 내림차순, 구매 이력이 없는 거래처는 맨 뒤, 동률은 거래처명순
    - 거래처명순 / 구매금액순(등급업 이후 공급가합계 내림차순)
    """
    if df is None or df.empty:
        return df

    out = df.copy()
    name = out["partner_name"].fillna("") if "partner_name" in out.columns else pd.Series("", index=out.index)

    if sort_key == PARTNER_SORT_NAME or "last_purchase_date" not in out.columns:
        if sort_key == PARTNER_SORT_AMOUNT and "purchase_supply_amount_sum" in out.columns:
            amount = pd.to_numeric(out["purchase_supply_amount_sum"], errors="coerce").fillna(0)
            return out.assign(_a=amount, _n=name).sort_values(["_a", "_n"], ascending=[False, True]).drop(columns=["_a", "_n"])
        return out.assign(_n=name).sort_values("_n").drop(columns=["_n"])

    if sort_key == PARTNER_SORT_AMOUNT:
        amount = pd.to_numeric(out.get("purchase_supply_amount_sum"), errors="coerce").fillna(0)
        return out.assign(_a=amount, _n=name).sort_values(["_a", "_n"], ascending=[False, True]).drop(columns=["_a", "_n"])

    last = pd.to_datetime(out["last_purchase_date"], errors="coerce")
    return (
        out.assign(_never=last.isna(), _last=last, _n=name)
        .sort_values(["_never", "_last", "_n"], ascending=[True, False, True])
        .drop(columns=["_never", "_last", "_n"])
    )


def load_grade_codes(conn) -> list[str]:
    if not table_exists(conn, "partner_grade_mst"):
        return []

    cols = get_table_columns(conn, "partner_grade_mst")
    code_col = find_existing_column(cols, ["grade_code", "partner_grade_code", "code"])
    use_col = find_existing_column(cols, ["is_active", "use_yn", "active_yn", "use_flag"])
    sort_col = find_existing_column(cols, ["sort_order", "display_order", "order_no", "sort_no"])
    if not code_col:
        return []

    sql = f"SELECT {code_col} AS grade_code FROM partner_grade_mst"
    if use_col:
        if use_col == "is_active":
            sql += f" WHERE COALESCE({use_col}, 1) IN (1, '1', 'Y', 'y', 'TRUE', 'true')"
        else:
            sql += f" WHERE COALESCE({use_col}, 'Y') IN ('Y', 'y', '1', 1, 'TRUE', 'true')"
    if sort_col:
        sql += f" ORDER BY {sort_col}, grade_code"
    else:
        sql += " ORDER BY grade_code"

    df = pd.read_sql(sql, conn)
    if df.empty:
        return []
    return [str(x) for x in df["grade_code"].dropna().tolist()]


def ensure_partner_grade_tables(conn):
    """
    파트너 등급 이력 테이블(partner_grade_history)을 생성하고,
    partner_grade_mst에 등급 기준 누적액(min_purchase_amount) 컬럼을 보강한다.
    partner_grade_mst가 완전히 비어있을 때만 PDF 기준(Bronze~Diamond) 기본값을 시드한다.
    """
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS partner_grade_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            partner_id INTEGER NOT NULL,
            grade_code TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT,
            base_amount REAL,
            notes TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
        """
    )
    conn.commit()

    if table_exists(conn, "partner_grade_mst"):
        existing = set(get_table_columns(conn, "partner_grade_mst"))
        if "min_purchase_amount" not in existing:
            cur.execute("ALTER TABLE partner_grade_mst ADD COLUMN min_purchase_amount REAL DEFAULT 0")
            conn.commit()
        if "grade_name" not in existing:
            cur.execute("ALTER TABLE partner_grade_mst ADD COLUMN grade_name TEXT")
            conn.commit()

        # 이미 등급 데이터가 있던 테이블이라 위 ALTER로 min_purchase_amount가 전부 0으로 깔린 경우를 보정.
        # 등급코드명(대소문자 무관)으로 PDF 기준 문턱값을 매칭해서, 아직 0으로 비어있는 행만 채운다.
        pdf_thresholds_by_name = {
            "bronze": 0,
            "silver": 6_000_000,
            "gold": 12_000_000,
            "platinum": 36_000_000,
            "diamond": 60_000_000,
        }
        code_col_df = pd.read_sql("SELECT grade_code, min_purchase_amount FROM partner_grade_mst", conn)
        for _, row in code_col_df.iterrows():
            code_norm = str(row["grade_code"] or "").strip().lower()
            current_min = _safe_float(row.get("min_purchase_amount", 0), 0)
            if code_norm in pdf_thresholds_by_name and current_min == 0 and pdf_thresholds_by_name[code_norm] != 0:
                cur.execute(
                    "UPDATE partner_grade_mst SET min_purchase_amount = ? WHERE grade_code = ?",
                    (pdf_thresholds_by_name[code_norm], row["grade_code"]),
                )
        conn.commit()
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS partner_grade_mst (
                grade_code TEXT PRIMARY KEY,
                grade_name TEXT,
                min_purchase_amount REAL DEFAULT 0,
                discount_rate REAL DEFAULT 0,
                sort_order INTEGER
            )
            """
        )
        conn.commit()

    cnt = pd.read_sql("SELECT COUNT(*) AS cnt FROM partner_grade_mst", conn)["cnt"].iloc[0]
    if int(cnt) == 0:
        # discount_rate는 퍼센트 표기(정수)로 시드 - 실제 운영 테이블과 동일한 관례
        default_grades = [
            ("bronze", "Bronze", 0, 0, 1),
            ("silver", "Silver", 6_000_000, 5, 2),
            ("gold", "Gold", 12_000_000, 10, 3),
            ("platinum", "Platinum", 36_000_000, 15, 4),
            ("diamond", "Diamond", 60_000_000, 20, 5),
        ]
        cur.executemany(
            """
            INSERT INTO partner_grade_mst
                (grade_code, grade_name, min_purchase_amount, discount_rate, sort_order)
            VALUES (?, ?, ?, ?, ?)
            """,
            default_grades,
        )
        conn.commit()


def load_partner_grade_thresholds(conn) -> pd.DataFrame:
    """
    등급코드 -> 최소 누적 구매액(min_purchase_amount) / 할인율, 임계값 오름차순 정렬.

    저장된 값이 1보다 크면(예: 5, 10, 20 = 퍼센트 표기) 자동으로 100으로 나눠 소수(0.05 등)로 정규화한다.
    """
    if not table_exists(conn, "partner_grade_mst"):
        return pd.DataFrame(columns=["grade_code", "min_purchase_amount", "discount_rate"])

    cols = get_table_columns(conn, "partner_grade_mst")
    code_col = find_existing_column(cols, ["grade_code", "partner_grade_code", "code"])
    min_col = find_existing_column(cols, ["min_purchase_amount", "min_amount", "threshold_amount"])
    rate_col = find_existing_column(cols, ["discount_rate"])
    if not code_col or not min_col or not rate_col:
        return pd.DataFrame(columns=["grade_code", "min_purchase_amount", "discount_rate"])

    sql = f"""
        SELECT {code_col} AS grade_code, {min_col} AS min_purchase_amount, {rate_col} AS discount_rate
        FROM partner_grade_mst
        ORDER BY {min_col} ASC
    """
    df = pd.read_sql(sql, conn)
    if df.empty:
        return df
    df["min_purchase_amount"] = df["min_purchase_amount"].apply(_safe_float)
    df["discount_rate"] = df["discount_rate"].apply(_safe_float)
    if (df["discount_rate"] > 1).any():
        # 5, 10, 20처럼 퍼센트 단위로 저장된 값을 0.05, 0.10, 0.20 소수로 정규화
        df["discount_rate"] = df["discount_rate"] / 100.0
    return df


def load_partner_active_grade_map(conn) -> pd.DataFrame:
    """전체 거래처의 '오늘 기준' 활성 등급 이력을 한 번에 조회 (거래처 목록 표시용)"""
    ensure_partner_grade_tables(conn)
    today_str = pd.Timestamp.today().strftime("%Y-%m-%d")
    sql = """
        SELECT partner_id, grade_code, start_date, end_date
        FROM partner_grade_history
        WHERE start_date <= ?
          AND (end_date IS NULL OR end_date >= ?)
    """
    df = pd.read_sql(sql, conn, params=[today_str, today_str])
    if df.empty:
        return pd.DataFrame(columns=["partner_id", "grade_code", "start_date", "end_date"])
    return df.sort_values("start_date").drop_duplicates(subset=["partner_id"], keep="last")


def compute_grade_reset_cumulative(
    conn, partner_id: int, order_date: str, join_date: str, thresholds: pd.DataFrame
) -> dict:
    """
    "등급업 이후부터 다시 누적" 정책에 따라, 이 주문일 기준 파트너의 현재 등급과
    그 등급을 달성한 날짜 이후부터 오늘까지의 누적 구매액(공급가액, 할인 이후 금액
    기준)을 계산한다.

    가입일부터 주문일까지의 일자별 구매액을 날짜 순서대로 누적하면서, 누적액이
    다음 등급 문턱을 넘는 날짜가 나오면 그 날짜까지는 통째로 해당 등급 달성으로
    인정하고(그날 넘긴 초과분은 서비스로 간주 - 거래 단위가 아닌 날짜 단위로
    반올림), 다음날부터 누적을 0부터 다시 시작한다. 최고 등급에 도달하면 더 이상
    리셋 없이 계속 누적한다.
    """
    tier_rows = thresholds.sort_values("min_purchase_amount").reset_index(drop=True)
    grade_codes = tier_rows["grade_code"].tolist()
    grade_thresholds = tier_rows["min_purchase_amount"].tolist()

    if not grade_codes:
        return {
            "grade_code": None,
            "grade_threshold": 0.0,
            "grade_start_date": join_date,
            "cumulative_since_start": 0.0,
        }

    ensure_wholesale_sales_columns(conn)
    daily_df = pd.read_sql(
        """
        SELECT sale_date, COALESCE(SUM(supply_amount), 0) AS day_total
        FROM wholesale_sales
        WHERE partner_id = ? AND sale_date >= ? AND sale_date <= ?
        GROUP BY sale_date
        ORDER BY sale_date
        """,
        conn,
        params=[partner_id, join_date, order_date],
    )

    tier_idx = 0
    reset_date = join_date
    cumulative = 0.0

    for _, row in daily_df.iterrows():
        cumulative += _safe_float(row["day_total"], 0)
        while tier_idx + 1 < len(grade_thresholds) and cumulative >= grade_thresholds[tier_idx + 1]:
            tier_idx += 1
            reset_date = (pd.Timestamp(str(row["sale_date"])) + pd.Timedelta(days=1)).strftime("%Y-%m-%d")
            cumulative = 0.0

    return {
        "grade_code": grade_codes[tier_idx],
        "grade_threshold": grade_thresholds[tier_idx],
        "grade_start_date": reset_date,
        "cumulative_since_start": cumulative,
    }


def compute_tiered_order_pricing(
    conn,
    partner_id: int,
    qty: float,
    base_unit_price: float,
    base_supply_price: float,
    order_date: str = None,
) -> dict:
    """
    한 건의 주문(상품 1종 × 수량)에 대해 누적 구매액 문턱을 파트너 프로그램 정책대로
    계산하여, 문턱을 넘는 지점을 기준으로 구간별 할인율을 정밀 적용한 혼합(가중평균)
    단가/공급가를 산출한다.

    파트너 프로그램 정책("등급업 이후부터 다시 누적"): 등급은 한 번 오르면 절대
    내려가지 않고, 다음 등급까지 필요한 금액은 "직전 등급을 달성한 날짜 다음날부터"
    다시 0부터 채워야 한다(예: Silver 달성 후에는 그 시점부터 새로 1,200만원을
    채워야 Gold). compute_grade_reset_cumulative()가 가입일부터 오늘까지 구매
    이력을 날짜 순으로 재생(replay)하며 이 리셋 지점들을 계산한다. 문턱을 넘긴
    바로 그 날짜치는 통째로 이전 등급 달성으로 인정하고(초과분은 서비스), 다음날
    0부터 다시 센다 - 거래 단위가 아닌 날짜 단위로 리셋한다.
    한 주문이 여러 등급 문턱을 동시에 넘는 경우 각 구간의 금액에 해당 구간의
    할인율을 각각 적용한 뒤 합산한다(기존 정밀 배분 기능은 그대로 유지).
    """
    ensure_partner_grade_tables(conn)

    # 파트너 프로그램 제외 대상(이미 별도 협의로 할인 중이라 등급 할인을 중복 적용하면
    # 안 되는 업체): 매출 누적액과 무관하게 항상 기본 단가/공급가를 그대로 사용한다.
    exempt_df = pd.read_sql(
        "SELECT program_exempt_yn FROM partner_mst WHERE id = ?", conn, params=[partner_id]
    )
    if not exempt_df.empty and int(exempt_df.iloc[0]["program_exempt_yn"] or 0) == 1:
        return {
            "blended_unit_price": base_unit_price,
            "blended_supply_price": base_supply_price,
            "segments": [],
            "cumulative_before": 0.0,
            "cumulative_after": 0.0,
            "discount_ratio": 0.0,
            "final_grade_code": None,
        }

    thresholds = load_partner_grade_thresholds(conn)

    if thresholds.empty or qty is None or float(qty) <= 0:
        return {
            "blended_unit_price": base_unit_price,
            "blended_supply_price": base_supply_price,
            "segments": [],
            "cumulative_before": 0.0,
            "cumulative_after": 0.0,
            "discount_ratio": 0.0,
            "final_grade_code": None,
        }

    order_date = order_date or pd.Timestamp.today().strftime("%Y-%m-%d")

    partner_row_df = pd.read_sql("SELECT join_date FROM partner_mst WHERE id = ?", conn, params=[partner_id])
    join_date = (
        str(partner_row_df.iloc[0]["join_date"])
        if not partner_row_df.empty and pd.notna(partner_row_df.iloc[0]["join_date"])
        else order_date
    )

    if pd.to_datetime(order_date) < pd.to_datetime(join_date):
        cumulative_before = 0.0  # 주문일이 가입일보다 이전인 경우 방어
    else:
        # 등급업 리셋 이후 누적 진행상황을 계산하고, 기존 구간 배분 로직이 그대로
        # 동작하도록 "현재 등급 문턱 + 리셋 이후 누적액"을 절대 기준 누적값으로 사용
        progress = compute_grade_reset_cumulative(conn, partner_id, order_date, join_date, thresholds)
        cumulative_before = progress["grade_threshold"] + progress["cumulative_since_start"]

    order_normal_supply_amount = float(qty) * float(base_supply_price)

    tier_rows = thresholds.sort_values("min_purchase_amount").reset_index(drop=True)
    boundaries = tier_rows["min_purchase_amount"].tolist() + [float("inf")]
    rates = tier_rows["discount_rate"].tolist()
    codes = tier_rows["grade_code"].tolist()

    remaining = order_normal_supply_amount
    cursor_amount = cumulative_before
    segments = []

    for i in range(len(rates)):
        seg_lower = boundaries[i]
        seg_upper = boundaries[i + 1]
        if cursor_amount >= seg_upper:
            continue
        seg_capacity = seg_upper - max(cursor_amount, seg_lower)
        if seg_capacity <= 0:
            continue
        seg_amount = min(remaining, seg_capacity)
        if seg_amount <= 0:
            continue
        segments.append({"grade_code": codes[i], "discount_rate": rates[i], "amount": seg_amount})
        remaining -= seg_amount
        cursor_amount += seg_amount
        if remaining <= 0:
            break

    if remaining > 0:
        if segments:
            segments[-1]["amount"] += remaining
        else:
            segments.append({"grade_code": codes[-1], "discount_rate": rates[-1], "amount": remaining})
        remaining = 0

    total_discount = sum(seg["amount"] * seg["discount_rate"] for seg in segments)
    blended_supply_amount = order_normal_supply_amount - total_discount
    blended_supply_price = blended_supply_amount / float(qty) if qty else base_supply_price

    discount_ratio = (total_discount / order_normal_supply_amount) if order_normal_supply_amount > 0 else 0.0
    blended_unit_price = base_unit_price * (1 - discount_ratio)

    final_grade_code = segments[-1]["grade_code"] if segments else None

    return {
        "blended_unit_price": round(blended_unit_price),
        "blended_supply_price": round(blended_supply_price),
        "segments": segments,
        "cumulative_before": cumulative_before,
        "cumulative_after": cumulative_before + order_normal_supply_amount,
        "discount_ratio": discount_ratio,
        "final_grade_code": final_grade_code,
    }


def build_tiered_pricing_note(pricing_result: dict) -> str:
    """
    compute_tiered_order_pricing 결과 중 2개 이상 등급 구간에 걸친 주문이면
    비고(notes)에 자동으로 남길 '혼합할인' 요약 문자열을 만든다.
    한 구간에만 속하는 일반 주문이면 빈 문자열을 반환한다(자동 기록 없음).
    """
    if not pricing_result:
        return ""
    segs = [s for s in (pricing_result.get("segments") or []) if s["amount"] > 0]
    if len(segs) <= 1:
        return ""
    parts = [f"{s['grade_code']}구간 ₩{s['amount']:,.0f}×{s['discount_rate']*100:.0f}%" for s in segs]
    return "[혼합할인] " + " + ".join(parts)


def load_cigar_products_for_wholesale(conn) -> pd.DataFrame:
    product_cols = get_table_columns(conn, "product_mst")
    code_col = find_existing_column(product_cols, ["product_code", "code", "item_code"])
    name_col = find_existing_column(product_cols, ["product_name", "name"])
    use_col = find_existing_column(product_cols, ["use_yn", "is_active", "active_yn", "use_flag"])

    if not name_col:
        return pd.DataFrame(columns=["id", "product_code", "product_name", "retail_price_krw", "supply_price_krw", "korea_cost_krw"])

    select_parts = [
        "id",
        f"{code_col} AS product_code" if code_col else "'' AS product_code",
        f"{name_col} AS product_name",
    ]

    sql = f"SELECT {', '.join(select_parts)} FROM product_mst"
    if use_col:
        if use_col == "use_yn":
            sql += f" WHERE COALESCE({use_col}, 'Y') = 'Y'"
        else:
            sql += f" WHERE COALESCE({use_col}, 'Y') IN ('Y', '1', 1, 'TRUE', 'true')"
    sql += " ORDER BY id"

    product_df = pd.read_sql(sql, conn)
    product_df["retail_price_krw"] = 0.0
    product_df["supply_price_krw"] = 0.0
    product_df["korea_cost_krw"] = 0.0

    if not table_exists(conn, "import_item"):
        return product_df

    import_cols = get_table_columns(conn, "import_item")
    retail_col = find_existing_column(import_cols, ["retail_price_krw", "retail_price", "retail_krw"])
    supply_col = find_existing_column(import_cols, ["supply_price_krw", "supply_price", "supply_krw"])
    korea_cost_col = find_existing_column(import_cols, ["korea_cost_krw", "korean_cost_krw", "korea_cost"])
    if not retail_col and not supply_col and not korea_cost_col:
        return product_df

    import_ref_candidates = [
        "product_id",
        "cigar_product_id",
        "product_mst_id",
        "product_code",
        "item_code",
        "code",
        "product_name",
        "item_name",
        "name",
    ]
    import_ref_col = find_existing_column(import_cols, import_ref_candidates)
    if not import_ref_col:
        return product_df

    import_select = [f"{import_ref_col} AS import_ref"]
    import_select.append(f"{retail_col} AS retail_price_krw" if retail_col else "0 AS retail_price_krw")
    import_select.append(f"{supply_col} AS supply_price_krw" if supply_col else "0 AS supply_price_krw")
    import_select.append(f"{korea_cost_col} AS korea_cost_krw" if korea_cost_col else "0 AS korea_cost_krw")

    order_candidates = [c for c in ["updated_at", "created_at", "import_date", "id"] if c in import_cols]
    for c in order_candidates:
        import_select.append(c)

    import_df = pd.read_sql(f"SELECT {', '.join(import_select)} FROM import_item", conn)
    if import_df.empty:
        return product_df

    sort_cols = [c for c in ["updated_at", "created_at", "import_date", "id"] if c in import_df.columns]
    if sort_cols:
        import_df = import_df.sort_values(sort_cols, ascending=False)

    if import_ref_col in ["product_id", "cigar_product_id", "product_mst_id"]:
        product_df["join_ref"] = product_df["id"].astype(str)
        import_df["join_ref"] = import_df["import_ref"].astype(str)
    elif import_ref_col in ["product_code", "item_code", "code"] and "product_code" in product_df.columns:
        product_df["join_ref"] = product_df["product_code"].fillna("").astype(str)
        import_df["join_ref"] = import_df["import_ref"].fillna("").astype(str)
    else:
        product_df["join_ref"] = product_df["product_name"].fillna("").astype(str)
        import_df["join_ref"] = import_df["import_ref"].fillna("").astype(str)

    import_df = import_df.drop_duplicates(subset=["join_ref"], keep="first")

    merged = product_df.merge(
        import_df[["join_ref", "retail_price_krw", "supply_price_krw", "korea_cost_krw"]],
        how="left",
        on="join_ref",
        suffixes=("", "_imp"),
    )

    if "retail_price_krw_imp" in merged.columns:
        merged["retail_price_krw"] = merged["retail_price_krw_imp"].combine_first(merged["retail_price_krw"])
    if "supply_price_krw_imp" in merged.columns:
        merged["supply_price_krw"] = merged["supply_price_krw_imp"].combine_first(merged["supply_price_krw"])
    if "korea_cost_krw_imp" in merged.columns:
        merged["korea_cost_krw"] = merged["korea_cost_krw_imp"].combine_first(merged["korea_cost_krw"])

    keep_cols = ["id", "product_code", "product_name", "retail_price_krw", "supply_price_krw", "korea_cost_krw"]
    result = merged[[c for c in keep_cols if c in merged.columns]].copy()
    result["retail_price_krw"] = result["retail_price_krw"].apply(_safe_float)
    result["supply_price_krw"] = result["supply_price_krw"].apply(_safe_float)
    result["korea_cost_krw"] = result["korea_cost_krw"].apply(_safe_float)
    return result.sort_values("id").reset_index(drop=True)


def load_non_cigar_products(conn) -> pd.DataFrame:
    """
    시가 외(non_cigar) 상품 마스터 목록.
    product_name 외에도 마스터 기준가(retail_price / wholesale_price)를
    retail_price_krw / supply_price_krw 로 함께 반환하여
    도매 판매 등록 시 기본값으로 사용할 수 있게 한다.
    """
    empty_cols = ["id", "product_name", "retail_price_krw", "supply_price_krw", "korea_cost_krw"]

    if not table_exists(conn, "non_cigar_product_mst"):
        return pd.DataFrame(columns=empty_cols)

    cols = get_table_columns(conn, "non_cigar_product_mst")
    name_col = find_existing_column(cols, ["product_name", "name"])
    if not name_col:
        return pd.DataFrame(columns=empty_cols)

    retail_col = find_existing_column(cols, ["retail_price", "retail_price_krw"])
    supply_col = find_existing_column(cols, ["wholesale_price", "supply_price_krw", "supply_price"])
    # 시가 외 상품 원가 = non_cigar_product_mst.purchase_price (소매 이익 계산과 동일 기준)
    cost_col = find_existing_column(cols, ["purchase_price", "korea_cost_krw", "unit_cost"])

    select_parts = ["id", f"{name_col} AS product_name"]
    select_parts.append(f"{retail_col} AS retail_price_krw" if retail_col else "0 AS retail_price_krw")
    select_parts.append(f"{supply_col} AS supply_price_krw" if supply_col else "0 AS supply_price_krw")
    select_parts.append(f"{cost_col} AS korea_cost_krw" if cost_col else "0 AS korea_cost_krw")

    sql = f"""
        SELECT {', '.join(select_parts)}
        FROM non_cigar_product_mst
        ORDER BY id
    """
    df = pd.read_sql(sql, conn)
    if df.empty:
        return pd.DataFrame(columns=empty_cols)

    df["retail_price_krw"] = df["retail_price_krw"].apply(_safe_float)
    df["supply_price_krw"] = df["supply_price_krw"].apply(_safe_float)
    df["korea_cost_krw"] = df["korea_cost_krw"].apply(_safe_float)
    return df


def load_wholesale_sales(conn) -> pd.DataFrame:
    ensure_wholesale_sales_columns(conn)
    return pd.read_sql(
        """
        SELECT *
        FROM wholesale_sales
        ORDER BY sale_date DESC, id DESC
        """,
        conn,
    )


def load_wholesale_sales_for_grid(conn) -> pd.DataFrame:
    ensure_wholesale_sales_columns(conn)
    sales = load_wholesale_sales(conn)
    if sales.empty:
        return sales

    df = sales.copy()

    for col in ["partner_id", "cigar_product_id", "non_cigar_product_id"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    partner_base = load_partners(conn).copy()
    partner_keep_cols = [
        c for c in ["id", "partner_name", "business_no", "owner_name", "contact_name", "phone", "email", "address"]
        if c in partner_base.columns
    ]
    partners = partner_base[partner_keep_cols].copy()
    partners["id"] = pd.to_numeric(partners["id"], errors="coerce")
    partners = partners.rename(columns={"id": "partner_id_lookup"})
    df = df.merge(partners, how="left", left_on="partner_id", right_on="partner_id_lookup")

    cigar_products = load_cigar_products_for_wholesale(conn)
    if not cigar_products.empty:
        cigar_products = cigar_products.copy()
        cigar_products["id"] = pd.to_numeric(cigar_products["id"], errors="coerce")
        cigar_products = cigar_products.rename(columns={"id": "cigar_product_id"})
        df = df.merge(
            cigar_products,
            how="left",
            on="cigar_product_id",
            suffixes=("", "_cigar"),
        )

    if table_exists(conn, "non_cigar_product_mst"):
        nc = load_non_cigar_products(conn).copy()
        if not nc.empty:
            nc["id"] = pd.to_numeric(nc["id"], errors="coerce")
            nc = nc.rename(
                columns={
                    "id": "non_cigar_product_id",
                    "product_name": "non_cigar_product_name",
                }
            )
            nc = nc[["non_cigar_product_id", "non_cigar_product_name"]]

            df = df.merge(
                nc,
                how="left",
                on="non_cigar_product_id",
                suffixes=("", "_non_cigar"),
            )

            if "product_name" in df.columns:
                df["product_name"] = df["product_name"].fillna(df["non_cigar_product_name"])
            else:
                df["product_name"] = df["non_cigar_product_name"]

    if "product_code" not in df.columns:
        df["product_code"] = ""

    if "sales_amount" not in df.columns and "supply_amount" in df.columns:
        df["sales_amount"] = df["supply_amount"]

    df["item_type_label"] = df["item_type"].map({
        "cigar": "시가",
        "non_cigar": "시가 외",
    }).fillna(df["item_type"])

    return df


def _validate_statement_rows(selected_df: pd.DataFrame) -> tuple[bool, str]:
    if selected_df.empty:
        return False, "거래명세서로 만들 항목을 하나 이상 선택해 주세요."

    partner_cnt = selected_df["partner_name"].fillna("").astype(str).nunique()
    sale_date_cnt = selected_df["sale_date"].fillna("").astype(str).nunique()

    if partner_cnt != 1:
        return False, "같은 거래처의 항목만 하나의 거래명세서로 출력할 수 있습니다."
    if sale_date_cnt != 1:
        return False, "같은 판매일자의 항목만 하나의 거래명세서로 출력할 수 있습니다."

    return True, ""


def _build_statement_from_rows(conn, selected_df: pd.DataFrame, document_no: str | None = None) -> tuple[dict, pd.DataFrame, dict]:
    sale_date_raw = str(selected_df.iloc[0].get("sale_date", "") or "")
    try:
        sale_dt = pd.to_datetime(sale_date_raw).to_pydatetime()
        sale_date_display = sale_dt.strftime("%Y-%m-%d")
        due_date_display = (sale_dt + pd.Timedelta(days=STATEMENT_PAYMENT_DUE_DAYS)).strftime("%Y-%m-%d")
    except Exception:
        sale_date_display = sale_date_raw
        due_date_display = sale_date_raw

    partner_id = _safe_int(selected_df.iloc[0].get("partner_id", 0), 0)
    partner_info = get_partner_detail_by_id(conn, partner_id)

    row0 = selected_df.iloc[0]

    partner_name = str(
        partner_info.get("partner_name", "") or row0.get("partner_name", "") or ""
    ).strip()

    owner_name = str(
        partner_info.get("owner_name", "") or row0.get("owner_name", "") or ""
    ).strip()

    contact_name = str(
        partner_info.get("contact_name", "") or row0.get("contact_name", "") or ""
    ).strip()

    phone = str(
        partner_info.get("phone", "") or row0.get("phone", "") or ""
    ).strip()

    address = str(
        partner_info.get("address", "") or row0.get("address", "") or ""
    ).strip()

    business_no = str(
        partner_info.get("business_no", "") or row0.get("business_no", "") or ""
    ).strip()

    partner_contact = owner_name or contact_name or partner_name

    statement_df = selected_df[
        [
            "product_name",
            "supply_price",
            "qty",
            "supply_amount",
            "vat_amount",
            "total_amount_vat",
        ]
    ].copy()

    statement_df = statement_df.rename(
        columns={
            "product_name": "품목",
            "supply_price": "단가",
            "qty": "수량",
            "supply_amount": "공급가액",
            "vat_amount": "세액",
            "total_amount_vat": "합계",
        }
    )

    statement_df["단가"] = pd.to_numeric(statement_df["단가"], errors="coerce").fillna(0)
    statement_df["수량"] = pd.to_numeric(statement_df["수량"], errors="coerce").fillna(0).astype(int)
    statement_df["공급가액"] = pd.to_numeric(statement_df["공급가액"], errors="coerce").fillna(0)
    statement_df["세액"] = pd.to_numeric(statement_df["세액"], errors="coerce").fillna(0)
    statement_df["합계"] = pd.to_numeric(statement_df["합계"], errors="coerce").fillna(0)

    totals = {
        "qty": int(statement_df["수량"].sum()),
        "supply_amount": float(statement_df["공급가액"].sum()),
        "vat_amount": float(statement_df["세액"].sum()),
        "total_amount": float(statement_df["합계"].sum()),
    }

    # 등급 할인 합계 계산 (정상가 합계 - 실제 청구 공급가액 합계)
    normal_supply_series = pd.to_numeric(
        selected_df.get("normal_supply_amount", pd.Series(dtype=float)), errors="coerce"
    )
    actual_supply_series = pd.to_numeric(
        selected_df.get("supply_amount", pd.Series(dtype=float)), errors="coerce"
    ).fillna(0)
    # normal_supply_amount가 기록 안 된(과거) 건은 실제 공급가액을 그대로 정상가로 간주(할인 0원 처리)
    normal_supply_series = normal_supply_series.fillna(actual_supply_series)

    normal_supply_sum = float(normal_supply_series.sum())
    discount_amount = max(0.0, round(normal_supply_sum - totals["supply_amount"]))

    applied_grades = sorted(
        {
            str(g).strip()
            for g in selected_df.get("grade_code_applied", pd.Series(dtype=str)).fillna("").tolist()
            if str(g).strip()
        }
    )

    totals["normal_supply_amount"] = normal_supply_sum
    totals["discount_amount"] = discount_amount
    totals["applied_grades"] = applied_grades

    discount_note = ""
    if discount_amount > 0:
        grade_label = ", ".join(applied_grades) if applied_grades else "등급"
        discount_note = (
            f"[{grade_label} 등급할인 -₩{discount_amount:,.0f} 반영 "
            f"(정상가 합계 ₩{normal_supply_sum:,.0f} → 공급가액 ₩{totals['supply_amount']:,.0f})]"
        )

    base_notes = " / ".join(
        sorted(
            {
                str(x).strip()
                for x in selected_df.get("notes", pd.Series(dtype=str)).fillna("").tolist()
                if str(x).strip()
            }
        )
    )

    header = {
        "partner_id": partner_id,
        "partner_name": partner_name,
        "partner_contact": partner_contact,
        "partner_phone": phone,
        "partner_address": address,
        "business_no": business_no,
        "sale_date": sale_date_display,
        "due_date": due_date_display,
        "document_no": document_no or "",
        "notes": (f"{discount_note} / {base_notes}" if base_notes else discount_note) if discount_note else base_notes,
    }

    return header, statement_df, totals


def _render_statement_preview(header: dict, statement_df: pd.DataFrame, totals: dict):
    st.divider()
    st.markdown("#### 거래명세서 미리보기")

    c1, c2, c3 = st.columns(3)
    with c1:
        st.write(f"**거래처**: {header['partner_name']}")
    with c2:
        st.write(f"**판매일자**: {header['sale_date']}")
    with c3:
        preview_doc_no = header.get("document_no") or "(다운로드 시 발급)"
        st.write(f"**문서번호**: {preview_doc_no}")

    if header.get("notes"):
        st.write(f"**비고**: {header['notes']}")

    st.dataframe(
        statement_df.style.format(
            {
                "단가": "₩{:,.0f}",
                "공급가액": "₩{:,.0f}",
                "세액": "₩{:,.0f}",
                "합계": "₩{:,.0f}",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    m1, m2, m3 = st.columns(3)
    with m1:
        st.metric("공급가액 합계", f"₩{totals['supply_amount']:,.0f}")
    with m2:
        st.metric("부가세 합계", f"₩{totals['vat_amount']:,.0f}")
    with m3:
        st.metric("총 합계", f"₩{totals['total_amount']:,.0f}")

    if totals.get("discount_amount", 0) > 0:
        d1, d2, d3 = st.columns(3)
        with d1:
            st.metric("정상가 합계", f"₩{totals['normal_supply_amount']:,.0f}")
        with d2:
            grade_label = ", ".join(totals.get("applied_grades", [])) or "등급"
            st.metric(f"{grade_label} 등급할인", f"-₩{totals['discount_amount']:,.0f}")
        with d3:
            discount_pct = (
                totals["discount_amount"] / totals["normal_supply_amount"] * 100
                if totals.get("normal_supply_amount")
                else 0
            )
            st.metric("평균 할인율", f"{discount_pct:.1f}%")


def _make_statement_excel_bytes(header: dict, statement_df: pd.DataFrame, totals: dict) -> bytes:
    if not STATEMENT_TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"거래명세서 템플릿 파일이 없습니다: {STATEMENT_TEMPLATE_PATH}"
        )

    wb = load_workbook(STATEMENT_TEMPLATE_PATH)
    if "거래명세서" not in wb.sheetnames:
        raise ValueError("거래명세서 템플릿에 '거래명세서' 시트가 없습니다.")

    ws = wb["거래명세서"]

    # 헤더
    set_merged_cell_value(ws, "I2", header.get("partner_name", ""))

    address_val = header.get("partner_address", "")
    set_merged_cell_value(ws, "J3", address_val)
    _apply_alignment_safe(ws, "J3", Alignment(horizontal="right", vertical="center", wrap_text=True))

    owner = str(header.get("partner_contact", "") or "").strip()
    partner_name = str(header.get("partner_name", "") or "").strip()
    display_owner = owner if owner else partner_name

    set_merged_cell_value(ws, "J4", f"{display_owner} 대표님")
    _apply_alignment_safe(ws, "J4", Alignment(horizontal="right", vertical="center"))

    set_merged_cell_value(ws, "J5", header.get("partner_phone", ""))
    _apply_alignment_safe(ws, "J5", Alignment(horizontal="right", vertical="center"))

    ws["C12"] = header.get("document_no", "")
    ws["C13"] = header.get("sale_date", "")
    ws["C14"] = header.get("due_date", "")

    ws["H12"] = STATEMENT_BANK_NAME
    ws["H13"] = STATEMENT_BANK_ACCOUNT
    ws["H14"] = STATEMENT_ACCOUNT_HOLDER

    item_count = len(statement_df)

    # ── 템플릿 실제 구조 반영: 품목 행 18~49 = 32행 ──
    TEMPLATE_ITEM_ROWS = 32   # 수정: 20 → 32
    start_row = 18
    template_end_row = start_row + TEMPLATE_ITEM_ROWS  # 50 (exclusive)

    # 품목이 32개 초과일 때만 행 삽입
    extra_rows = max(0, item_count - TEMPLATE_ITEM_ROWS)
    if extra_rows > 0:
        ws.insert_rows(template_end_row, amount=extra_rows)

    # 품목 영역 전체 초기화
    total_item_rows = max(item_count, TEMPLATE_ITEM_ROWS)
    for row in range(start_row, start_row + total_item_rows):
        ws[f"B{row}"] = None
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = 0
        ws[f"H{row}"] = 0
        ws[f"I{row}"] = 0
        ws[f"J{row}"] = 0

    # 품목 입력
    for idx, item in statement_df.reset_index(drop=True).iterrows():
        row = start_row + idx
        ws[f"B{row}"] = item.get("품목", "")
        ws[f"F{row}"] = _safe_float(item.get("단가", 0), 0)
        ws[f"G{row}"] = _safe_int(item.get("수량", 0), 0)
        ws[f"H{row}"] = _safe_float(item.get("공급가액", 0), 0)
        ws[f"I{row}"] = _safe_float(item.get("세액", 0), 0)
        ws[f"J{row}"] = _safe_float(item.get("합계", 0), 0)

    # 남은 행 0 처리
    for row in range(start_row + item_count, start_row + TEMPLATE_ITEM_ROWS):
        ws[f"F{row}"] = 0
        ws[f"G{row}"] = 0
        ws[f"H{row}"] = 0
        ws[f"I{row}"] = 0
        ws[f"J{row}"] = 0

    # 숫자 포맷 유지
    for row in range(start_row, start_row + total_item_rows):
        for col in "FGHIJ":
            ws[f"{col}{row}"].number_format = '#,##0'

    # ── 합계 셀: 템플릿 기준 51~53행 (extra_rows만큼 밀림 반영) ──
    totals_base = 51 + extra_rows
    ws[f"I{totals_base}"]     = totals.get("supply_amount", 0)
    ws[f"I{totals_base + 1}"] = totals.get("vat_amount", 0)
    ws[f"I{totals_base + 2}"] = totals.get("total_amount", 0)
    for t_row in range(totals_base, totals_base + 3):
        ws[f"I{t_row}"].number_format = '#,##0'

    # ── 비고: extra_rows 반영하여 정확한 행에 기록 ──
    notes_value = str(header.get("notes", "") or "")
    notes_content_row = 56 + extra_rows   # 템플릿 기준 56행
    ws[f"B{notes_content_row}"] = notes_value

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _make_bulk_statement_zip_bytes(conn, selected_df: pd.DataFrame) -> tuple[bytes, int, list[str]]:
    zip_buf = BytesIO()
    count = 0
    errors: list[str] = []

    groups = selected_df.groupby(["partner_name", "sale_date"], sort=False)

    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for (partner_name, sale_date), group_df in groups:
            label = f"{partner_name} / {sale_date}"
            try:
                document_no = issue_statement_document_no(conn, str(sale_date))
                header, stmt_df, totals = _build_statement_from_rows(conn, group_df.copy(), document_no)
                header["document_no"] = document_no

                excel_bytes = _make_statement_excel_bytes(header, stmt_df, totals)

                safe_partner = (
                    "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(partner_name))
                    .strip("_") or "거래처"
                )
                file_name = f"거래명세서_{safe_partner}_{document_no}.xlsx"
                zf.writestr(file_name, excel_bytes)

                try:
                    save_statement_history(conn, header, totals, file_name)
                except sqlite3.IntegrityError:
                    pass

                count += 1
            except Exception as exc:
                errors.append(f"{label}: {exc}")

    return zip_buf.getvalue(), count, errors


# -----------------------------
# Partner CRUD
# -----------------------------
def insert_partner(
    conn,
    partner_name: str,
    partner_type: str,
    business_no: str,
    owner_name: str,
    contact_name: str,
    phone: str,
    email: str,
    address: str,
    join_date: str,
    current_grade_code: str,
    notes: str,
    program_exempt_yn: int = 0,
):
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO partner_mst (
            partner_name,
            partner_type,
            business_no,
            owner_name,
            contact_name,
            phone,
            email,
            address,
            join_date,
            status,
            current_grade_code,
            grade_acquired_date,
            notes,
            program_exempt_yn
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            partner_name,
            partner_type,
            business_no,
            owner_name,
            contact_name,
            phone,
            email,
            address,
            join_date,
            "active",
            current_grade_code if current_grade_code else None,
            join_date if current_grade_code else None,
            notes,
            int(program_exempt_yn or 0),
        ),
    )
    conn.commit()


def update_partner(
    conn,
    partner_id: int,
    partner_name: str,
    partner_type: str,
    business_no: str,
    owner_name: str,
    contact_name: str,
    phone: str,
    email: str,
    address: str,
    join_date: str,
    current_grade_code: str,
    notes: str,
    status: str,
    program_exempt_yn: int = 0,
):
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE partner_mst
        SET
            partner_name = ?,
            partner_type = ?,
            business_no = ?,
            owner_name = ?,
            contact_name = ?,
            phone = ?,
            email = ?,
            address = ?,
            join_date = ?,
            current_grade_code = ?,
            notes = ?,
            status = ?,
            program_exempt_yn = ?,
            grade_acquired_date = CASE
                WHEN COALESCE(?, '') <> '' AND COALESCE(current_grade_code, '') <> COALESCE(?, '') THEN ?
                ELSE grade_acquired_date
            END
        WHERE id = ?
        """,
        (
            partner_name,
            partner_type,
            business_no,
            owner_name,
            contact_name,
            phone,
            email,
            address,
            join_date,
            current_grade_code if current_grade_code else None,
            notes,
            status,
            int(program_exempt_yn or 0),
            current_grade_code if current_grade_code else None,
            current_grade_code if current_grade_code else None,
            join_date,
            partner_id,
        ),
    )
    conn.commit()


def delete_partner(conn, partner_id: int):
    """
    거래처를 영구 삭제한다.

    도매판매(wholesale_sales) / 명세서(statement_history) / 출고이력(stock_out)에
    해당 거래처를 참조하는 데이터가 하나라도 있으면 삭제를 막는다 — 삭제 시
    기존 판매/재무/재고 이력의 거래처 정보가 깨지기 때문. 이 경우 '상태'를
    '비활성(inactive)'으로 바꾸는 것으로 대체해야 한다.

    등급이력(partner_grade_history)은 거래처에만 종속된 메타데이터라 함께 삭제한다.
    """
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM wholesale_sales WHERE partner_id = ?", (partner_id,))
    sales_cnt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM statement_history WHERE partner_id = ?", (partner_id,))
    stmt_cnt = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM stock_out WHERE partner_id = ?", (partner_id,))
    stockout_cnt = cur.fetchone()[0]

    if sales_cnt > 0 or stmt_cnt > 0 or stockout_cnt > 0:
        raise ValueError(
            f"이 거래처는 연관된 데이터가 있어 삭제할 수 없습니다 "
            f"(도매판매 {sales_cnt}건 / 명세서 {stmt_cnt}건 / 출고이력 {stockout_cnt}건). "
            "삭제 대신 '상태'를 '비활성(inactive)'으로 변경해 주세요."
        )

    cur.execute("DELETE FROM partner_grade_history WHERE partner_id = ?", (partner_id,))
    cur.execute("DELETE FROM partner_mst WHERE id = ?", (partner_id,))
    conn.commit()


# -----------------------------
# Wholesale CRUD
# -----------------------------
def insert_wholesale_sale(
    conn,
    sale_date: str,
    partner_id: int,
    item_type: str,
    cigar_product_id,
    non_cigar_product_id,
    qty: int,
    unit_price: float,
    supply_price: float,
    unit_cost: float,
    supply_amount: float,
    vat_amount: float,
    total_amount_vat: float,
    profit_amount: float,
    notes: str,
    grade_code_applied: str = None,
    discount_rate_applied: float = 0.0,
    normal_supply_amount: float = None,
):
    ensure_wholesale_sales_columns(conn)
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO wholesale_sales (
            sale_date,
            partner_id,
            item_type,
            cigar_product_id,
            non_cigar_product_id,
            qty,
            unit_price,
            supply_price,
            unit_cost,
            sales_amount,
            supply_amount,
            vat_amount,
            total_amount_vat,
            profit_amount,
            notes,
            grade_code_applied,
            discount_rate_applied,
            normal_supply_amount,
            updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        """,
        (
            sale_date,
            partner_id,
            item_type,
            cigar_product_id,
            non_cigar_product_id,
            qty,
            unit_price,
            supply_price,
            unit_cost,
            supply_amount,
            supply_amount,
            vat_amount,
            total_amount_vat,
            profit_amount,
            notes,
            grade_code_applied,
            discount_rate_applied,
            normal_supply_amount,
        ),
    )
    conn.commit()


def update_wholesale_sale(
    conn,
    sale_id: int,
    sale_date: str,
    partner_id: int,
    item_type: str,
    cigar_product_id,
    non_cigar_product_id,
    qty: int,
    unit_price: float,
    supply_price: float,
    unit_cost: float,
    supply_amount: float,
    vat_amount: float,
    total_amount_vat: float,
    profit_amount: float,
    notes: str,
):
    ensure_wholesale_sales_columns(conn)
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE wholesale_sales
        SET
            sale_date = ?,
            partner_id = ?,
            item_type = ?,
            cigar_product_id = ?,
            non_cigar_product_id = ?,
            qty = ?,
            unit_price = ?,
            supply_price = ?,
            unit_cost = ?,
            sales_amount = ?,
            supply_amount = ?,
            vat_amount = ?,
            total_amount_vat = ?,
            profit_amount = ?,
            notes = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            sale_date,
            partner_id,
            item_type,
            cigar_product_id,
            non_cigar_product_id,
            qty,
            unit_price,
            supply_price,
            unit_cost,
            supply_amount,
            supply_amount,
            vat_amount,
            total_amount_vat,
            profit_amount,
            notes,
            sale_id,
        ),
    )
    conn.commit()


def delete_wholesale_sale(conn, sale_id: int):
    cur = conn.cursor()
    cur.execute("DELETE FROM wholesale_sales WHERE id = ?", (sale_id,))
    conn.commit()


# -----------------------------
# UI helpers
# -----------------------------
def _render_amount_summary(qty: int, unit_price: float, supply_price: float, unit_cost: float):
    supply_amount = float(qty) * float(supply_price)
    vat_amount = round(supply_amount * 0.1, 2)
    total_amount_vat = supply_amount + vat_amount
    profit_amount = float(qty) * (float(supply_price) - float(unit_cost))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("공급가액", f"₩{supply_amount:,.0f}")
    c2.metric("부가세", f"₩{vat_amount:,.0f}")
    c3.metric("부가세 포함 금액", f"₩{total_amount_vat:,.0f}")
    c4.metric("예상마진", f"₩{profit_amount:,.0f}")
    return supply_amount, vat_amount, total_amount_vat, profit_amount


def _bind_price_cost_by_selection(state_prefix: str, selected_product_key: str, auto_unit_price: float, auto_supply_price: float, auto_unit_cost: float):
    prev_key = f"{state_prefix}_selected_product_prev"
    price_key = f"{state_prefix}_unit_price"
    supply_key = f"{state_prefix}_supply_price"
    cost_key = f"{state_prefix}_unit_cost"

    prev_selected = st.session_state.get(prev_key)
    if prev_selected != selected_product_key:
        st.session_state[price_key] = int(auto_unit_price)
        st.session_state[supply_key] = int(auto_supply_price)
        st.session_state[cost_key] = int(auto_unit_cost)
        st.session_state[prev_key] = selected_product_key
    else:
        st.session_state.setdefault(price_key, int(auto_unit_price))
        st.session_state.setdefault(supply_key, int(auto_supply_price))
        st.session_state.setdefault(cost_key, int(auto_unit_cost))


# -----------------------------
# Render sections
# -----------------------------
def render_partner_registration(conn):
    st.markdown("### 거래처 관리")

    ensure_partner_grade_tables(conn)

    partners = load_partners(conn)
    grade_codes = load_grade_codes(conn)
    grade_active_map = load_partner_active_grade_map(conn)

    mode = st.radio("작업", ["신규 등록", "기존 업체 수정"], horizontal=True)

    purchase_summary = load_partner_purchase_summary_from_grade(conn)

    if not purchase_summary.empty:
        partners = partners.merge(
            purchase_summary,
            how="left",
            left_on="id",
            right_on="partner_id",
        )
    else:
        partners["purchase_item_names"] = ""
        partners["purchase_supply_amount_sum"] = 0
        partners["grade_start_date"] = None
        partners["next_grade_code"] = None
        partners["next_grade_remaining"] = None

    last_purchase = load_partner_last_purchase(conn)
    if not last_purchase.empty:
        partners = partners.merge(
            last_purchase,
            how="left",
            left_on="id",
            right_on="partner_id",
            suffixes=("", "_lp"),
        )
    else:
        partners["last_purchase_date"] = None
        partners["purchase_count"] = 0

    if not grade_active_map.empty:
        partners = partners.merge(
            grade_active_map.rename(columns={
                "grade_code": "active_grade_code",
                "start_date": "active_start_date",
                "end_date": "active_end_date",
            }),
            how="left",
            left_on="id",
            right_on="partner_id",
            suffixes=("", "_grade"),
        )
    else:
        partners["active_grade_code"] = None
        partners["active_start_date"] = None
        partners["active_end_date"] = None

    selected_partner = None
    selected_row = None
    if mode == "기존 업체 수정":
        if partners.empty:
            st.info("수정할 거래처가 없습니다.")
            return

        option_df = partners[["id", "partner_name"]].copy()
        option_df["label"] = option_df["partner_name"].fillna("") + " (ID:" + option_df["id"].astype(str) + ")"
        selected_label = st.selectbox("수정할 업체 선택", option_df["label"].tolist())
        selected_partner = option_df.loc[option_df["label"] == selected_label].iloc[0]
        selected_row = partners.loc[partners["id"] == selected_partner["id"]].iloc[0]

        active_grade_display = selected_row.get("active_grade_code") or selected_row.get("current_grade_code") or "-"
        active_end_display = selected_row.get("active_end_date") or "-"
        st.caption(
            f"현재 등급: **{active_grade_display}** / 만료일: {active_end_display} "
            "(등급 변경은 '판매관리 > 거래처등급관리' 화면에서 처리해 주세요)"
        )

        with st.expander("⚠️ 거래처 삭제"):
            del_partner_id = int(selected_row["id"])
            del_cur = conn.cursor()
            del_cur.execute("SELECT COUNT(*) FROM wholesale_sales WHERE partner_id = ?", (del_partner_id,))
            del_sales_cnt = del_cur.fetchone()[0]
            del_cur.execute("SELECT COUNT(*) FROM statement_history WHERE partner_id = ?", (del_partner_id,))
            del_stmt_cnt = del_cur.fetchone()[0]
            del_cur.execute("SELECT COUNT(*) FROM stock_out WHERE partner_id = ?", (del_partner_id,))
            del_stockout_cnt = del_cur.fetchone()[0]
            del_cur.execute("SELECT COUNT(*) FROM partner_grade_history WHERE partner_id = ?", (del_partner_id,))
            del_grade_cnt = del_cur.fetchone()[0]

            st.caption(
                f"연관 데이터 — 도매판매 {del_sales_cnt}건 / 명세서 {del_stmt_cnt}건 / "
                f"출고이력 {del_stockout_cnt}건 / 등급이력 {del_grade_cnt}건"
            )

            can_delete = del_sales_cnt == 0 and del_stmt_cnt == 0 and del_stockout_cnt == 0

            if not can_delete:
                st.warning(
                    "도매판매/명세서/출고 이력이 있는 거래처는 데이터 무결성 보호를 위해 삭제할 수 없습니다. "
                    "위 '거래처 수정' 폼에서 상태를 '비활성(inactive)'으로 변경해 주세요."
                )
            else:
                confirm_label = f"'{selected_row['partner_name']}' 거래처를 영구 삭제하는 것에 동의합니다"
                if del_grade_cnt:
                    confirm_label += f" (등급이력 {del_grade_cnt}건도 함께 삭제됩니다)"
                confirm_delete = st.checkbox(confirm_label, key=f"confirm_delete_partner_{del_partner_id}")
                if st.button(
                    "거래처 영구 삭제",
                    type="primary",
                    disabled=not confirm_delete,
                    key=f"delete_partner_btn_{del_partner_id}",
                ):
                    try:
                        deleted_name = str(selected_row["partner_name"])
                        delete_partner(conn, del_partner_id)
                        st.success(f"'{deleted_name}' 거래처가 삭제되었습니다.")
                        st.rerun()
                    except ValueError as e:
                        st.error(str(e))

    with st.form("partner_registration_form", clear_on_submit=(mode == "신규 등록")):
        c1, c2 = st.columns(2)
        partner_type_list = ["wholesale", "retail_partner", "bar", "cafe", "shop", "etc"]
        status_list = ["active", "inactive"]

        with c1:
            partner_name = st.text_input("거래처명 *", value="" if selected_row is None else str(selected_row["partner_name"] or ""))
            default_partner_type = "wholesale" if selected_row is None else str(selected_row["partner_type"] or "wholesale")
            partner_type = st.selectbox(
                "거래처 유형",
                partner_type_list,
                index=partner_type_list.index(default_partner_type) if default_partner_type in partner_type_list else 0,
            )
            business_no = st.text_input("사업자번호", value="" if selected_row is None else str(selected_row.get("business_no", "") or ""))
            owner_name = st.text_input("대표자명", value="" if selected_row is None else str(selected_row["owner_name"] or ""))
            contact_name = st.text_input("담당자명", value="" if selected_row is None else str(selected_row["contact_name"] or ""))

        with c2:
            phone = st.text_input("연락처", value="" if selected_row is None else str(selected_row["phone"] or ""))
            email = st.text_input("이메일", value="" if selected_row is None else str(selected_row["email"] or ""))
            address = st.text_input("주소", value="" if selected_row is None else str(selected_row.get("address", "") or ""))

            default_join_date = pd.to_datetime("today").date()
            if selected_row is not None and pd.notna(selected_row["join_date"]):
                default_join_date = pd.to_datetime(selected_row["join_date"]).date()
            join_date = str(st.date_input("가입일", value=default_join_date))

            if grade_codes:
                grade_list = [""] + grade_codes
                default_grade = "" if selected_row is None else str(selected_row["current_grade_code"] or "")
                current_grade_code = st.selectbox(
                    "현재 등급",
                    grade_list,
                    index=grade_list.index(default_grade) if default_grade in grade_list else 0,
                )
            else:
                current_grade_code = st.text_input("현재 등급", value="" if selected_row is None else str(selected_row["current_grade_code"] or ""))

            default_status = "active" if selected_row is None else str(selected_row["status"] or "active")
            status = st.selectbox(
                "상태",
                status_list,
                index=status_list.index(default_status) if default_status in status_list else 0,
            )
            notes = st.text_area("비고", value="" if selected_row is None else str(selected_row.get("notes", "") or ""))
            program_exempt_yn = st.checkbox(
                "파트너 프로그램 제외 (등급별 자동 할인 적용 안 함)",
                value=False if selected_row is None else bool(int(selected_row.get("program_exempt_yn") or 0)),
                help="이미 별도 협의로 할인 중인 업체 등, 매출 누적액과 무관하게 등급 할인을 적용하면 안 되는 거래처에 체크하세요.",
            )

        submitted = st.form_submit_button(
            "거래처 등록" if mode == "신규 등록" else "거래처 수정 저장",
            use_container_width=True,
        )

        if submitted:
            if not partner_name.strip():
                st.error("거래처명은 필수입니다.")
                return

            if mode == "신규 등록":
                dup = pd.read_sql(
                    "SELECT COUNT(*) AS cnt FROM partner_mst WHERE partner_name = ?",
                    conn,
                    params=(partner_name.strip(),),
                )["cnt"].iloc[0]
                if dup > 0:
                    st.error("이미 같은 거래처명이 등록되어 있습니다.")
                    return

                insert_partner(
                    conn=conn,
                    partner_name=partner_name.strip(),
                    partner_type=partner_type,
                    business_no=business_no.strip(),
                    owner_name=owner_name.strip(),
                    contact_name=contact_name.strip(),
                    phone=phone.strip(),
                    email=email.strip(),
                    address=address.strip(),
                    join_date=join_date,
                    current_grade_code=current_grade_code.strip() if current_grade_code else "",
                    notes=notes.strip(),
                    program_exempt_yn=int(program_exempt_yn),
                )
                st.success(f"거래처가 등록되었습니다: {partner_name}")
            else:
                dup = pd.read_sql(
                    "SELECT COUNT(*) AS cnt FROM partner_mst WHERE partner_name = ? AND id <> ?",
                    conn,
                    params=(partner_name.strip(), int(selected_partner["id"])),
                )["cnt"].iloc[0]
                if dup > 0:
                    st.error("이미 같은 거래처명이 등록되어 있습니다.")
                    return

                update_partner(
                    conn=conn,
                    partner_id=int(selected_partner["id"]),
                    partner_name=partner_name.strip(),
                    partner_type=partner_type,
                    business_no=business_no.strip(),
                    owner_name=owner_name.strip(),
                    contact_name=contact_name.strip(),
                    phone=phone.strip(),
                    email=email.strip(),
                    address=address.strip(),
                    join_date=join_date,
                    current_grade_code=current_grade_code.strip() if current_grade_code else "",
                    notes=notes.strip(),
                    status=status,
                    program_exempt_yn=int(program_exempt_yn),
                )
                st.success(f"거래처 정보가 수정되었습니다: {partner_name}")

            st.rerun()

    if not partners.empty:
        st.divider()
        st.markdown("#### 등록된 거래처")
        sort_key = st.selectbox(
            "정렬", list(PARTNER_SORT_OPTIONS), index=0, key="partner_list_sort",
            help="기본값은 최근구매순입니다. 구매 이력이 없는 거래처는 맨 뒤에 표시됩니다.",
        )
        view_df = sort_partners(partners, sort_key).rename(columns={
            "last_purchase_date": "최근구매일",
            "purchase_count": "구매횟수",
            "partner_name": "거래처명",
            "partner_type": "유형",
            "phone": "연락처",
            "current_grade_code": "등급(수동값)",
            "active_grade_code": "등급(자동계산)",
            "active_end_date": "등급만료일",
            "status": "상태",
            "join_date": "가입일",
            "grade_acquired_date": "등급업일(수동값·미사용)",
            "grade_start_date": "등급업(현재등급) 시작일",
            "purchase_item_names": "등급업 이후 구매물품",
            "purchase_supply_amount_sum": "등급업 이후 공급가합계",
            "next_grade_code": "다음 등급",
            "next_grade_remaining": "다음 등급까지 남은 금액",
            "notes": "비고",
            "business_no": "사업자번호",
            "owner_name": "대표자명",
            "contact_name": "담당자명",
            "email": "이메일",
            "address": "주소",
            "program_exempt_yn": "프로그램제외",
        })

        if "등급업 이후 공급가합계" in view_df.columns:
            view_df["등급업 이후 공급가합계"] = (
                pd.to_numeric(view_df["등급업 이후 공급가합계"], errors="coerce")
                .fillna(0)
                .map(lambda x: f"{x:,.0f}")
            )

        if "다음 등급까지 남은 금액" in view_df.columns:
            view_df["다음 등급까지 남은 금액"] = (
                pd.to_numeric(view_df["다음 등급까지 남은 금액"], errors="coerce")
                .map(lambda x: "-" if pd.isna(x) else f"{x:,.0f}")
            )

        if "다음 등급" in view_df.columns:
            view_df["다음 등급"] = view_df["다음 등급"].fillna("최고 등급")

        if "프로그램제외" in view_df.columns:
            view_df["프로그램제외"] = (
                pd.to_numeric(view_df["프로그램제외"], errors="coerce")
                .fillna(0)
                .map(lambda x: "Y" if int(x) == 1 else "")
            )

        if "최근구매일" in view_df.columns:
            view_df["최근구매일"] = view_df["최근구매일"].fillna("구매이력없음")
        if "구매횟수" in view_df.columns:
            view_df["구매횟수"] = (
                pd.to_numeric(view_df["구매횟수"], errors="coerce").fillna(0).map(lambda x: f"{int(x):,}회")
            )

        preferred_order = [
            "거래처명",
            "최근구매일",
            "구매횟수",
            "유형",
            "연락처",
            "등급(자동계산)",
            "등급만료일",
            "등급(수동값)",
            "프로그램제외",
            "상태",
            "가입일",
            "등급업(현재등급) 시작일",
            "등급업 이후 공급가합계",
            "다음 등급",
            "다음 등급까지 남은 금액",
            "등급업 이후 구매물품",
            "등급업일(수동값·미사용)",
            "사업자번호",
            "비고",
            "이메일",
            "대표자명",
            "담당자명",
            "주소",
        ]

        existing_order = [c for c in preferred_order if c in view_df.columns]
        remaining_cols = [c for c in view_df.columns if c not in existing_order]
        view_df = view_df[existing_order + remaining_cols]

        st.dataframe(view_df, use_container_width=True, hide_index=True)
    else:
        st.info("등록된 거래처가 없습니다.")

def render_wholesale_management(conn):
    st.markdown("### 도매 판매 관리")
    st.caption("상단에서 등록하고, 하단 표에서 행을 선택하여 수정/삭제하거나 거래명세서를 출력할 수 있습니다.")

    partners = load_partners(conn)
    cigar_products = load_cigar_products_for_wholesale(conn)
    non_cigar_products = load_non_cigar_products(conn)

    if partners.empty:
        st.warning("먼저 거래처를 등록해 주세요.")
        return

    with st.expander("신규 도매 판매 등록", expanded=True):
        item_type_label = st.radio("상품 구분", ["시가", "시가 외"], horizontal=True, key="wh_item_type_label")
        item_type = "cigar" if item_type_label == "시가" else "non_cigar"

        c1, c2 = st.columns(2)
        with c1:
            partner_name = st.selectbox("거래처", partners["partner_name"].tolist(), key="wh_partner_name")
            sale_date = str(st.date_input("판매일자", key="wh_sale_date"))
            qty = st.number_input("수량", min_value=1, value=1, step=1, format="%d", key="wh_qty")

        # 선택된 거래처의 실시간 누적 구매액 기준 구간별(문턱값) 할인 계산 준비
        selected_partner_row = partners.loc[partners["partner_name"] == partner_name].iloc[0]
        partner_id_for_default = _safe_int(selected_partner_row["id"])

        cigar_product_id = None
        non_cigar_product_id = None
        product_code = ""
        auto_unit_price = 0.0
        auto_supply_price = 0.0
        auto_unit_cost = 0.0
        pricing_result = None
        base_supply_price_ref = 0.0

        if item_type == "cigar":
            if cigar_products.empty:
                st.warning("시가 상품 기준정보가 없습니다.")
                return

            display_options = cigar_products.copy()
            display_options["display_name"] = display_options.apply(
                lambda r: f"{str(r['product_code']).strip()} | {str(r['product_name']).strip()}"
                if str(r["product_code"]).strip() else str(r["product_name"]).strip(),
                axis=1,
            )

            selected_display = st.selectbox("시가 상품", display_options["display_name"].tolist(), key="wh_cigar_product")
            selected_product = display_options.loc[display_options["display_name"] == selected_display].iloc[0]
            cigar_product_id = _safe_int(selected_product["id"])
            product_code = str(selected_product["product_code"] or "")

            base_unit_price = _safe_float(selected_product.get("retail_price_krw", 0))
            base_supply_price = _safe_float(selected_product.get("supply_price_krw", 0))
            base_supply_price_ref = base_supply_price
            auto_unit_cost = _safe_float(selected_product.get("korea_cost_krw", 0))  # 원가는 할인 미적용

            pricing_result = compute_tiered_order_pricing(
                conn,
                partner_id=partner_id_for_default,
                qty=qty,
                base_unit_price=base_unit_price,
                base_supply_price=base_supply_price,
                order_date=sale_date,
            )
            auto_unit_price = pricing_result["blended_unit_price"]
            auto_supply_price = pricing_result["blended_supply_price"]

            _bind_price_cost_by_selection(
                state_prefix="wh",
                selected_product_key=f"cigar::{cigar_product_id}::partner{partner_id_for_default}::qty{int(qty)}::date{sale_date}",
                auto_unit_price=auto_unit_price,
                auto_supply_price=auto_supply_price,
                auto_unit_cost=auto_unit_cost,
            )
        else:
            if non_cigar_products.empty:
                st.warning("시가 외 상품 기준정보가 없습니다.")
                return

            selected_name = st.selectbox("시가 외 상품", non_cigar_products["product_name"].tolist(), key="wh_non_cigar_product")
            selected_product = non_cigar_products.loc[non_cigar_products["product_name"] == selected_name].iloc[0]
            non_cigar_product_id = _safe_int(selected_product["id"])

            base_unit_price = _safe_float(selected_product.get("retail_price_krw", 0))
            base_supply_price = _safe_float(selected_product.get("supply_price_krw", 0))
            base_supply_price_ref = base_supply_price
            # 원가 기본값 = 시가 외 상품 매입가(purchase_price). 0으로 두면 이익이 공급가액 전체로
            # 과대계상되어 소매 쪽(매입가 기준 원가)과 기준이 어긋나므로 매입가를 기본값으로 사용.
            auto_unit_cost = _safe_float(selected_product.get("korea_cost_krw", 0))

            pricing_result = compute_tiered_order_pricing(
                conn,
                partner_id=partner_id_for_default,
                qty=qty,
                base_unit_price=base_unit_price,
                base_supply_price=base_supply_price,
                order_date=sale_date,
            )
            auto_unit_price = pricing_result["blended_unit_price"]
            auto_supply_price = pricing_result["blended_supply_price"]

            _bind_price_cost_by_selection(
                state_prefix="wh",
                selected_product_key=f"non_cigar::{non_cigar_product_id}::partner{partner_id_for_default}::qty{int(qty)}::date{sale_date}",
                auto_unit_price=auto_unit_price,
                auto_supply_price=auto_supply_price,
                auto_unit_cost=auto_unit_cost,
            )

        normal_total_for_display = base_supply_price_ref * float(qty) if base_supply_price_ref > 0 else None

        if pricing_result and normal_total_for_display is not None:
            discount_ratio_display = pricing_result.get("discount_ratio", 0.0)
            if discount_ratio_display > 0:
                discount_amt_display = normal_total_for_display - pricing_result["blended_supply_price"] * qty
                st.info(
                    f"⚠️ **등급 할인이 자동 반영된 값입니다** — 아래 '공급가'는 할인 적용 후 금액입니다.\n\n"
                    f"정상 공급가 합계: ₩{normal_total_for_display:,.0f} → 자동계산 공급가 합계: "
                    f"₩{pricing_result['blended_supply_price'] * qty:,.0f} "
                    f"(할인 ₩{discount_amt_display:,.0f}, 평균 {discount_ratio_display*100:.1f}%)"
                )
            else:
                st.caption(f"정상 공급가 합계: ₩{normal_total_for_display:,.0f} (현재 등급 기준 할인 없음)")

        with c2:
            unit_price = st.number_input(
                "단가 (₩)", min_value=0, step=100, format="%d", key="wh_unit_price",
                help="상품 정상 단가에 거래처 등급 할인이 자동 반영된 값입니다. 필요시 직접 수정 가능합니다.",
            )
            supply_price = st.number_input(
                "공급가 (₩) — 등급할인 자동반영됨", min_value=0, step=100, format="%d", key="wh_supply_price",
                help="상품 정상 공급가에 거래처 등급 할인이 자동 반영된 값입니다. 필요시 직접 수정 가능합니다.",
            )
            unit_cost = st.number_input("원가 (₩)", min_value=0, step=100, format="%d", key="wh_unit_cost")
            notes = st.text_area("비고", key="wh_notes")

        st.caption(
            f"단가: ₩{unit_price:,.0f} / "
            f"공급가: ₩{supply_price:,.0f} / "
            f"원가: ₩{unit_cost:,.0f}"
        )

        if product_code:
            st.caption(f"상품코드: {product_code}")

        if pricing_result:
            segs = pricing_result.get("segments") or []
            multi_tier = len([s for s in segs if s["amount"] > 0]) > 1 or (segs and segs[0]["discount_rate"] > 0)
            if multi_tier:
                cum_before = pricing_result["cumulative_before"]
                cum_after = pricing_result["cumulative_after"]
                st.caption(f"이 거래처의 누적 공급가액(할인 판단 기준): ₩{cum_before:,.0f} → 이번 주문 후 ₩{cum_after:,.0f}")
                for seg in segs:
                    if seg["amount"] <= 0:
                        continue
                    st.caption(
                        f"　· {seg['grade_code']} 구간 ₩{seg['amount']:,.0f} × "
                        f"{seg['discount_rate']*100:.0f}% 할인"
                    )
                st.caption("→ 이번 주문이 등급 문턱에 걸쳐있어 구간별로 나눠 계산된 결과입니다")

        if item_type == "cigar":
            st.caption("정상가 기준 = import_item.retail_price_krw / import_item.supply_price_krw (원가는 할인 미반영: import_item.korea_cost_krw)")
        else:
            st.caption("정상가 기준 = non_cigar_product_mst.retail_price / wholesale_price (원가 = purchase_price)")

        if st.button("도매 판매 저장", use_container_width=True, key="wh_insert_btn"):
            partner_id = int(partners.loc[partners["partner_name"] == partner_name, "id"].iloc[0])

            supply_amount, vat_amount, total_amount_vat, profit_amount = _render_amount_summary(qty, unit_price, supply_price, unit_cost)

            auto_note = build_tiered_pricing_note(pricing_result)
            final_notes = notes.strip()
            if auto_note:
                final_notes = f"{final_notes} / {auto_note}" if final_notes else auto_note

            grade_code_applied = pricing_result.get("final_grade_code") if pricing_result else None
            if base_supply_price_ref > 0:
                # 실제 저장되는 공급가(수동 수정 가능) 기준으로 적용할인율을 재계산 — 표시값과 실제 저장값을 일치시킴
                discount_rate_applied = max(0.0, 1 - (float(supply_price) / base_supply_price_ref))
            else:
                discount_rate_applied = pricing_result.get("discount_ratio", 0.0) if pricing_result else 0.0

            normal_supply_amount = float(qty) * base_supply_price_ref if base_supply_price_ref > 0 else float(supply_amount)

            insert_wholesale_sale(
                conn=conn,
                sale_date=sale_date,
                partner_id=partner_id,
                item_type=item_type,
                cigar_product_id=cigar_product_id,
                non_cigar_product_id=non_cigar_product_id,
                qty=int(qty),
                unit_price=float(unit_price),
                supply_price=float(supply_price),
                unit_cost=float(unit_cost),
                supply_amount=float(supply_amount),
                vat_amount=float(vat_amount),
                total_amount_vat=float(total_amount_vat),
                profit_amount=float(profit_amount),
                notes=final_notes,
                grade_code_applied=grade_code_applied,
                discount_rate_applied=float(discount_rate_applied),
                normal_supply_amount=float(normal_supply_amount),
            )
            st.success("도매 판매 이력이 저장되었습니다." + (" (비고에 혼합할인 내역 자동 기록됨)" if auto_note else ""))
            st.rerun()
        else:
            _render_amount_summary(qty, unit_price, supply_price, unit_cost)

    st.divider()
    st.markdown("#### 도매 판매 내역")

    df = load_wholesale_sales_for_grid(conn)
    if df.empty:
        st.info("등록된 도매 판매 이력이 없습니다.")
        return

    today = pd.Timestamp.today().date()

    with st.expander("검색 / 필터", expanded=True):
        f1, f2 = st.columns(2)
        with f1:
            keyword = st.text_input("거래처 / 상품 / 코드 검색", key="manage_keyword")
            item_type_filter = st.selectbox("상품구분", ["전체", "시가", "시가 외"], key="manage_item_type")
        with f2:
            partner_options = ["전체"] + sorted([str(x) for x in df["partner_name"].dropna().unique().tolist()])
            partner_filter = st.selectbox("거래처", partner_options, key="manage_partner")
            d1, d2 = st.columns(2)
            with d1:
                date_from = st.date_input("판매일 시작", value=today, key="manage_date_from")
            with d2:
                date_to = st.date_input("판매일 종료", value=today, key="manage_date_to")

    filtered = df.copy()
    filtered["sale_date"] = pd.to_datetime(filtered["sale_date"], errors="coerce")
    filtered = filtered[
        (filtered["sale_date"].dt.date >= date_from) &
        (filtered["sale_date"].dt.date <= date_to)
    ]
    filtered["sale_date"] = filtered["sale_date"].dt.strftime("%Y-%m-%d")

    if keyword.strip():
        kw = keyword.strip().lower()
        mask = (
            filtered.get("partner_name", pd.Series("", index=filtered.index)).fillna("").astype(str).str.lower().str.contains(kw)
            | filtered.get("product_name", pd.Series("", index=filtered.index)).fillna("").astype(str).str.lower().str.contains(kw)
            | filtered.get("product_code", pd.Series("", index=filtered.index)).fillna("").astype(str).str.lower().str.contains(kw)
        )
        filtered = filtered[mask]

    if item_type_filter != "전체":
        filtered = filtered[filtered["item_type_label"] == item_type_filter]

    if partner_filter != "전체":
        filtered = filtered[filtered["partner_name"] == partner_filter]

    if filtered.empty:
        st.warning("검색 결과가 없습니다.")
        return

    sel_col1, sel_col2 = st.columns([3, 1])
    with sel_col1:
        select_all_filtered = st.checkbox(
            f"검색 결과 전체 선택 ({len(filtered):,}건) — 일괄 거래명세서 다운로드",
            key="wholesale_select_all",
        )
    with sel_col2:
        st.caption("※ 거래처·판매일자별로 각각 발행됩니다.")

    # ── st.dataframe 기반 조회 테이블 ──
    display_cols = {
        "id": "ID",
        "sale_date": "판매일자",
        "partner_name": "거래처명",
        "item_type_label": "상품구분",
        "product_code": "상품코드",
        "product_name": "상품명",
        "qty": "수량",
        "supply_price": "공급가(₩)",
        "supply_amount": "공급가액(₩)",
        "grade_code_applied": "적용등급",
        "discount_rate_applied": "적용할인율",
        "vat_amount": "부가세(₩)",
        "total_amount_vat": "부가세포함(₩)",
        "profit_amount": "예상마진(₩)",
        "notes": "비고",
    }
    available_cols = [c for c in display_cols if c in filtered.columns]
    grid_df = filtered[available_cols].rename(columns=display_cols).copy()
    if "적용할인율" in grid_df.columns:
        grid_df["적용할인율"] = pd.to_numeric(grid_df["적용할인율"], errors="coerce").map(
            lambda x: f"{x*100:.1f}%" if pd.notna(x) and x > 0 else ("-" if pd.notna(x) else "")
        )
    for col in ["공급가(₩)", "공급가액(₩)", "부가세(₩)", "부가세포함(₩)", "예상마진(₩)"]:
        if col in grid_df.columns:
            grid_df[col] = pd.to_numeric(grid_df[col], errors="coerce").map(
                lambda x: f"{x:,.0f}" if pd.notna(x) else ""
            )
    st.dataframe(grid_df, use_container_width=True, hide_index=True, height=380)

    # ── 행 선택: multiselect (ID 기준) ──
    id_label_map = {
        row["id"]: (
            f"{row['id']} | {row.get('sale_date','')} | {row.get('partner_name','')} | "
            f"{row.get('product_name','')} | {int(_safe_float(row.get('qty', 0)))}개"
        )
        for _, row in filtered.iterrows()
    }

    if not select_all_filtered:
        selected_labels = st.multiselect(
            "거래명세서 발행 / 수정·삭제할 행 선택 (ID 기준)",
            options=list(id_label_map.values()),
            key="wholesale_row_select",
        )
        selected_ids = [k for k, v in id_label_map.items() if v in selected_labels]
        selected_grid_df = filtered[filtered["id"].isin(selected_ids)].copy() if selected_ids else pd.DataFrame()
    else:
        selected_grid_df = filtered.copy()

    if not selected_grid_df.empty:
        groups_for_stmt = selected_grid_df.groupby(["partner_name", "sale_date"], sort=False)
        group_count = groups_for_stmt.ngroups

        if group_count == 1:
            ok, msg = _validate_statement_rows(selected_grid_df)
            if ok:
                preview_header, statement_df, totals = _build_statement_from_rows(conn, selected_grid_df)
                _render_statement_preview(preview_header, statement_df, totals)

                safe_partner = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in preview_header["partner_name"]).strip("_") or "거래처"

                try:
                    document_no = issue_statement_document_no(conn, preview_header["sale_date"])
                    final_header = dict(preview_header)
                    final_header["document_no"] = document_no

                    excel_bytes = _make_statement_excel_bytes(final_header, statement_df, totals)
                    file_name = f"거래명세서_{safe_partner}_{document_no}.xlsx"

                    if st.download_button(
                        "거래명세서 엑셀 다운로드",
                        data=excel_bytes,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="download_statement_excel",
                    ):
                        try:
                            save_statement_history(conn, final_header, totals, file_name)
                        except sqlite3.IntegrityError:
                            pass
                        st.success(f"거래명세서가 발행되었습니다. 문서번호: {document_no}")
                except Exception as e:
                    st.error(f"거래명세서 생성 중 오류가 발생했습니다: {e}")
                    st.caption(
                        f"템플릿 파일 경로: {STATEMENT_TEMPLATE_PATH} "
                        "(프로젝트 루트의 templates/statement_template.xlsx 권장)"
                    )
            else:
                st.warning(msg)

        else:
            st.divider()
            st.markdown(f"#### 거래명세서 일괄 발행 — {group_count}건 (거래처·판매일자별)")
            try:
                zip_bytes, ok_count, errors = _make_bulk_statement_zip_bytes(conn, selected_grid_df)
                zip_file_name = "거래명세서_일괄.zip"

                st.download_button(
                    f"📥 거래명세서 ZIP 다운로드 ({ok_count}건)",
                    data=zip_bytes,
                    file_name=zip_file_name,
                    mime="application/zip",
                    use_container_width=True,
                    key="download_bulk_statement_zip",
                )
                if ok_count:
                    st.success(f"{ok_count}건의 거래명세서가 ZIP으로 준비되었습니다.")
                if errors:
                    for err_msg in errors:
                        st.warning(f"⚠️ 건너뜀: {err_msg}")
            except Exception as e:
                st.error(f"일괄 거래명세서 생성 중 오류가 발생했습니다: {e}")

    if not select_all_filtered and not selected_grid_df.empty is False and selected_grid_df.empty:
        st.info("수정/삭제 또는 거래명세서 출력을 위해 행을 하나 이상 선택하거나, 전체 선택 체크박스를 사용하세요.")
        return

    if select_all_filtered or selected_grid_df.empty:
        return

    # 단건 수정/삭제는 multiselect에서 1개만 선택된 경우
    if len(selected_grid_df) != 1:
        st.info("수정/삭제는 행을 정확히 1개 선택하세요. 거래명세서 발행은 위에서 확인하세요.")
        return

    selected_id = _safe_int(selected_grid_df.iloc[0]["id"])
    base_row = load_wholesale_sales(conn)
    base_row = base_row.loc[base_row["id"] == selected_id]
    if base_row.empty:
        st.error("선택한 도매 판매 건을 찾을 수 없습니다.")
        return
    row = base_row.iloc[0]

    partner_names = partners["partner_name"].tolist()
    partner_name_map = dict(zip(partners["id"], partners["partner_name"]))
    current_partner_name = partner_name_map.get(row["partner_id"], partner_names[0] if partner_names else "")

    st.divider()
    st.markdown("#### 선택 건 수정 / 삭제")

    item_type = str(row["item_type"])
    item_type_label = "시가" if item_type == "cigar" else "시가 외"
    cigar_product_id = row["cigar_product_id"] if "cigar_product_id" in row.index else None
    non_cigar_product_id = row["non_cigar_product_id"] if "non_cigar_product_id" in row.index else None

    c1, c2 = st.columns(2)
    with c1:
        sale_date = str(st.date_input("판매일자", value=pd.to_datetime(row["sale_date"]).date(), key=f"edit_sale_date_{selected_id}"))
        partner_name = st.selectbox(
            "거래처",
            partner_names,
            index=partner_names.index(current_partner_name) if current_partner_name in partner_names else 0,
            key=f"edit_partner_{selected_id}",
        )
        qty = st.number_input("수량", min_value=1, value=_safe_int(row["qty"], 1), step=1, format="%d", key=f"edit_qty_{selected_id}")

    selected_product_code = ""
    auto_unit_price = _safe_float(row.get("unit_price", 0))
    auto_supply_price = _safe_float(row.get("supply_price", 0))
    auto_unit_cost = _safe_float(row.get("unit_cost", 0))

    if item_type == "cigar":
        if cigar_products.empty:
            st.warning("시가 상품 기준정보가 없습니다.")
            return

        product_options = cigar_products.copy()
        product_options["display_name"] = product_options.apply(
            lambda r: f"{str(r['product_code']).strip()} | {str(r['product_name']).strip()}"
            if str(r["product_code"]).strip() else str(r["product_name"]).strip(),
            axis=1,
        )

        default_idx = 0
        if pd.notna(cigar_product_id):
            matched_idx = product_options.index[product_options["id"] == _safe_int(cigar_product_id)].tolist()
            if matched_idx:
                default_idx = product_options.index.get_loc(matched_idx[0])

        selected_display = st.selectbox(
            "시가 상품",
            product_options["display_name"].tolist(),
            index=default_idx,
            key=f"edit_cigar_product_{selected_id}",
        )
        selected_product = product_options.loc[product_options["display_name"] == selected_display].iloc[0]
        cigar_product_id = _safe_int(selected_product["id"])
        non_cigar_product_id = None
        selected_product_code = str(selected_product["product_code"] or "")
        auto_unit_price = _safe_float(selected_product.get("retail_price_krw", auto_unit_price))
        auto_supply_price = _safe_float(selected_product.get("supply_price_krw", auto_supply_price))
        auto_unit_cost = _safe_float(selected_product.get("korea_cost_krw", auto_unit_cost))

        _bind_price_cost_by_selection(
            state_prefix=f"edit_{selected_id}",
            selected_product_key=f"cigar::{cigar_product_id}",
            auto_unit_price=auto_unit_price,
            auto_supply_price=auto_supply_price,
            auto_unit_cost=auto_unit_cost,
        )
    else:
        if non_cigar_products.empty:
            st.warning("시가 외 상품 기준정보가 없습니다.")
            return

        default_nc_idx = 0
        if pd.notna(non_cigar_product_id):
            matched_idx = non_cigar_products.index[non_cigar_products["id"] == _safe_int(non_cigar_product_id)].tolist()
            if matched_idx:
                default_nc_idx = non_cigar_products.index.get_loc(matched_idx[0])

        selected_product_name = st.selectbox(
            "시가 외 상품",
            non_cigar_products["product_name"].tolist(),
            index=default_nc_idx,
            key=f"edit_non_cigar_product_{selected_id}",
        )
        selected_product = non_cigar_products.loc[non_cigar_products["product_name"] == selected_product_name].iloc[0]
        non_cigar_product_id = _safe_int(selected_product["id"])
        cigar_product_id = None

        _bind_price_cost_by_selection(
            state_prefix=f"edit_{selected_id}",
            selected_product_key=f"non_cigar::{non_cigar_product_id}",
            auto_unit_price=_safe_float(row.get("unit_price", 0)),
            auto_supply_price=_safe_float(row.get("supply_price", 0)),
            auto_unit_cost=_safe_float(row.get("unit_cost", 0)),
        )

    with c2:
        unit_price = st.number_input("단가 (₩)", min_value=0, step=100, format="%d", key=f"edit_{selected_id}_unit_price")
        supply_price = st.number_input("공급가 (₩)", min_value=0, step=100, format="%d", key=f"edit_{selected_id}_supply_price")
        unit_cost = st.number_input("원가 (₩)", min_value=0, step=100, format="%d", key=f"edit_{selected_id}_unit_cost")

    st.caption(
        f"단가: ₩{unit_price:,.0f} / "
        f"공급가: ₩{supply_price:,.0f} / "
        f"원가: ₩{unit_cost:,.0f}"
    )

    notes = st.text_area(
        "비고",
        value=str(row.get("notes", "") or ""),
        key=f"edit_notes_{selected_id}"
    )

    st.caption(f"상품구분: {item_type_label}")
    if selected_product_code:
        st.caption(f"상품코드: {selected_product_code}")

    supply_amount, vat_amount, total_amount_vat, profit_amount = _render_amount_summary(qty, unit_price, supply_price, unit_cost)

    b1, b2 = st.columns(2)
    with b1:
        if st.button("선택 건 수정 저장", use_container_width=True, key=f"save_sale_{selected_id}"):
            partner_id = int(partners.loc[partners["partner_name"] == partner_name, "id"].iloc[0])
            update_wholesale_sale(
                conn=conn,
                sale_id=selected_id,
                sale_date=sale_date,
                partner_id=partner_id,
                item_type=item_type,
                cigar_product_id=int(cigar_product_id) if cigar_product_id is not None else None,
                non_cigar_product_id=int(non_cigar_product_id) if non_cigar_product_id is not None else None,
                qty=int(qty),
                unit_price=float(unit_price),
                supply_price=float(supply_price),
                unit_cost=float(unit_cost),
                supply_amount=float(supply_amount),
                vat_amount=float(vat_amount),
                total_amount_vat=float(total_amount_vat),
                profit_amount=float(profit_amount),
                notes=notes.strip(),
            )
            st.success("도매 판매 건이 수정되었습니다.")
            st.rerun()

    with b2:
        confirm_delete = st.checkbox("삭제 확인", key=f"confirm_delete_{selected_id}")
        if st.button("선택 건 삭제", use_container_width=True, type="secondary", key=f"delete_sale_{selected_id}"):
            if not confirm_delete:
                st.error("삭제 확인을 체크해 주세요.")
            else:
                delete_wholesale_sale(conn, selected_id)
                st.success("도매 판매 건이 삭제되었습니다.")
                st.rerun()


def render_daily_sales_summary(conn):
    st.markdown("### 일자별 판매 현황")
    st.caption("기간/거래처/상품으로 검색하고, 일자별 판매 금액 추이를 그래프로 확인할 수 있습니다.")

    df = load_wholesale_sales_for_grid(conn)
    if df.empty:
        st.info("등록된 도매 판매 이력이 없습니다.")
        return

    # 데이터 정제
    work_df = df.copy()
    work_df["sale_date"] = pd.to_datetime(work_df["sale_date"], errors="coerce")
    work_df = work_df.dropna(subset=["sale_date"])

    if work_df.empty:
        st.info("유효한 판매일자 데이터가 없습니다.")
        return

    for col in ["qty", "supply_amount", "vat_amount", "total_amount_vat", "profit_amount"]:
        if col in work_df.columns:
            work_df[col] = pd.to_numeric(work_df[col], errors="coerce").fillna(0)
        else:
            work_df[col] = 0

    # 거래건수 집계용 카운트 컬럼 (id 컬럼 유무와 무관하게 동작 보장)
    work_df["_row_count"] = 1

    # 검색/필터 영역
    today = pd.Timestamp.today().date()
    min_date = work_df["sale_date"].min().date()
    max_date = work_df["sale_date"].max().date()
    default_from = max(min_date, (pd.Timestamp.today() - pd.Timedelta(days=30)).date())
    default_to = max(max_date, today)

    with st.expander("검색 / 필터", expanded=True):
        f1, f2 = st.columns(2)
        with f1:
            d1, d2 = st.columns(2)
            with d1:
                date_from = st.date_input(
                    "조회 시작일",
                    value=default_from,
                    key="daily_summary_date_from",
                )
            with d2:
                date_to = st.date_input(
                    "조회 종료일",
                    value=default_to,
                    key="daily_summary_date_to",
                )
            keyword = st.text_input(
                "거래처 / 상품 / 코드 검색",
                key="daily_summary_keyword",
            )
        with f2:
            partner_options = ["전체"] + sorted(
                [str(x) for x in work_df["partner_name"].dropna().unique().tolist()]
            )
            partner_filter = st.selectbox(
                "거래처",
                partner_options,
                key="daily_summary_partner",
            )
            item_type_filter = st.selectbox(
                "상품구분",
                ["전체", "시가", "시가 외"],
                key="daily_summary_item_type",
            )
            amount_metric_label = st.selectbox(
                "금액 기준",
                ["공급가액", "부가세포함 금액", "이익"],
                key="daily_summary_metric",
            )

    metric_col_map = {
        "공급가액": "supply_amount",
        "부가세포함 금액": "total_amount_vat",
        "이익": "profit_amount",
    }
    metric_col = metric_col_map[amount_metric_label]

    # 필터 적용
    if date_from > date_to:
        st.error("조회 시작일은 종료일보다 빠르거나 같아야 합니다.")
        return

    filtered = work_df[
        (work_df["sale_date"].dt.date >= date_from)
        & (work_df["sale_date"].dt.date <= date_to)
    ].copy()

    if keyword.strip():
        kw = keyword.strip().lower()
        mask = (
            filtered.get("partner_name", pd.Series("", index=filtered.index))
            .fillna("").astype(str).str.lower().str.contains(kw)
            | filtered.get("product_name", pd.Series("", index=filtered.index))
            .fillna("").astype(str).str.lower().str.contains(kw)
            | filtered.get("product_code", pd.Series("", index=filtered.index))
            .fillna("").astype(str).str.lower().str.contains(kw)
        )
        filtered = filtered[mask]

    if partner_filter != "전체":
        filtered = filtered[filtered["partner_name"] == partner_filter]

    if item_type_filter != "전체":
        filtered = filtered[filtered["item_type_label"] == item_type_filter]

    if filtered.empty:
        st.warning("검색 결과가 없습니다.")
        return

    # 요약 지표
    total_supply = float(filtered["supply_amount"].sum())
    total_vat = float(filtered["vat_amount"].sum())
    total_with_vat = float(filtered["total_amount_vat"].sum())
    total_profit = float(filtered["profit_amount"].sum())
    total_qty = int(filtered["qty"].sum())
    sale_days = filtered["sale_date"].dt.date.nunique()

    m1, m2, m3, m4 = st.columns(4)
    with m1:
        st.metric("판매일수", f"{sale_days:,}일")
    with m2:
        st.metric("총 수량", f"{total_qty:,}")
    with m3:
        st.metric("공급가액 합계", f"₩{total_supply:,.0f}")
    with m4:
        st.metric("부가세포함 합계", f"₩{total_with_vat:,.0f}")

    m5, m6, m7 = st.columns(3)
    with m5:
        st.metric("부가세 합계", f"₩{total_vat:,.0f}")
    with m6:
        st.metric("이익 합계", f"₩{total_profit:,.0f}")
    with m7:
        avg_per_day = (total_supply / sale_days) if sale_days else 0
        st.metric("일평균 공급가액", f"₩{avg_per_day:,.0f}")

    st.divider()

    # 일자별 집계
    grouped_src = filtered.copy()
    grouped_src["판매일자"] = grouped_src["sale_date"].dt.date

    daily_df = (
        grouped_src.groupby("판매일자", as_index=False)
        .agg(
            qty=("qty", "sum"),
            supply_amount=("supply_amount", "sum"),
            vat_amount=("vat_amount", "sum"),
            total_amount_vat=("total_amount_vat", "sum"),
            profit_amount=("profit_amount", "sum"),
            transaction_count=("_row_count", "sum"),
        )
        .sort_values("판매일자")
        .reset_index(drop=True)
    )

    # 그래프
    st.markdown(f"#### 일자별 {amount_metric_label} 추이")

    chart_df = daily_df[["판매일자", metric_col]].copy()
    chart_df["판매일자"] = pd.to_datetime(chart_df["판매일자"])
    chart_df = chart_df.rename(columns={metric_col: amount_metric_label})

    chart_type = st.radio(
        "그래프 형태",
        ["막대그래프", "선그래프"],
        horizontal=True,
        key="daily_summary_chart_type",
    )

    import plotly.graph_objects as go

    fig = go.Figure()
    if chart_type == "막대그래프":
        fig.add_trace(
            go.Bar(
                x=chart_df["판매일자"],
                y=chart_df[amount_metric_label],
                name=amount_metric_label,
                marker_color="#4C78A8",
                hovertemplate="%{x|%Y-%m-%d}<br>₩%{y:,.0f}<extra></extra>",
            )
        )
    else:
        fig.add_trace(
            go.Scatter(
                x=chart_df["판매일자"],
                y=chart_df[amount_metric_label],
                mode="lines+markers",
                name=amount_metric_label,
                line=dict(color="#4C78A8", width=2),
                marker=dict(size=6),
                hovertemplate="%{x|%Y-%m-%d}<br>₩%{y:,.0f}<extra></extra>",
            )
        )

    fig.update_layout(
        height=320,
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(
            title=None,
            fixedrange=True,  # 줌/팬 비활성화
            tickformat="%Y-%m-%d",
        ),
        yaxis=dict(
            title=amount_metric_label,
            fixedrange=True,  # 줌/팬 비활성화
            tickformat=",.0f",
        ),
        showlegend=False,
        dragmode=False,
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "displayModeBar": False,  # Plotly 툴바 숨김
            "scrollZoom": False,
            "doubleClick": False,
            "staticPlot": False,  # 호버는 유지
        },
    )

    st.divider()

    # 일자별 표
    st.markdown("#### 일자별 판매 집계")

    display_daily = daily_df.copy()
    display_daily["판매일자"] = pd.to_datetime(display_daily["판매일자"]).dt.strftime("%Y-%m-%d")
    display_daily = display_daily.rename(
        columns={
            "qty": "수량",
            "supply_amount": "공급가액(₩)",
            "vat_amount": "부가세(₩)",
            "total_amount_vat": "부가세포함(₩)",
            "profit_amount": "이익(₩)",
            "transaction_count": "거래건수",
        }
    )[["판매일자", "거래건수", "수량", "공급가액(₩)", "부가세(₩)", "부가세포함(₩)", "이익(₩)"]]

    # 합계 행 추가
    total_row = pd.DataFrame(
        [
            {
                "판매일자": "합계",
                "거래건수": int(display_daily["거래건수"].sum()),
                "수량": int(display_daily["수량"].sum()),
                "공급가액(₩)": float(display_daily["공급가액(₩)"].sum()),
                "부가세(₩)": float(display_daily["부가세(₩)"].sum()),
                "부가세포함(₩)": float(display_daily["부가세포함(₩)"].sum()),
                "이익(₩)": float(display_daily["이익(₩)"].sum()),
            }
        ]
    )
    display_daily_with_total = pd.concat([display_daily, total_row], ignore_index=True)

    st.dataframe(
        display_daily_with_total.style.format(
            {
                "거래건수": "{:,.0f}",
                "수량": "{:,.0f}",
                "공급가액(₩)": "₩{:,.0f}",
                "부가세(₩)": "₩{:,.0f}",
                "부가세포함(₩)": "₩{:,.0f}",
                "이익(₩)": "₩{:,.0f}",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    # CSV 다운로드
    csv_bytes = display_daily.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "일자별 집계 CSV 다운로드",
        data=csv_bytes,
        file_name=f"daily_sales_{date_from}_{date_to}.csv",
        mime="text/csv",
        use_container_width=True,
        key="daily_summary_csv_download",
    )

    # 상세 내역 (선택적으로 펼쳐보기)
    with st.expander(f"상세 거래 내역 보기 ({len(filtered):,}건)", expanded=False):
        detail_cols = {
            "sale_date": "판매일자",
            "partner_name": "거래처명",
            "item_type_label": "상품구분",
            "product_code": "상품코드",
            "product_name": "상품명",
            "qty": "수량",
            "supply_price": "공급가(₩)",
            "supply_amount": "공급가액(₩)",
            "vat_amount": "부가세(₩)",
            "total_amount_vat": "부가세포함(₩)",
            "profit_amount": "이익(₩)",
        }
        existing_detail_cols = [c for c in detail_cols.keys() if c in filtered.columns]
        detail_df = filtered[existing_detail_cols].copy()
        detail_df["sale_date"] = pd.to_datetime(detail_df["sale_date"]).dt.strftime("%Y-%m-%d")
        detail_df = detail_df.rename(columns=detail_cols).sort_values(
            ["판매일자", "거래처명"], ascending=[False, True]
        )

        money_cols = [c for c in ["공급가(₩)", "공급가액(₩)", "부가세(₩)", "부가세포함(₩)", "이익(₩)"] if c in detail_df.columns]
        format_dict = {c: "₩{:,.0f}" for c in money_cols}
        if "수량" in detail_df.columns:
            format_dict["수량"] = "{:,.0f}"

        st.dataframe(
            detail_df.style.format(format_dict),
            use_container_width=True,
            hide_index=True,
        )


def render_partner_sales_summary(conn):
    st.markdown("### 거래처별 매출 현황")
    st.caption("기간을 설정하면 해당 기간 동안 거래처별 매출/이익을 한눈에 비교할 수 있습니다.")

    df = load_wholesale_sales_for_grid(conn)
    if df.empty:
        st.info("등록된 도매 판매 이력이 없습니다.")
        return

    work_df = df.copy()
    work_df["sale_date"] = pd.to_datetime(work_df["sale_date"], errors="coerce")
    work_df = work_df.dropna(subset=["sale_date"])

    if work_df.empty:
        st.info("유효한 판매일자 데이터가 없습니다.")
        return

    for col in ["qty", "supply_amount", "vat_amount", "total_amount_vat", "profit_amount"]:
        if col in work_df.columns:
            work_df[col] = pd.to_numeric(work_df[col], errors="coerce").fillna(0)
        else:
            work_df[col] = 0

    work_df["_row_count"] = 1
    work_df["partner_name"] = work_df.get("partner_name", "").fillna("(거래처 미상)")

    today = pd.Timestamp.today().date()
    min_date = work_df["sale_date"].min().date()
    max_date = work_df["sale_date"].max().date()
    default_from = max(min_date, (pd.Timestamp.today() - pd.Timedelta(days=30)).date())
    default_to = max(max_date, today)

    with st.expander("검색 / 필터", expanded=True):
        f1, f2 = st.columns(2)
        with f1:
            d1, d2 = st.columns(2)
            with d1:
                date_from = st.date_input(
                    "조회 시작일",
                    value=default_from,
                    key="partner_summary_date_from",
                )
            with d2:
                date_to = st.date_input(
                    "조회 종료일",
                    value=default_to,
                    key="partner_summary_date_to",
                )
        with f2:
            item_type_filter = st.selectbox(
                "상품구분",
                ["전체", "시가", "시가 외"],
                key="partner_summary_item_type",
            )
            sort_metric_label = st.selectbox(
                "정렬 기준",
                ["공급가액", "부가세포함 금액", "이익"],
                key="partner_summary_sort_metric",
            )

    metric_col_map = {
        "공급가액": "supply_amount",
        "부가세포함 금액": "total_amount_vat",
        "이익": "profit_amount",
    }
    sort_col = metric_col_map[sort_metric_label]

    if date_from > date_to:
        st.error("조회 시작일은 종료일보다 빠르거나 같아야 합니다.")
        return

    filtered = work_df[
        (work_df["sale_date"].dt.date >= date_from)
        & (work_df["sale_date"].dt.date <= date_to)
    ].copy()

    if item_type_filter != "전체" and "item_type_label" in filtered.columns:
        filtered = filtered[filtered["item_type_label"] == item_type_filter]

    if filtered.empty:
        st.warning("검색 결과가 없습니다.")
        return

    total_supply = float(filtered["supply_amount"].sum())
    total_with_vat = float(filtered["total_amount_vat"].sum())
    total_profit = float(filtered["profit_amount"].sum())
    total_qty = int(filtered["qty"].sum())
    partner_count = filtered["partner_name"].nunique()

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("거래처 수", f"{partner_count:,}곳")
    m2.metric("총 수량", f"{total_qty:,}")
    m3.metric("공급가액 합계", f"₩{total_supply:,.0f}")
    m4.metric("이익 합계", f"₩{total_profit:,.0f}")

    st.divider()

    # 거래처별 집계
    partner_df = (
        filtered.groupby("partner_name", as_index=False)
        .agg(
            transaction_count=("_row_count", "sum"),
            qty=("qty", "sum"),
            supply_amount=("supply_amount", "sum"),
            vat_amount=("vat_amount", "sum"),
            total_amount_vat=("total_amount_vat", "sum"),
            profit_amount=("profit_amount", "sum"),
        )
        .sort_values(sort_col, ascending=False)
        .reset_index(drop=True)
    )
    partner_df["profit_margin_pct"] = partner_df.apply(
        lambda r: (r["profit_amount"] / r["supply_amount"] * 100) if r["supply_amount"] else 0,
        axis=1,
    )

    # 그래프 (상위 거래처 랭킹)
    st.markdown(f"#### 거래처별 {sort_metric_label} 순위")

    top_n = min(20, len(partner_df))
    chart_df = partner_df.head(top_n).sort_values(sort_col, ascending=True)

    import plotly.graph_objects as go

    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=chart_df[sort_col],
            y=chart_df["partner_name"],
            orientation="h",
            marker_color="#4C78A8",
            hovertemplate="%{y}<br>₩%{x:,.0f}<extra></extra>",
        )
    )
    fig.update_layout(
        height=max(320, 28 * top_n),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(title=sort_metric_label, fixedrange=True, tickformat=",.0f"),
        yaxis=dict(title=None, fixedrange=True),
        showlegend=False,
        dragmode=False,
    )
    st.plotly_chart(
        fig,
        use_container_width=True,
        config={
            "displayModeBar": False,
            "scrollZoom": False,
            "doubleClick": False,
            "staticPlot": False,
        },
    )

    st.divider()

    # 거래처별 표
    st.markdown("#### 거래처별 매출 집계")

    display_df = partner_df.rename(
        columns={
            "partner_name": "거래처명",
            "transaction_count": "거래건수",
            "qty": "수량",
            "supply_amount": "공급가액(₩)",
            "vat_amount": "부가세(₩)",
            "total_amount_vat": "부가세포함(₩)",
            "profit_amount": "이익(₩)",
            "profit_margin_pct": "이익률(%)",
        }
    )[["거래처명", "거래건수", "수량", "공급가액(₩)", "부가세(₩)", "부가세포함(₩)", "이익(₩)", "이익률(%)"]]

    total_row = pd.DataFrame(
        [
            {
                "거래처명": "합계",
                "거래건수": int(display_df["거래건수"].sum()),
                "수량": int(display_df["수량"].sum()),
                "공급가액(₩)": float(display_df["공급가액(₩)"].sum()),
                "부가세(₩)": float(display_df["부가세(₩)"].sum()),
                "부가세포함(₩)": float(display_df["부가세포함(₩)"].sum()),
                "이익(₩)": float(display_df["이익(₩)"].sum()),
                "이익률(%)": (
                    float(display_df["이익(₩)"].sum()) / float(display_df["공급가액(₩)"].sum()) * 100
                    if display_df["공급가액(₩)"].sum()
                    else 0
                ),
            }
        ]
    )
    display_df_with_total = pd.concat([display_df, total_row], ignore_index=True)

    st.dataframe(
        display_df_with_total.style.format(
            {
                "거래건수": "{:,.0f}",
                "수량": "{:,.0f}",
                "공급가액(₩)": "₩{:,.0f}",
                "부가세(₩)": "₩{:,.0f}",
                "부가세포함(₩)": "₩{:,.0f}",
                "이익(₩)": "₩{:,.0f}",
                "이익률(%)": "{:,.1f}%",
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    csv_bytes = display_df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        "거래처별 매출 집계 CSV 다운로드",
        data=csv_bytes,
        file_name=f"partner_sales_{date_from}_{date_to}.csv",
        mime="text/csv",
        use_container_width=True,
        key="partner_summary_csv_download",
    )

    # 거래처 선택 시 품목별 상세
    st.divider()
    st.markdown("#### 거래처별 품목 상세")
    partner_pick = st.selectbox(
        "품목 상세를 볼 거래처 선택",
        options=partner_df["partner_name"].tolist(),
        key="partner_summary_detail_partner",
    )
    if partner_pick:
        item_detail = (
            filtered[filtered["partner_name"] == partner_pick]
            .groupby("product_name", as_index=False)
            .agg(
                qty=("qty", "sum"),
                supply_amount=("supply_amount", "sum"),
                profit_amount=("profit_amount", "sum"),
            )
            .sort_values("supply_amount", ascending=False)
            .rename(columns={
                "product_name": "상품명",
                "qty": "수량",
                "supply_amount": "공급가액(₩)",
                "profit_amount": "이익(₩)",
            })
        )
        st.dataframe(
            item_detail.style.format({"수량": "{:,.0f}", "공급가액(₩)": "₩{:,.0f}", "이익(₩)": "₩{:,.0f}"}),
            use_container_width=True,
            hide_index=True,
        )


def render():
    st.subheader("도매 관리")

    conn = get_conn()
    try:
        missing_tables = ensure_required_tables(conn)
        if missing_tables:
            st.error("필수 테이블이 없습니다: " + ", ".join(missing_tables))
            st.info("먼저 DB 초기화 SQL을 실행해 주세요.")
            return

        ensure_wholesale_sales_columns(conn)

        tab1, tab2, tab3, tab4 = st.tabs([
            "거래처 관리",
            "도매 판매 관리",
            "일자별 판매 현황",
            "거래처별 매출 현황",
        ])

        with tab1:
            render_partner_registration(conn)

        with tab2:
            render_wholesale_management(conn)

        with tab3:
            render_daily_sales_summary(conn)

        with tab4:
            render_partner_sales_summary(conn)

    finally:
        conn.close()

def issue_document_no(conn, doc_type: str, prefix: str, sale_date_str: str) -> str:
    ensure_statement_tables(conn)

    try:
        doc_date = pd.to_datetime(sale_date_str).strftime("%Y%m%d")
    except Exception:
        doc_date = pd.Timestamp.now().strftime("%Y%m%d")

    cur = conn.cursor()
    cur.execute(
        """
        SELECT last_seq
        FROM statement_doc_sequence
        WHERE doc_type = ? AND doc_date = ?
        """,
        (doc_type, doc_date),
    )
    row = cur.fetchone()

    if row:
        next_seq = _safe_int(row[0], 0) + 1
        cur.execute(
            """
            UPDATE statement_doc_sequence
            SET last_seq = ?, updated_at = CURRENT_TIMESTAMP
            WHERE doc_type = ? AND doc_date = ?
            """,
            (next_seq, doc_type, doc_date),
        )
    else:
        next_seq = 1
        cur.execute(
            """
            INSERT INTO statement_doc_sequence (doc_type, doc_date, last_seq)
            VALUES (?, ?, ?)
            """,
            (doc_type, doc_date, next_seq),
        )

    conn.commit()
    return f"{prefix}-{doc_date}-{next_seq:03d}"

def get_partner_detail_by_id(conn, partner_id: int) -> dict:
    if not partner_id:
        return {}

    df = pd.read_sql(
        """
        SELECT
            id,
            partner_name,
            owner_name,
            contact_name,
            phone,
            address,
            business_no
        FROM partner_mst
        WHERE id = ?
        """,
        conn,
        params=[partner_id],
    )
    return df.iloc[0].to_dict() if not df.empty else {}