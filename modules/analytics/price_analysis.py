
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import altair as alt
import io
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from copy import copy
from openpyxl.styles import Alignment
from matplotlib.ticker import FuncFormatter
from db import get_all_import_batch, get_price_analysis_view

def align_product_name_center(ws, start_row: int, end_row: int, product_col: int = 1):
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=product_col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def merge_same_product_name(ws, start_row: int, end_row: int, product_col: int = 1):
    """
    product_col=1 -> A열
    연속된 행에서 같은 제품명이면 세로 병합
    """
    if end_row < start_row:
        return

    merge_start = start_row
    prev_value = ws.cell(row=start_row, column=product_col).value

    for row in range(start_row + 1, end_row + 1):
        current_value = ws.cell(row=row, column=product_col).value

        if current_value != prev_value:
            if prev_value not in (None, "") and row - 1 > merge_start:
                ws.merge_cells(
                    start_row=merge_start,
                    start_column=product_col,
                    end_row=row - 1,
                    end_column=product_col,
                )
                top_cell = ws.cell(row=merge_start, column=product_col)
                top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            merge_start = row
            prev_value = current_value

    # 마지막 구간 처리
    if prev_value not in (None, "") and end_row > merge_start:
        ws.merge_cells(
            start_row=merge_start,
            start_column=product_col,
            end_row=end_row,
            end_column=product_col,
        )
        top_cell = ws.cell(row=merge_start, column=product_col)
        top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def setup_korean_font():
    candidates = [
        "Malgun Gothic",
        "AppleGothic",
        "NanumGothic",
        "NanumBarunGothic",
        "DejaVu Sans",
    ]
    available = {f.name for f in fm.fontManager.ttflist}

    for name in candidates:
        if name in available:
            plt.rcParams["font.family"] = name
            break

    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["font.size"] = 9


def format_krw(value):
    if pd.isna(value):
        return ""
    return f"₩{value:,.0f}"


def metric_with_caption(column, label: str, value: str, caption: str):
    column.metric(label, value)
    column.caption(caption)


def _safe_num(v, default=0):
    try:
        if pd.isna(v):
            return default
        return float(v)
    except Exception:
        return default


def _set_basic_border_and_font(ws, start_row: int, end_row: int, start_col: int = 1, end_col: int = 24):
    thin = Side(style="thin", color="000000")
    basic_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)

            if isinstance(cell, MergedCell):
                continue

            cell.font = copy(cell.font)
            cell.font = Font(
                name=cell.font.name or "맑은 고딕",
                size=10,
                bold=cell.font.bold,
                italic=cell.font.italic,
                underline=cell.font.underline,
                strike=cell.font.strike,
                color=cell.font.color
            )
            cell.border = basic_border


def _apply_number_formats(ws, row_idx: int):
    krw_cols = ["D", "F", "G", "H", "I", "J", "K", "L", "O", "P", "Q", "R", "S", "T", "U", "V"]
    php_cols = ["M", "N"]
    pct_cols = ["W", "X"]
    usd_cols = ["C"]

    for col in krw_cols:
        ws[f"{col}{row_idx}"].number_format = '₩#,##0'
    for col in php_cols:
        ws[f"{col}{row_idx}"].number_format = '₱#,##0'
    for col in pct_cols:
        ws[f"{col}{row_idx}"].number_format = '0.0%'
    for col in usd_cols:
        ws[f"{col}{row_idx}"].number_format = '$#,##0.00'

    ws[f"E{row_idx}"].number_format = '#,##0.00'


def build_price_analysis_excel(df: pd.DataFrame) -> io.BytesIO:
    template_path = Path("templates/가격분석_template.xlsx")

    if not template_path.exists():
        raise FileNotFoundError("templates 폴더에 가격분석 템플릿 파일이 없습니다. (예: templates/가격분석_template.xlsx)")

    wb = load_workbook(template_path)
    ws = wb["가격분석"]

    start_row = 3
    end_col = 24  # A:X
    export_df = df.copy()

    numeric_cols = [
        "export_unit_price_usd",
        "import_unit_cost_krw",
        "unit_weight_g",
        "individual_tax_krw",
        "tobacco_tax_krw",
        "local_education_tax_krw",
        "health_charge_krw",
        "import_vat_krw",
        "tax_total_krw",
        "korea_cost_krw",
        "local_box_price_php",
        "local_unit_price_php",
        "local_unit_price_krw",
        "proposal_retail_price_krw",
        "retail_price_krw",
        "store_retail_price_krw",
        "supply_price_krw",
        "supply_vat_krw",
        "margin_krw",
        "retail_margin_rate",
        "wholesale_margin_rate",
    ]
    for col in numeric_cols:
        if col in export_df.columns:
            export_df[col] = pd.to_numeric(export_df[col], errors="coerce")

    # 기존 데이터 영역 삭제
    if ws.max_row >= start_row:
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=1, max_col=end_col):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                cell.value = None
                cell.border = Border()
                cell.number_format = "General"

    # 데이터 입력
    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=start_row):
        ws[f"A{row_idx}"] = row.get("product_name", "")
        ws[f"B{row_idx}"] = row.get("size_name", "")
        ws[f"C{row_idx}"] = row.get("export_unit_price_usd", "")
        ws[f"D{row_idx}"] = row.get("import_unit_cost_krw", "")
        ws[f"E{row_idx}"] = row.get("unit_weight_g", "")

        ws[f"F{row_idx}"] = row.get("individual_tax_krw", "")
        ws[f"G{row_idx}"] = row.get("tobacco_tax_krw", "")
        ws[f"H{row_idx}"] = row.get("local_education_tax_krw", "")
        ws[f"I{row_idx}"] = row.get("health_charge_krw", "")
        ws[f"J{row_idx}"] = row.get("import_vat_krw", "")
        ws[f"K{row_idx}"] = row.get("tax_total_krw", "")
        ws[f"L{row_idx}"] = row.get("korea_cost_krw", "")

        ws[f"M{row_idx}"] = row.get("local_box_price_php", "")
        ws[f"N{row_idx}"] = row.get("local_unit_price_php", "")
        ws[f"O{row_idx}"] = row.get("local_unit_price_krw", "")

        ws[f"P{row_idx}"] = row.get("proposal_retail_price_krw", "")
        ws[f"Q{row_idx}"] = row.get("retail_price_krw", "")
        ws[f"R{row_idx}"] = row.get("store_retail_price_krw", "")
        ws[f"S{row_idx}"] = row.get("supply_price_krw", "")
        ws[f"T{row_idx}"] = row.get("supply_vat_krw", "")
        ws[f"U{row_idx}"] = f'=IFERROR(S{row_idx}+T{row_idx},0)'
        ws[f"V{row_idx}"] = row.get("margin_krw", "")
        ws[f"W{row_idx}"] = f'=IFERROR((Q{row_idx}-L{row_idx})/Q{row_idx},0)'
        ws[f"X{row_idx}"] = f'=IFERROR((S{row_idx}-L{row_idx})/S{row_idx},0)'

        # 정렬/표시
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if col in [1, 2]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")

        _apply_number_formats(ws, row_idx)

    final_row = max(start_row, start_row + len(export_df) - 1)

    # 헤더 포함 폰트 10 / 기본 테두리
    if len(export_df) > 0:
        _set_basic_border_and_font(ws, start_row=start_row, end_row=final_row, start_col=1, end_col=end_col)

    final_row = start_row + len(export_df) - 1

    if len(export_df) > 0:
        align_product_name_center(ws, start_row=start_row, end_row=final_row, product_col=1)
        merge_same_product_name(ws, start_row=start_row, end_row=final_row, product_col=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render():
    st.set_page_config(page_title="가격분석 통합조회", layout="wide")
    st.subheader("가격분석 통합조회")

    setup_korean_font()

    batch_df = get_all_import_batch()

    batch_options = {"전체": None}
    for _, row in batch_df.iterrows():
        batch_options[f'{row["id"]} | {row["version_name"]}'] = row["id"]

    c1, c2 = st.columns([2, 3])

    with c1:
        selected = st.selectbox("버전 선택", list(batch_options.keys()))
        batch_id = batch_options[selected]

    with c2:
        keyword = st.text_input("검색", placeholder="상품명 / 사이즈 / 코드")

    df = get_price_analysis_view(batch_id, keyword)

    if df.empty:
        st.info("조회된 데이터가 없습니다.")
        st.stop()

    numeric_cols = [
        "korea_cost_krw",
        "supply_price_krw",
        "retail_price_krw",
        "store_retail_price_krw",
        "proposal_retail_price_krw",
        "import_total_cost_krw",
        "tax_total_all_krw",
        "margin_krw",
        "local_box_price_php",
        "local_unit_price_php",
        "local_unit_price_krw",
        "individual_tax_krw",
        "tobacco_tax_krw",
        "local_education_tax_krw",
        "health_charge_krw",
        "import_vat_krw",
        "import_unit_qty",
        "unit_weight_g",
    ]

    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "supply_price_krw" in df.columns and "korea_cost_krw" in df.columns:
        df["wholesale_margin_krw"] = (
            df["supply_price_krw"].fillna(0) - df["korea_cost_krw"].fillna(0)
        ).clip(lower=0)
    else:
        df["wholesale_margin_krw"] = 0

    if "retail_price_krw" in df.columns and "korea_cost_krw" in df.columns:
        df["retail_margin_krw"] = (
            df["retail_price_krw"].fillna(0) - df["korea_cost_krw"].fillna(0)
        ).clip(lower=0)
    else:
        df["retail_margin_krw"] = 0

    if "retail_price_krw" in df.columns and "supply_price_krw" in df.columns:
        df["retail_extra_margin_krw"] = (
            df["retail_price_krw"].fillna(0) - df["supply_price_krw"].fillna(0)
        ).clip(lower=0)
    else:
        df["retail_extra_margin_krw"] = 0

    if "proposal_retail_price_krw" in df.columns and "korea_cost_krw" in df.columns:
        df["proposal_retail_margin_krw"] = (
            df["proposal_retail_price_krw"].fillna(0) - df["korea_cost_krw"].fillna(0)
        ).clip(lower=0)
    else:
        df["proposal_retail_margin_krw"] = 0

    if "proposal_retail_price_krw" in df.columns and "supply_price_krw" in df.columns:
        df["proposal_retail_extra_margin_krw"] = (
            df["proposal_retail_price_krw"].fillna(0) - df["supply_price_krw"].fillna(0)
        ).clip(lower=0)
    else:
        df["proposal_retail_extra_margin_krw"] = 0

    df["wholesale_margin_rate"] = df.apply(
        lambda r: (r["wholesale_margin_krw"] / r["supply_price_krw"] * 100)
        if pd.notna(r.get("supply_price_krw")) and r.get("supply_price_krw", 0) > 0
        else 0,
        axis=1,
    )
    df["retail_margin_rate"] = df.apply(
        lambda r: (r["retail_margin_krw"] / r["retail_price_krw"] * 100)
        if pd.notna(r.get("retail_price_krw")) and r.get("retail_price_krw", 0) > 0
        else 0,
        axis=1,
    )
    df["proposal_retail_margin_rate"] = df.apply(
        lambda r: (r["proposal_retail_margin_krw"] / r["proposal_retail_price_krw"] * 100)
        if pd.notna(r.get("proposal_retail_price_krw")) and r.get("proposal_retail_price_krw", 0) > 0
        else 0,
        axis=1,
    )

    st.subheader("핵심 지표")

    total_cost = df["korea_cost_krw"].fillna(0).sum() if "korea_cost_krw" in df.columns else 0
    avg_wholesale = df["wholesale_margin_rate"].fillna(0).mean() if "wholesale_margin_rate" in df.columns else 0
    avg_retail = df["retail_margin_rate"].fillna(0).mean() if "retail_margin_rate" in df.columns else 0

    k1, k2, k3, k4 = st.columns(4)
    metric_with_caption(k1, "총 제품 수", f"{len(df):,}개", "조회 결과 기준")
    metric_with_caption(k2, "총 한국원가", format_krw(total_cost), "조회 제품 원가 합계")
    metric_with_caption(k3, "평균 도매 마진율", f"{avg_wholesale:.1f}%", "조회 제품 평균")
    metric_with_caption(k4, "평균 소매 마진율", f"{avg_retail:.1f}%", "조회 제품 평균")

    d1, d2 = st.columns([1.2, 6])
    with d1:
        try:
            excel_data = build_price_analysis_excel(df)
            st.download_button(
                label="엑셀 다운로드",
                data=excel_data,
                file_name="price_analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"엑셀 다운로드 생성 오류: {e}")

    st.subheader("가격 분석 테이블")

    table_df = df.copy()
    table_df["소매마진율"] = table_df["retail_margin_rate"].fillna(0)
    table_df["도매마진율"] = table_df["wholesale_margin_rate"].fillna(0)

    preferred_order = [
        "version_name",
        "product_code",
        "product_name",
        "size_name",
        "discount_rate",
        "export_unit_price_usd",
        "import_unit_cost_krw",
        "unit_weight_g",
        "individual_tax_krw",
        "tobacco_tax_krw",
        "local_education_tax_krw",
        "health_charge_krw",
        "import_vat_krw",
        "tax_total_krw",
        "korea_cost_krw",
        "local_box_price_php",
        "local_unit_price_php",
        "local_unit_price_krw",
        "retail_price_krw",
        "supply_price_krw",
        "supply_vat_krw",
        "margin_krw",
        "retail_margin_rate",
        "wholesale_margin_rate",
        "store_retail_price_krw",
    ]

    existing_preferred = [c for c in preferred_order if c in table_df.columns]
    remaining_cols = [c for c in table_df.columns if c not in existing_preferred]
    table_df = table_df[existing_preferred + remaining_cols]

    styled_df = table_df.style.apply(highlight_columns, axis=None)

    st.dataframe(
        styled_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "version_name": "버전",
            "product_code": "코드",
            "product_name": "상품명",
            "size_name": "사이즈",
            "discount_rate": st.column_config.NumberColumn("할인율", format="%.1f%%"),
            "export_unit_price_usd": st.column_config.NumberColumn("수출가격(USD)", format="$%.2f"),
            "import_unit_cost_krw": st.column_config.NumberColumn("수입가격(KRW)", format="₩%.0f"),
            "unit_weight_g": st.column_config.NumberColumn("무게(g)", format="%.2f"),
            "individual_tax_krw": st.column_config.NumberColumn("개별소비세", format="₩%.0f"),
            "tobacco_tax_krw": st.column_config.NumberColumn("담배세", format="₩%.0f"),
            "local_education_tax_krw": st.column_config.NumberColumn("지방교육세", format="₩%.0f"),
            "health_charge_krw": st.column_config.NumberColumn("건강부담금", format="₩%.0f"),
            "import_vat_krw": st.column_config.NumberColumn("부가세", format="₩%.0f"),
            "tax_total_krw": st.column_config.NumberColumn("세금합계", format="₩%.0f"),
            "korea_cost_krw": st.column_config.NumberColumn("한국원가", format="₩%.0f"),
            "local_box_price_php": st.column_config.NumberColumn("현지박스가격(PHP)", format="₱%.0f"),
            "local_unit_price_php": st.column_config.NumberColumn("현지가격(PHP)", format="₱%.0f"),
            "local_unit_price_krw": st.column_config.NumberColumn("현지가격(KRW)", format="₩%.0f"),
            "retail_price_krw": st.column_config.NumberColumn("소비자가격", format="₩%.0f"),
            "supply_price_krw": st.column_config.NumberColumn("공급가", format="₩%.0f"),
            "supply_vat_krw": st.column_config.NumberColumn("공급가(VAT)", format="₩%.0f"),
            "margin_krw": st.column_config.NumberColumn("파트너마진", format="₩%.0f"),
            "retail_margin_rate": st.column_config.NumberColumn("소매마진율", format="%.1f%%"),
            "wholesale_margin_rate": st.column_config.NumberColumn("도매마진율", format="%.1f%%"),
            "store_retail_price_krw": st.column_config.NumberColumn("매장운영가격", format="₩%.0f"),
            "proposal_retail_price_krw": st.column_config.NumberColumn("제안소비자가", format="₩%.0f"),
            "소매마진율": None,
            "도매마진율": None,
            "retail_extra_margin_krw": None,
            "wholesale_margin_krw": None,
            "retail_margin_krw": None,
            "현지가격": None,
            "import_unit_qty": None,
            "total_weight_g": None,
        },
    )

    st.subheader("제품별 가격 구조 그래프")

    o1, o2, o3 = st.columns([1.6, 1.0, 1.6])

    with o1:
        view_mode = st.radio(
            "그래프 기준",
            ["소비자가 기준", "제안소비자가 기준", "도매 기준"],
            horizontal=True,
        )

    with o2:
        top10_only = st.toggle("Top 10만 보기", value=False)

    with o3:
        sort_by = st.selectbox(
            "정렬 기준",
            ["소비자가", "제안소비자가", "공급가", "한국원가", "도매마진", "소매마진"],
        )

    chart_df = df.copy()

    product_code = chart_df["product_code"].astype(str).fillna("") if "product_code" in chart_df.columns else ""
    product_name = chart_df["product_name"].astype(str).fillna("") if "product_name" in chart_df.columns else ""
    size_name = chart_df["size_name"].astype(str).fillna("") if "size_name" in chart_df.columns else ""

    chart_df["label"] = product_code
    empty_mask = chart_df["label"].str.strip().eq("")

    chart_df.loc[empty_mask, "label"] = (
        product_name[empty_mask].str.strip() + "\n" + size_name[empty_mask].str.strip()
    )

    sort_map = {
        "소비자가": "retail_price_krw",
        "제안소비자가": "proposal_retail_price_krw",
        "공급가": "supply_price_krw",
        "한국원가": "korea_cost_krw",
        "도매마진": "wholesale_margin_krw",
        "소매마진": (
            "retail_margin_krw"
            if view_mode == "소비자가 기준"
            else "proposal_retail_margin_krw"
            if view_mode == "제안소비자가 기준"
            else "wholesale_margin_krw"
        ),
    }

    sort_col = sort_map.get(sort_by)
    if sort_col in chart_df.columns:
        chart_df = chart_df.sort_values(sort_col, ascending=False)

    if top10_only:
        chart_df = chart_df.head(10)

    if view_mode == "소비자가 기준":
        chart_title = "제품별 가격 구조 (소비자가 기준)"
        required_cols = ["korea_cost_krw", "wholesale_margin_krw", "retail_extra_margin_krw"]
    elif view_mode == "제안소비자가 기준":
        chart_title = "제품별 가격 구조 (제안소비자가 기준)"
        required_cols = ["korea_cost_krw", "wholesale_margin_krw", "proposal_retail_extra_margin_krw"]
    else:
        chart_title = "제품별 가격 구조 (도매 기준)"
        required_cols = ["korea_cost_krw", "wholesale_margin_krw"]

    for col in required_cols:
        if col not in chart_df.columns:
            chart_df[col] = 0
        chart_df[col] = pd.to_numeric(chart_df[col], errors="coerce").fillna(0)

    render_pretty_stacked_bar_chart(
        chart_df=chart_df,
        view_mode=view_mode,
        chart_title=chart_title,
        sort_by=sort_by,
    )

    s1, s2 = st.columns(2)

    with s1:
        if "wholesale_margin_krw" in chart_df.columns:
            top_wholesale = chart_df.nlargest(
                min(5, len(chart_df)), "wholesale_margin_krw"
            )[["product_name", "size_name", "wholesale_margin_krw", "wholesale_margin_rate"]].copy()

            top_wholesale["도매마진"] = top_wholesale["wholesale_margin_krw"].apply(format_krw)
            top_wholesale["도매마진율"] = top_wholesale["wholesale_margin_rate"].fillna(0).map(lambda x: f"{x:.1f}%")
            top_wholesale = top_wholesale[["product_name", "size_name", "도매마진", "도매마진율"]]

            st.caption("도매 마진 상위 제품")
            st.table(top_wholesale)

    with s2:
        if view_mode == "제안소비자가 기준":
            margin_col = "proposal_retail_margin_krw"
            margin_rate_col = "proposal_retail_margin_rate"
            caption_text = "제안소비자가 기준 마진 상위 제품"
        else:
            margin_col = "retail_margin_krw"
            margin_rate_col = "retail_margin_rate"
            caption_text = "소매 마진 상위 제품"

        if margin_col in chart_df.columns:
            top_retail = chart_df.nlargest(
                min(5, len(chart_df)), margin_col
            )[["product_name", "size_name", margin_col, margin_rate_col]].copy()

            top_retail["소매마진"] = top_retail[margin_col].apply(format_krw)
            top_retail["소매마진율"] = top_retail[margin_rate_col].fillna(0).map(lambda x: f"{x:.1f}%")
            top_retail = top_retail[["product_name", "size_name", "소매마진", "소매마진율"]]

            st.caption(caption_text)
            st.table(top_retail)


def highlight_columns(dataframe: pd.DataFrame):
    styles = pd.DataFrame("", index=dataframe.index, columns=dataframe.columns)

    style_map = {
        "individual_tax_krw": "background-color: #FFF8FA; color: #111827;",
        "tobacco_tax_krw": "background-color: #FFF8FA; color: #111827;",
        "local_education_tax_krw": "background-color: #FFF8FA; color: #111827;",
        "health_charge_krw": "background-color: #FFF8FA; color: #111827;",
        "import_vat_krw": "background-color: #FFF8FA; color: #111827;",
        "tax_total_krw": "background-color: #FFF8FA; color: #111827;",
        "korea_cost_krw": "background-color: #FDECEF; color: #111827;",
        "retail_price_krw": "background-color: #EEF9E8; color: #111827;",
        "supply_price_krw": "background-color: #EAF3FF; color: #111827;",
        "store_retail_price_krw": "background-color: #FFF7E8; color: #111827;",
    }

    for col, style in style_map.items():
        if col in styles.columns:
            styles[col] = style

    return styles


def render_pretty_stacked_bar_chart(
    chart_df: pd.DataFrame,
    view_mode: str,
    chart_title: str,
    sort_by: str,
):
    if chart_df.empty:
        st.info("그래프 데이터가 없습니다.")
        return

    work = chart_df.copy()

    if view_mode == "소비자가 기준":
        part_defs = [
            ("한국원가", "korea_cost_krw"),
            ("도매 마진", "wholesale_margin_krw"),
            ("추가 소매 마진", "retail_extra_margin_krw"),
        ]
    elif view_mode == "제안소비자가 기준":
        part_defs = [
            ("한국원가", "korea_cost_krw"),
            ("도매 마진", "wholesale_margin_krw"),
            ("추가 소매 마진", "proposal_retail_extra_margin_krw"),
        ]
    else:
        part_defs = [
            ("한국원가", "korea_cost_krw"),
            ("도매 마진", "wholesale_margin_krw"),
        ]

    rows = []
    for _, r in work.iterrows():
        label = r["label"]
        for part_name, col in part_defs:
            val = pd.to_numeric(r.get(col, 0), errors="coerce")
            rows.append({"label": label, "구성": part_name, "금액": 0 if pd.isna(val) else float(val)})

    plot_df = pd.DataFrame(rows)

    if plot_df.empty:
        st.info("그래프 데이터가 없습니다.")
        return

    sort_map = {
        "소비자가": "retail_price_krw",
        "제안소비자가": "proposal_retail_price_krw",
        "공급가": "supply_price_krw",
        "한국원가": "korea_cost_krw",
        "도매마진": "wholesale_margin_krw",
        "소매마진": (
            "retail_margin_krw"
            if view_mode == "소비자가 기준"
            else "proposal_retail_margin_krw"
            if view_mode == "제안소비자가 기준"
            else "wholesale_margin_krw"
        ),
    }

    sort_col = sort_map.get(sort_by)

    if sort_col and sort_col in work.columns:
        order_df = (
            work[["label", sort_col]]
            .copy()
            .assign(_sort_value=lambda x: pd.to_numeric(x[sort_col], errors="coerce").fillna(0))
            .sort_values("_sort_value", ascending=False)
        )
        label_order = order_df["label"].tolist()
    else:
        order_df = plot_df.groupby("label", as_index=False)["금액"].sum().sort_values("금액", ascending=False)
        label_order = order_df["label"].tolist()

    chart = (
        alt.Chart(plot_df)
        .mark_bar(cornerRadiusEnd=4)
        .encode(
            y=alt.Y(
                "label:N",
                sort=label_order,
                title=None,
                scale=alt.Scale(paddingInner=0.12, paddingOuter=0),
                axis=alt.Axis(labelLimit=1000, labelFontSize=13, labelPadding=10, labelOverlap=False),
            ),
            x=alt.X("sum(금액):Q", title="금액(원)", axis=alt.Axis(format=",.0f", grid=True)),
            color=alt.Color(
                "구성:N",
                title=None,
                legend=alt.Legend(orient="top", direction="horizontal"),
                scale=alt.Scale(
                    domain=["한국원가", "도매 마진", "추가 소매 마진"],
                    range=["#dbeafe", "#93c5fd", "#2563eb"],
                ),
            ),
            tooltip=[
                alt.Tooltip("label:N", title="제품"),
                alt.Tooltip("구성:N", title="구성"),
                alt.Tooltip("금액:Q", title="금액", format=",.0f"),
            ],
        )
        .properties(title=chart_title, width=900, height=max(320, min(1700, len(label_order) * 22)))
        .configure_view(strokeWidth=0)
        .configure_axis(labelFontSize=12, titleFontSize=12, gridColor="#e5e7eb")
        .configure_title(fontSize=14, anchor="start", color="#111827")
        .configure_legend(labelFontSize=12, symbolSize=140)
    )

    st.altair_chart(chart, use_container_width=True)
