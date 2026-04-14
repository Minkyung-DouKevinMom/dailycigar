import streamlit as st
import pandas as pd
import os
from io import BytesIO
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

from db import (
    get_all_import_batch,
    get_import_batch_one,
    get_import_item_list_filtered,
    get_import_item_detail,
    delete_import_item,
    get_export_price_product_names,
    get_export_price_sizes_by_product,
    get_export_price_package_options,
    get_product_mst_one,
    get_latest_tax_rule_for_import_calc,
    get_import_batch_tax_rule,
    upsert_import_item_full,
)

USD_TO_KRW_DEFAULT = 1500.0
PHP_TO_KRW_DEFAULT = 26.0


def _n(v, default=0.0):
    try:
        if v is None or v == "":
            return float(default)
        return float(v)
    except Exception:
        return float(default)


def _i(v, default=0):
    try:
        if v is None or v == "":
            return int(default)
        return int(v)
    except Exception:
        return int(default)


def _none_if_blank_text(v):
    s = str(v).strip()
    return s if s else None


def _none_if_zero_num(v):
    try:
        if float(v) == 0:
            return None
        return float(v)
    except Exception:
        return None


def _calc_values(
    export_price_usd: float,
    discount_percent: float,
    import_unit_qty: int,
    package_qty: int,
    unit_weight_g: float,
    usd_to_krw_rate: float,
    tax_rule: dict,
    local_box_price_php: float = 0.0,
    local_unit_price_php: float = 0.0,
    php_to_krw_rate: float = 0.0,
    retail_price_krw: float = 0.0,
    supply_price_krw: float = 0.0,
    store_retail_price_krw=None,
):
    box_qty = package_qty if package_qty and package_qty > 0 else 1

    # 화면 입력값: 20 => 20% 할인, DB 저장값: 0.8
    discount_factor = 1 - (discount_percent / 100.0)
    if discount_factor < 0:
        discount_factor = 0.0

    # 박스/개당 가격
    discounted_box_price_usd = export_price_usd * discount_factor
    export_unit_price_usd = discounted_box_price_usd / box_qty if box_qty > 0 else 0.0

    # 개당 / 합계 수입원가
    import_unit_cost_krw = export_unit_price_usd * usd_to_krw_rate
    import_total_cost_krw = import_unit_cost_krw * import_unit_qty

    # 현지가 계산
    local_unit_price_krw = local_unit_price_php * php_to_krw_rate * 1.12

    # 무게: total만 합계
    total_weight_g = unit_weight_g * import_unit_qty

    # 세금: 전부 개당 기준
    individual_tax_krw = unit_weight_g * _n(tax_rule.get("individual_tax_per_g"))
    tobacco_tax_krw = unit_weight_g * _n(tax_rule.get("tobacco_tax_per_g"))
    local_education_tax_krw = tobacco_tax_krw * _n(tax_rule.get("local_education_rate"))
    health_charge_krw = unit_weight_g * _n(tax_rule.get("health_charge_per_g"))
    import_vat_krw = (import_unit_cost_krw + individual_tax_krw) * _n(tax_rule.get("import_vat_rate"))

    # 개당 세금합계 / 합계 세금합계
    tax_total_krw = (
        individual_tax_krw
        + tobacco_tax_krw
        + local_education_tax_krw
        + health_charge_krw
        + import_vat_krw
    )
    tax_total_all_krw = tax_total_krw * import_unit_qty

    # 한국원가: 개당
    korea_cost_krw = import_unit_cost_krw + tax_total_krw

    # 공급가 관련
    supply_vat_krw = supply_price_krw * 0.1
    supply_total_krw = supply_price_krw + supply_vat_krw

    # 마진율: 개당 한국원가 기준
    retail_margin_rate = ((retail_price_krw - korea_cost_krw) / retail_price_krw) if retail_price_krw else 0.0
    wholesale_margin_rate = ((supply_price_krw - korea_cost_krw) / supply_price_krw) if supply_price_krw else 0.0

    # 매장 소매가는 기본적으로 소비자가와 동일하게 시작
    if store_retail_price_krw in (None, ""):
        store_retail_price_krw = retail_price_krw

    return {
        "discount_factor": discount_factor,
        "discounted_box_price_usd": discounted_box_price_usd,
        "export_unit_price_usd": export_unit_price_usd,
        "import_unit_cost_krw": import_unit_cost_krw,
        "import_total_cost_krw": import_total_cost_krw,
        "local_box_price_php": local_box_price_php,
        "local_unit_price_php": local_unit_price_php,
        "local_unit_price_krw": local_unit_price_krw,
        "total_weight_g": total_weight_g,
        "individual_tax_krw": individual_tax_krw,
        "tobacco_tax_krw": tobacco_tax_krw,
        "local_education_tax_krw": local_education_tax_krw,
        "health_charge_krw": health_charge_krw,
        "import_vat_krw": import_vat_krw,
        "tax_total_krw": tax_total_krw,
        "tax_total_all_krw": tax_total_all_krw,
        "korea_cost_krw": korea_cost_krw,
        "supply_vat_krw": supply_vat_krw,
        "supply_total_krw": supply_total_krw,
        "retail_margin_rate": retail_margin_rate,
        "wholesale_margin_rate": wholesale_margin_rate,
        "store_retail_price_krw": store_retail_price_krw,
    }

def _excel_num(v, default=0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except Exception:
        return default


def _excel_int(v, default=0):
    try:
        if v is None or v == "":
            return default
        return int(float(v))
    except Exception:
        return default


def _find_header_row(ws, max_scan_row=20):
    """
    템플릿에서 데이터 헤더 행 자동 탐색
    """
    target_candidates = {"상품코드", "상품명", "사이즈", "수입개수"}
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        values = {
            str(ws.cell(r, c).value).strip()
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        }
        if len(target_candidates & values) >= 2:
            return r
    return 2


def _build_header_map(ws, header_row):
    header_map = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None:
            header_map[str(v).strip()] = c
    return header_map


def _set_value_next_to_label(ws, label_text, value, max_scan_row=10, max_scan_col=10):
    """
    예: '버전명' 셀을 찾고 오른쪽 칸에 값 입력
    """
    for r in range(1, min(ws.max_row, max_scan_row) + 1):
        for c in range(1, min(ws.max_column, max_scan_col) + 1):
            cell_val = ws.cell(r, c).value
            if cell_val is not None and str(cell_val).strip() == label_text:
                ws.cell(r, c + 1).value = value
                return True
    return False


def _collect_import_price_rows(df: pd.DataFrame):
    """
    목록 df의 id 기준으로 상세 재조회 후
    수입가격분석_template.xlsx 전용 데이터 구성
    """
    rows = []
    if df is None or df.empty:
        return rows

    for _, row in df.iterrows():
        item_id = row.get("id")
        if pd.isna(item_id):
            continue

        detail_df = get_import_item_detail(int(item_id))
        if detail_df is None or detail_df.empty:
            continue

        d = detail_df.iloc[0].to_dict()

        qty = _excel_int(d.get("import_unit_qty"), 0)
        proposal_retail = _excel_num(d.get("proposal_retail_price_krw"), 0)
        retail_price = _excel_num(d.get("retail_price_krw"), 0)
        store_retail = _excel_num(d.get("store_retail_price_krw"), 0)
        supply_price = _excel_num(d.get("supply_price_krw"), 0)
        margin = _excel_num(d.get("margin_krw"), 0)

        item = {
            "원본행번호": _excel_int(d.get("source_row_no"), 0),

            "제품명": d.get("product_name", ""),
            "사이즈": d.get("size_name", ""),
            "박스수출원가($)": _excel_num(d.get("export_box_price_usd"), 0),
            "할인가($)": _excel_num(d.get("discounted_box_price_usd"), 0),
            "수입개수": qty,
            "수출가격($/unit)": _excel_num(d.get("export_unit_price_usd"), 0),
            "수입원가(KRW)": _excel_num(d.get("import_unit_cost_krw"), 0),
            "수입가격합계": _excel_num(d.get("import_total_cost_krw"), 0),

            "무게": _excel_num(d.get("unit_weight_g"), 0),
            "전체무게": _excel_num(d.get("total_weight_g"), 0),

            "개별소비세": _excel_num(d.get("individual_tax_krw"), 0),
            "담배소비세": _excel_num(d.get("tobacco_tax_krw"), 0),
            "지방교육세": _excel_num(d.get("local_education_tax_krw"), 0),
            "국민건강": _excel_num(d.get("health_charge_krw"), 0),
            "부가세": _excel_num(d.get("import_vat_krw"), 0),

            "세금합계": _excel_num(d.get("tax_total_krw"), 0),
            "세금합계(총합계)": _excel_num(d.get("tax_total_all_krw"), 0),
            "한국 원가(KRW)": _excel_num(d.get("korea_cost_krw"), 0),

            "Box(페소)": _excel_num(d.get("local_box_price_php"), 0),
            "Unit(페소)": _excel_num(d.get("local_unit_price_php"), 0),
            "Unit(KRW)": _excel_num(d.get("local_unit_price_krw"), 0),

            "제안소비자가": proposal_retail,
            "소비자가격": retail_price,
            "매장운영가격": store_retail,

            "공급가": supply_price,
            "공급가(VAT)": round(supply_price * 1.1, 2) if supply_price else 0,
            "공급가(합계)": round(supply_price * qty, 2) if supply_price and qty else 0,

            "마진": margin,
            "소매이익률": round((margin / proposal_retail) * 100, 2) if proposal_retail else 0,
            "도매이익률": round((margin / supply_price) * 100, 2) if supply_price else 0,
        }
        rows.append(item)

    rows.sort(key=lambda x: (x.get("원본행번호", 0), x.get("제품명", ""), x.get("사이즈", "")))
    return rows


def _make_import_price_analysis_excel(batch_row: dict, export_df: pd.DataFrame) -> bytes:
    template_path = Path("templates") / "수입가격분석_template.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {template_path}")

    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    export_rows = _collect_import_price_rows(export_df)

    # 상단 기본 정보 자동 입력
    _set_value_next_to_label(ws, "버전명", batch_row.get("version_name", ""))
    _set_value_next_to_label(ws, "USD 환율", _n(batch_row.get("usd_to_krw_rate"), USD_TO_KRW_DEFAULT))
    _set_value_next_to_label(ws, "PHP 환율", _n(batch_row.get("php_to_krw_rate"), PHP_TO_KRW_DEFAULT))
    _set_value_next_to_label(ws, "수입일", batch_row.get("import_date", ""))

    # 현재 템플릿 기준: 2행 헤더, 3행부터 데이터
    start_row = 3

    # 기존 데이터 삭제 (3행부터 AD열 정도까지)
    for r in range(start_row, max(ws.max_row, start_row) + 1):
        for c in range(1, 31):  # A~AD
            ws.cell(r, c).value = None

    for row_idx, item in enumerate(export_rows, start=start_row):
        ws[f"A{row_idx}"] = item.get("제품명", "")
        ws[f"B{row_idx}"] = item.get("사이즈", "")
        ws[f"C{row_idx}"] = item.get("박스수출원가($)", 0)
        ws[f"D{row_idx}"] = item.get("할인가($)", 0)
        ws[f"E{row_idx}"] = item.get("수입개수", 0)
        ws[f"F{row_idx}"] = item.get("수출가격($/unit)", 0)
        ws[f"G{row_idx}"] = item.get("수입원가(KRW)", 0)
        ws[f"H{row_idx}"] = item.get("수입가격합계", 0)
        ws[f"I{row_idx}"] = item.get("무게", 0)
        ws[f"J{row_idx}"] = item.get("전체무게", 0)
        ws[f"K{row_idx}"] = item.get("개별소비세", 0)
        ws[f"L{row_idx}"] = item.get("담배소비세", 0)
        ws[f"M{row_idx}"] = item.get("지방교육세", 0)
        ws[f"N{row_idx}"] = item.get("국민건강", 0)
        ws[f"O{row_idx}"] = item.get("부가세", 0)
        ws[f"P{row_idx}"] = item.get("세금합계", 0)
        ws[f"Q{row_idx}"] = item.get("세금합계(총합계)", 0)
        ws[f"R{row_idx}"] = item.get("한국 원가(KRW)", 0)
        ws[f"S{row_idx}"] = item.get("Box(페소)", 0)
        ws[f"T{row_idx}"] = item.get("Unit(페소)", 0)
        ws[f"U{row_idx}"] = item.get("Unit(KRW)", 0)
        ws[f"V{row_idx}"] = item.get("제안소비자가", 0)
        ws[f"W{row_idx}"] = item.get("소비자가격", 0)
        ws[f"X{row_idx}"] = item.get("매장운영가격", 0)
        ws[f"Y{row_idx}"] = item.get("공급가", 0)
        ws[f"Z{row_idx}"] = item.get("공급가(VAT)", 0)
        ws[f"AA{row_idx}"] = item.get("공급가(합계)", 0)
        ws[f"AB{row_idx}"] = item.get("마진", 0)
        ws[f"AC{row_idx}"] = item.get("소매이익률", 0)
        ws[f"AD{row_idx}"] = item.get("도매이익률", 0)

        # 데이터 영역 스타일 적용 (타이틀 제외)
    thin = Side(style="thin", color="000000")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    last_row = start_row + len(export_rows) - 1
    if last_row >= start_row:
        for r in range(start_row, last_row + 1):
            for c in range(1, 31):  # A~AD
                cell = ws.cell(r, c)
                cell.font = Font(
                    name=cell.font.name if cell.font and cell.font.name else "Calibri",
                    size=10,
                    bold=cell.font.bold if cell.font else False,
                    italic=cell.font.italic if cell.font else False,
                    color=cell.font.color if cell.font and cell.font.color else None,
                )
                cell.border = data_border
                
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def render():
    st.subheader("수입제품 관리")

    batch_df = get_all_import_batch()
    if batch_df.empty:
        st.info("수입 버전 데이터가 없습니다.")
        return

    batch_options = {"선택하세요": None}
    for _, row in batch_df.iterrows():
        batch_options[f'{row["id"]} | {row["version_name"]}'] = row["id"]

    c1, c2 = st.columns([2, 3])
    with c1:
        selected_label = st.selectbox("수입 버전 선택", list(batch_options.keys()))
        selected_batch_id = batch_options[selected_label]
    with c2:
        keyword = st.text_input("검색", placeholder="상품명 / 사이즈 / 코드")

    if selected_batch_id is None:
        st.info("먼저 수입 버전을 선택해 주세요.")
        return

    batch_row = get_import_batch_one(selected_batch_id)
    tax_rule = get_import_batch_tax_rule(selected_batch_id)

    info1, info2, info3, info4 = st.columns(4)

    info1.metric("USD 환율", f'{_n(batch_row.get("usd_to_krw_rate"), USD_TO_KRW_DEFAULT):,.2f}')
    info2.metric("PHP 환율", f'{_n(batch_row.get("php_to_krw_rate"), PHP_TO_KRW_DEFAULT):,.2f}')
    info3.metric("세금 규칙", tax_rule.get("rule_name", "최신 규칙"))

    df = get_import_item_list_filtered(selected_batch_id, keyword)

    # 다운로드는 검색조건과 무관하게 선택한 수입버전 전체 기준
    export_df = get_import_item_list_filtered(selected_batch_id, "")

    with info4:
        st.write("")
        st.write("")
        if export_df is None or export_df.empty:
            st.button("수입가격 다운로드", disabled=True, use_container_width=True)
        else:
            try:
                excel_bytes = _make_import_price_analysis_excel(batch_row, export_df)
                version_name = str(batch_row.get("version_name", f"batch_{selected_batch_id}"))
                st.download_button(
                    label="수입가격 다운로드",
                    data=excel_bytes,
                    file_name=f"수입가격분석_{version_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"엑셀 생성 오류: {e}")

    st.markdown("### 조회 결과")
    
    if not df.empty:
        show_df = df.rename(columns={
            "id": "ID",
            "batch_id": "버전ID",
            "product_name": "상품명",
            "size_name": "사이즈",
            "product_code": "상품코드",
            "import_unit_qty": "수입개수",
            "import_total_cost_krw": "총수입금액",
            "total_weight_g": "총무게(g)",
            "retail_price_krw": "소비자가",
            "proposal_retail_price_krw": "제안소비자가",
            "store_retail_price_krw": "매장운영가",
            "supply_price_krw": "공급가",
            "margin_krw": "마진",
            "source_row_no": "원본행번호",
            "notes": "코멘트",
        }).copy()

        for col in ["총수입금액", "제안소비자가", "소비자가", "매장운영가", "공급가", "마진"]:
            if col in show_df.columns:
                show_df[col] = show_df[col].apply(
                    lambda x: f"₩{x:,.0f}" if pd.notna(x) else ""
                )

        preferred_order = [
            "ID",
            "버전ID",
            "상품코드",
            "상품명",
            "사이즈",
            "제안소비자가",
            "소비자가",
            "매장운영가",
            "공급가",
        ]

        remaining_cols = [col for col in show_df.columns if col not in preferred_order]
        show_df = show_df[preferred_order + remaining_cols]

        st.dataframe(show_df, use_container_width=True, hide_index=True, height=340)
        st.caption(f"{len(show_df):,}건 조회되었습니다.")
    else:
        st.info("조회 결과가 없습니다.")

    st.divider()
    render_editor(df, selected_batch_id, batch_row, tax_rule)


def render_editor(df: pd.DataFrame, selected_batch_id: int, batch_row: dict, tax_rule: dict):
    edit_mode = st.radio("작업 구분", ["신규 추가", "기존 수정"], horizontal=True)

    detail_row = {}
    selected_item_id = None

    if edit_mode == "기존 수정":
        if df.empty:
            st.warning("수정할 수입제품이 없습니다.")
            return

        item_options = {
            f'{row["id"]} | {row["product_name"]} | {row["size_name"]} | row:{row["source_row_no"]}': row["id"]
            for _, row in df.iterrows()
        }
        selected_item_label = st.selectbox("수정할 수입제품 선택", list(item_options.keys()))
        selected_item_id = item_options[selected_item_label]

        detail_df = get_import_item_detail(selected_item_id)
        if detail_df.empty:
            st.error("상세 데이터를 찾을 수 없습니다.")
            return
        detail_row = detail_df.iloc[0].to_dict()

    product_names = get_export_price_product_names() or []

    st.markdown("### 수입제품 입력")
    
    manual_default = False
    if detail_row and detail_row.get("product_name") not in product_names:
        manual_default = True

    manual_input = st.checkbox(
        "목록에 없는 상품/사이즈 직접 입력",
        value=manual_default,
        key=f"manual_input_mode_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
    )

    # ---- 의존 선택값은 form 밖에서 즉시 반영되도록 처리 ----
    export_box_price_usd = 0.0
    package_qty = 1
    export_package_type = ""
    export_package_qty = 1
    product_info = {}
    product_name = ""
    size_name = ""
    product_code = ""

    selection_col1, selection_col2, selection_col3 = st.columns(3)

    with selection_col1:
        if manual_input:
            product_name = st.text_input(
                "상품명",
                value=str(detail_row.get("product_name") or ""),
                key=f"product_name_manual_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
            )

            size_name = st.text_input(
                "사이즈",
                value=str(detail_row.get("size_name") or ""),
                key=f"size_name_manual_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
            )

            package_qty = st.number_input(
                "포장 수량(개)",
                min_value=1,
                value=max(_i(detail_row.get("export_package_qty"), 1), 1),
                step=1,
                key=f"package_qty_manual_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
            )
            export_package_type = str(detail_row.get("export_package_type") or "MANUAL").strip() or "MANUAL"
            export_package_qty = package_qty

            export_box_price_usd = st.number_input(
                "본사 수출가격(USD, 박스기준)",
                min_value=0.0,
                value=_n(detail_row.get("export_box_price_usd"), 0.0),
                step=0.01,
                key=f"export_usd_manual_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
            )
        else:
            if not product_names:
                st.warning("본사수출배포가격 데이터가 없습니다. 직접 입력을 사용해 주세요.")
                product_name = st.text_input(
                    "상품명",
                    value=str(detail_row.get("product_name") or ""),
                    key=f"product_name_fallback_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
                )
                size_name = st.text_input(
                    "사이즈",
                    value=str(detail_row.get("size_name") or ""),
                    key=f"size_name_fallback_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
                )
                package_qty = st.number_input(
                    "포장 수량(개)",
                    min_value=1,
                    value=max(_i(detail_row.get("export_package_qty"), 1), 1),
                    step=1,
                    key=f"package_qty_fallback_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
                )
                export_package_type = str(detail_row.get("export_package_type") or "MANUAL").strip() or "MANUAL"
                export_package_qty = package_qty
                export_box_price_usd = st.number_input(
                    "본사 수출가격(USD, 박스기준)",
                    min_value=0.0,
                    value=_n(detail_row.get("export_box_price_usd"), 0.0),
                    step=0.01,
                    key=f"export_usd_fallback_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
                )
            else:
                default_product = detail_row.get("product_name") if detail_row else product_names[0]
                if default_product not in product_names:
                    default_product = product_names[0]

                product_name = st.selectbox(
                    "상품명",
                    product_names,
                    index=product_names.index(default_product),
                    key=f"product_select_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}",
                )

                size_options = get_export_price_sizes_by_product(product_name) or []
                default_size = detail_row.get("size_name") if detail_row else (size_options[0] if size_options else "")

                if size_options:
                    if default_size not in size_options:
                        default_size = size_options[0]

                    size_name = st.selectbox(
                        "사이즈",
                        size_options,
                        index=size_options.index(default_size) if default_size in size_options else 0,
                        key=f"size_select_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}",
                    )
                else:
                    st.warning("해당 상품의 사이즈 목록이 없습니다. 직접 입력을 사용해 주세요.")
                    size_name = st.text_input(
                        "사이즈",
                        value=str(detail_row.get("size_name") or ""),
                        key=f"size_text_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}",
                    )

                package_df = get_export_price_package_options(product_name, size_name)
                package_labels = []
                package_map = {}

                if package_df is not None and not package_df.empty:
                    for _, r in package_df.iterrows():
                        label = f'{r["package_type"]} / {int(r["package_qty"])}개 / USD {float(r["export_price_usd"]):,.2f}'
                        package_labels.append(label)
                        package_map[label] = r.to_dict()

                if package_labels:
                    default_package_label = package_labels[0]

                    widget_suffix = f"{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}_{size_name}"
                    package_key = f"package_select_{widget_suffix}"
                    edit_token = f"{selected_batch_id}_{selected_item_id}_{edit_mode}_{product_name}_{size_name}"

                    if st.session_state.get("last_import_package_token") != edit_token:
                        st.session_state["last_import_package_token"] = edit_token
                        st.session_state.pop(package_key, None)

                    if detail_row:
                        saved_pkg_type = str(detail_row.get("export_package_type") or "").strip()
                        saved_pkg_qty = _i(detail_row.get("export_package_qty"), 0)
                        saved_export_usd = _n(detail_row.get("export_box_price_usd"), 0.0)

                        matched_label = None

                        for label, pkg in package_map.items():
                            pkg_type = str(pkg.get("package_type") or "").strip()
                            pkg_qty = _i(pkg.get("package_qty"), 0)

                            if pkg_type == saved_pkg_type and pkg_qty == saved_pkg_qty:
                                matched_label = label
                                break

                        if matched_label is None:
                            for label, pkg in package_map.items():
                                pkg_qty = _i(pkg.get("package_qty"), 0)
                                pkg_export_usd = _n(pkg.get("export_price_usd"), 0.0)

                                if pkg_qty == saved_pkg_qty and abs(pkg_export_usd - saved_export_usd) < 0.001:
                                    matched_label = label
                                    break

                        if matched_label is not None:
                            default_package_label = matched_label
                    
                    package_label = st.selectbox(
                        "수출 포장유형 / 개수",
                        package_labels,
                        index=package_labels.index(default_package_label),
                        key=package_key,
                    )
                    pkg = package_map[package_label]
                    export_package_type = str(pkg.get("package_type") or "").strip()
                    export_package_qty = _i(pkg.get("package_qty"), 1)
                    export_box_price_usd = _n(pkg.get("export_price_usd"))
                    package_qty = export_package_qty
                else:
                    st.warning("해당 상품/사이즈의 포장유형 데이터가 없습니다. 직접 입력을 사용하거나 수출가격을 수동 입력하세요.")
                    package_qty = st.number_input(
                        "포장 수량(개)",
                        min_value=1,
                        value=max(_i(detail_row.get("export_package_qty"), 1), 1),
                        step=1,
                        key=f"package_qty_no_pkg_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}_{size_name}",
                    )
                    export_package_type = str(detail_row.get("export_package_type") or "MANUAL").strip() or "MANUAL"
                    export_package_qty = package_qty
                    export_box_price_usd = st.number_input(
                        "본사 수출가격(USD, 박스기준)",
                        min_value=0.0,
                        value=_n(detail_row.get("export_box_price_usd"), 0.0),
                        step=0.01,
                        key=f"export_usd_no_pkg_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}_{size_name}",
                    )

                product_info = get_product_mst_one(product_name, size_name) or {}

    with selection_col2:
        default_product_code = (
            detail_row.get("product_code")
            or product_info.get("product_code")
            or ""
        )

        product_code = st.text_input(
            "상품코드",
            value=str(default_product_code),
            disabled=False if manual_input else True,
            key=f"product_code_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}_{product_name}_{size_name}",
        )

        st.metric("자동조회 본사 수출가격(USD)", f"{export_box_price_usd:,.2f}")
        st.metric("자동조회 포장 수량", f"{package_qty:,}개")

    with selection_col3:
        product_weight = _n(product_info.get("unit_weight_g"), 0.0)
        st.metric("상품마스터 개당 무게(g)", f"{product_weight:,.1f}")

    st.divider()

    with st.form(f"import_item_form_{selected_batch_id}_{edit_mode}_{selected_item_id or 'new'}", clear_on_submit=False):
        col1, col2, col3 = st.columns(3)

        with col1:
            import_unit_qty = st.number_input(
                "총 수입 개수",
                min_value=0,
                value=max(_i(detail_row.get("import_unit_qty"), package_qty), 0),
                step=1,
            )

            saved_discount_factor = _n(detail_row.get("discount_rate"), 1.0)
            default_discount_percent = (1 - saved_discount_factor) * 100 if detail_row else 0.0

            discount_rate = st.number_input(
                "할인율(%)",
                min_value=0.0,
                max_value=100.0,
                value=default_discount_percent,
                step=0.1,
            )

            discounted_box_price_usd_manual = st.number_input(
                "최종 수출가격(USD, 박스기준)",
                min_value=0.0,
                value=export_box_price_usd * (1 - discount_rate / 100.0),
                step=0.01,
                disabled=True,
            )

            source_row_no = st.number_input(
                "원본 행번호",
                min_value=0,
                value=_i(detail_row.get("source_row_no"), 0),
                step=1,
            )

            item_notes = st.text_area(
                "특징/코멘트",
                value=str(detail_row.get("notes") or ""),
                height=100,
                placeholder="예: 시트러스 향이 강함 / 래퍼 상태 좋음 / 추천 포인트 등",
            )

        with col2:
            usd_to_krw_rate = _n(batch_row.get("usd_to_krw_rate"), USD_TO_KRW_DEFAULT)
            php_to_krw_rate = _n(batch_row.get("php_to_krw_rate"), PHP_TO_KRW_DEFAULT)

            st.number_input("달러환율", value=usd_to_krw_rate, disabled=True)
            st.number_input("페소환율", value=php_to_krw_rate, disabled=True)

            local_box_price_php = st.number_input(
                "현지 박스가격(PHP)",
                min_value=0.0,
                value=_n(detail_row.get("local_box_price_php"), 0.0),
                step=0.01,
            )

            local_unit_price_php = st.number_input(
                "현지 유닛가격(PHP)",
                min_value=0.0,
                value=_n(detail_row.get("local_unit_price_php"), 0.0),
                step=0.01,
            )

            local_unit_price_krw_preview = local_unit_price_php * php_to_krw_rate
            st.number_input(
                "현지 유닛가격(원화)",
                min_value=0.0,
                value=float(local_unit_price_krw_preview),
                step=1.0,
                disabled=True,
            )

        with col3:
            default_unit_weight = _n(
                detail_row.get("unit_weight_g"),
                _n(product_info.get("unit_weight_g"), 0.0)
            )
            unit_weight_g = st.number_input(
                "개당 무게(g)",
                min_value=0.0,
                value=default_unit_weight,
                step=0.1,
            )

            retail_price_krw = st.number_input(
                "소비자가",
                min_value=0,
                value=int(round(_n(detail_row.get("retail_price_krw"), 0.0))),
                step=100,
                format="%d",
            )

            proposal_retail_default = _n(
                detail_row.get("proposal_retail_price_krw"),
                _n(detail_row.get("retail_price_krw"), 0.0)
            )
            if proposal_retail_default == 0 and retail_price_krw > 0:
                proposal_retail_default = retail_price_krw

            supply_price_krw = st.number_input(
                "공급가",
                min_value=0,
                value=int(round(_n(detail_row.get("supply_price_krw"), 0.0))),
                step=100,
                format="%d",
            )

            store_retail_default = _n(
                detail_row.get("store_retail_price_krw"),
                _n(detail_row.get("retail_price_krw"), 0.0)
            )
            if store_retail_default == 0 and retail_price_krw > 0:
                store_retail_default = retail_price_krw

            store_retail_price_krw = st.number_input(
                "매장 소매가",
                min_value=0,
                value=int(round(store_retail_default)),
                step=100,
                format="%d",
            )

            proposal_retail_price_krw = st.number_input(
                "제안소비자가",
                min_value=0,
                value=int(round(proposal_retail_default)),
                step=100,
                format="%d",
            )

        calc = _calc_values(
            export_price_usd=export_box_price_usd,
            discount_percent=discount_rate,
            import_unit_qty=import_unit_qty,
            package_qty=package_qty,
            unit_weight_g=unit_weight_g,
            usd_to_krw_rate=usd_to_krw_rate,
            tax_rule=tax_rule,
            local_box_price_php=local_box_price_php,
            local_unit_price_php=local_unit_price_php,
            php_to_krw_rate=php_to_krw_rate,
            retail_price_krw=retail_price_krw,
            supply_price_krw=supply_price_krw,
            store_retail_price_krw=store_retail_price_krw,
        )

        margin_krw = retail_price_krw - supply_price_krw if (retail_price_krw or supply_price_krw) else 0.0

        st.info(
            f"""
세금 계산식
- 개별소비세(개당) = 개당무게 × 개별소비세비율 = {calc['individual_tax_krw']:,.0f}원
- 담배소비세(개당) = 개당무게 × 담배소비세비율 = {calc['tobacco_tax_krw']:,.0f}원
- 지방교육세(개당) = 담배소비세(개당) × 지방교육세비율 = {calc['local_education_tax_krw']:,.0f}원
- 국민건강증진금(개당) = 개당무게 × 국민건강비율 = {calc['health_charge_krw']:,.0f}원
- 수입부가세(개당) = (개당수입원가 + 개별소비세) × 부가세율 = {calc['import_vat_krw']:,.0f}원
- 세금합계(개당) = {calc['tax_total_krw']:,.0f}원
- 세금합계(총) = 개당세금합계 × 수입개수 = {calc['tax_total_all_krw']:,.0f}원
- 한국원가(개당) = 개당수입원가 + 개당세금합계 = {calc['korea_cost_krw']:,.0f}원
            """
        )

        st.markdown("### 자동 계산 결과")
        r1, r2, r3, r4 = st.columns(4)
        r1.metric("본사 박스가격(USD)", f"{export_box_price_usd:,.2f}")
        r2.metric("할인적용 박스가격(USD)", f"{calc['discounted_box_price_usd']:,.2f}")
        r3.metric("개당 수출단가(USD)", f"{calc['export_unit_price_usd']:,.2f}")
        r4.metric("개당 수입원가(KRW)", f"₩{calc['import_unit_cost_krw']:,.0f}")

        r5, r6, r7, r8 = st.columns(4)
        r5.metric("총 수입금액(KRW)", f"₩{calc['import_total_cost_krw']:,.0f}")
        r6.metric("총 무게(g)", f"{calc['total_weight_g']:,.1f}")
        r7.metric("공급가 VAT", f"₩{calc['supply_vat_krw']:,.0f}")
        r8.metric("공급가 합계", f"₩{calc['supply_total_krw']:,.0f}")

        r9, r10, r11, r12 = st.columns(4)
        r9.metric("세금합계(개당)", f"₩{calc['tax_total_krw']:,.0f}")
        r10.metric("한국원가(개당)", f"₩{calc['korea_cost_krw']:,.0f}")
        r11.metric("소비자가 마진율", f"{calc['retail_margin_rate']*100:,.1f}%")
        r12.metric("공급가 마진율", f"{calc['wholesale_margin_rate']*100:,.1f}%")

        r13, r14, r15, r16 = st.columns(4)
        r13.metric("개별소비세(개당)", f"₩{calc['individual_tax_krw']:,.0f}")
        r14.metric("담배소비세(개당)", f"₩{calc['tobacco_tax_krw']:,.0f}")
        r15.metric("지방교육세(개당)", f"₩{calc['local_education_tax_krw']:,.0f}")
        r16.metric("국민건강증진금(개당)", f"₩{calc['health_charge_krw']:,.0f}")

        r17, r18, r19 = st.columns(3)
        r17.metric("수입부가세(개당)", f"₩{calc['import_vat_krw']:,.0f}")
        r18.metric("매장 소매가", f"₩{calc['store_retail_price_krw']:,.0f}")
        r19.metric("마진", f"₩{margin_krw:,.0f}")

        raw_row_json = detail_row.get("raw_row_json") or ""
        raw_formula_json = detail_row.get("raw_formula_json") or ""

        c1, c2 = st.columns(2)
        save_btn = c1.form_submit_button("저장", use_container_width=True, type="primary")
        delete_btn = c2.form_submit_button("삭제", use_container_width=True)
        # 원본행번호 중복 체크
        dup_df = df.copy() if df is not None else pd.DataFrame()

        if save_btn:
            if not str(product_name).strip():
                st.warning("상품명을 입력해 주세요.")
                st.stop()

            if not str(size_name).strip():
                st.warning("사이즈를 입력해 주세요.")
                st.stop()

            if manual_input and not str(product_code).strip():
                st.warning("직접 입력 모드에서는 상품코드를 입력해 주세요.")
                st.stop()
            
            if not dup_df.empty and "source_row_no" in dup_df.columns:
                # 기존 수정이면 자기 자신은 제외
                if edit_mode == "기존 수정" and selected_item_id is not None and "id" in dup_df.columns:
                    dup_df = dup_df[dup_df["id"] != selected_item_id]

                dup_rows = dup_df[dup_df["source_row_no"] == source_row_no]

                if not dup_rows.empty:
                    dup_names = []
                    if "product_name" in dup_rows.columns and "size_name" in dup_rows.columns:
                        dup_names = [
                            f'{r["product_name"]} / {r["size_name"]}'
                            for _, r in dup_rows.iterrows()
                        ]

                    dup_msg = ", ".join(dup_names) if dup_names else "기존 품목"

                    st.error(
                        f"원본 행번호 {source_row_no} 는 이미 사용 중입니다. "
                        f"중복 품목: {dup_msg}. 다른 번호를 입력해 주세요."
                    )
                    st.stop()

            upsert_import_item_full(
                item_id=selected_item_id,
                batch_id=selected_batch_id,
                product_name=product_name,
                size_name=size_name,
                product_code=_none_if_blank_text(product_code),
                export_package_type=_none_if_blank_text(export_package_type),
                export_package_qty=None if _i(export_package_qty, 0) == 0 else _i(export_package_qty, 0),
                export_box_price_usd=_none_if_zero_num(export_box_price_usd),
                discounted_box_price_usd=_none_if_zero_num(calc["discounted_box_price_usd"]),
                discount_rate=_none_if_zero_num(calc["discount_factor"]),
                import_unit_qty=import_unit_qty,
                export_unit_price_usd=_none_if_zero_num(calc["export_unit_price_usd"]),
                import_unit_cost_krw=_none_if_zero_num(calc["import_unit_cost_krw"]),
                import_total_cost_krw=_none_if_zero_num(calc["import_total_cost_krw"]),
                unit_weight_g=_none_if_zero_num(unit_weight_g),
                total_weight_g=_none_if_zero_num(calc["total_weight_g"]),
                individual_tax_krw=_none_if_zero_num(calc["individual_tax_krw"]),
                tobacco_tax_krw=_none_if_zero_num(calc["tobacco_tax_krw"]),
                local_education_tax_krw=_none_if_zero_num(calc["local_education_tax_krw"]),
                health_charge_krw=_none_if_zero_num(calc["health_charge_krw"]),
                import_vat_krw=_none_if_zero_num(calc["import_vat_krw"]),
                tax_total_krw=_none_if_zero_num(calc["tax_total_krw"]),
                tax_total_all_krw=_none_if_zero_num(calc["tax_total_all_krw"]),
                korea_cost_krw=_none_if_zero_num(calc["korea_cost_krw"]),
                local_box_price_php=_none_if_zero_num(calc["local_box_price_php"]),
                local_unit_price_php=_none_if_zero_num(calc["local_unit_price_php"]),
                local_unit_price_krw=_none_if_zero_num(calc["local_unit_price_krw"]),
                retail_price_krw=_none_if_zero_num(retail_price_krw),
                proposal_retail_price_krw=_none_if_zero_num(proposal_retail_price_krw),
                supply_price_krw=_none_if_zero_num(supply_price_krw),
                supply_vat_krw=_none_if_zero_num(calc["supply_vat_krw"]),
                supply_total_krw=_none_if_zero_num(calc["supply_total_krw"]),
                retail_margin_rate=_none_if_zero_num(calc["retail_margin_rate"]),
                wholesale_margin_rate=_none_if_zero_num(calc["wholesale_margin_rate"]),
                store_retail_price_krw=_none_if_zero_num(calc["store_retail_price_krw"]),
                margin_krw=_none_if_zero_num(margin_krw),
                source_row_no=None if source_row_no == 0 else source_row_no,
                notes=_none_if_blank_text(item_notes),
                raw_row_json=raw_row_json,
                raw_formula_json=raw_formula_json,
            )
            st.success("저장되었습니다.")
            st.rerun()

        if delete_btn:
            if selected_item_id is None:
                st.warning("신규 추가 상태에서는 삭제할 항목이 없습니다.")
            else:
                delete_import_item(selected_item_id)
                st.success("삭제되었습니다.")
                st.rerun()