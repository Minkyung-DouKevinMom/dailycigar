import io
import os
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from db import (
    get_product_intro_export_data,
    get_all_partner_for_select,
    get_partner_detail_by_id,
    get_estimate_cigar_items,
    get_estimate_non_cigar_items,
)

TEMPLATE_PATH = "templates/견적서_template.xlsx"
TEMPLATE_PATH_A3 = "templates/견적서_template_A3.xlsx"
PRODUCT_INTRO_TEMPLATE_PATH = "templates/상품소개서_template.xlsx"

# -----------------------------
# 견적서 템플릿별 레이아웃 정의
# -----------------------------
# A4: 데이터 영역이 B~I열, 10~64행
# A3: 데이터 영역이 D~K열, 10~51행
ESTIMATE_LAYOUTS = {
    "A4": {
        "sheet_cigar": "견적서_기본",
        "sheet_non_cigar": "시가 외",
        # 헤더 셀
        "date_cell": "A3",
        "partner_name_cell_1": "G3",
        "partner_name_cell_2": "G4",
        "address_cell": "G5",
        "contact_cell": "G6",
        "phone_cell": "G7",
        "amount_text_cell": "A8",
        "grade_discount_cell": "H8",
        "label_cell": "A10",
        # 데이터 영역
        "data_start_row": 10,
        "data_end_row": 64,
        "total_row": 65,
        # 컬럼 (display_name은 B:D 병합되어 B에 입력)
        "col_display": "B",
        "col_qty": "E",
        "col_retail": "F",
        "col_supply": "G",
        "col_vat": "H",
        "col_total": "I",
        # 그룹 외곽 테두리에 사용할 컬럼 범위
        "border_columns": ["B", "C", "D", "E", "F", "G", "H", "I"],
    },
    "A3": {
        "sheet_cigar": "견적서_기본",
        "sheet_non_cigar": "견적서_시가 외",
        # 헤더 셀
        "date_cell": "C3",
        "partner_name_cell_1": "I3",
        "partner_name_cell_2": "I4",
        "address_cell": "I5",
        "contact_cell": "I6",
        "phone_cell": "I7",
        "amount_text_cell": "C8",
        "grade_discount_cell": "J8",
        "label_cell": "C10",
        # 데이터 영역
        "data_start_row": 10,
        "data_end_row": 51,
        "total_row": 52,
        # 컬럼 (display_name은 D:F 병합되어 D에 입력)
        "col_display": "D",
        "col_qty": "G",
        "col_retail": "H",
        "col_supply": "I",
        "col_vat": "J",
        "col_total": "K",
        # 그룹 외곽 테두리에 사용할 컬럼 범위
        "border_columns": ["D", "E", "F", "G", "H", "I", "J", "K"],
    },
}

# -----------------------------
# 상품소개서 기존 설정
# -----------------------------
HEADER_MAPPING = {
    "상품명": "product_name",
    "사이즈": "size_name",
    "Flavor": "flavor",
    "Strength": "strength",
    "Length": "length_text",
    "RG": "rg",
    "Time": "time_text",
    "Guide": "guide_text",
    "소비자가(KRW)": "proposal_retail_price_krw",
    "공급가(KRW)": "supply_price_krw",
    "공급가합계(KRW)": "supply_total_krw",
}

DISPLAY_RENAME = {
    "product_name": "상품명",
    "size_name": "사이즈",
    "flavor": "Flavor",
    "strength": "Strength",
    "length_text": "Length",
    "rg": "RG",
    "time_text": "Time",
    "guide_text": "Guide",
    "proposal_retail_price_krw": "소비자가(KRW)",
    "supply_price_krw": "공급가(KRW)",
    "supply_total_krw": "공급가합계(KRW)",
}


def _safe_value(value):
    if pd.isna(value):
        return ""
    return value


# =========================================================
# 상품소개서 기존 기능
# =========================================================
def _load_product_intro_template_workbook():
    if not os.path.exists(PRODUCT_INTRO_TEMPLATE_PATH):
        raise FileNotFoundError(
            f"템플릿 파일이 없습니다: {PRODUCT_INTRO_TEMPLATE_PATH}\n"
            "프로젝트 폴더의 templates 아래에 파일을 넣어주세요."
        )
    return load_workbook(PRODUCT_INTRO_TEMPLATE_PATH)


def _get_header_map(ws, header_row=1):
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if value is None:
            continue
        header_map[str(value).strip()] = col_idx
    return header_map


def _merge_cells(ws, col_idx, start_row, end_row):
    if col_idx and start_row < end_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=col_idx,
            end_row=end_row,
            end_column=col_idx,
        )


def _get_product_groups(ws, header_map, start_row, end_row):
    product_col = header_map.get("상품명")
    if not product_col or end_row < start_row:
        return []

    groups = []
    group_start = start_row
    prev_value = ws.cell(row=start_row, column=product_col).value

    for row_idx in range(start_row + 1, end_row + 1):
        current_value = ws.cell(row=row_idx, column=product_col).value
        if current_value != prev_value:
            groups.append((group_start, row_idx - 1))
            group_start = row_idx
            prev_value = current_value

    groups.append((group_start, end_row))
    return groups


def _merge_within_group_same_value(ws, header_map, header_name, group_start, group_end):
    col_idx = header_map.get(header_name)
    if not col_idx or group_end < group_start:
        return

    merge_start = group_start
    prev_value = ws.cell(row=group_start, column=col_idx).value

    for row_idx in range(group_start + 1, group_end + 1):
        current_value = ws.cell(row=row_idx, column=col_idx).value
        if current_value != prev_value:
            if prev_value not in (None, ""):
                _merge_cells(ws, col_idx, merge_start, row_idx - 1)
            merge_start = row_idx
            prev_value = current_value

    if prev_value not in (None, ""):
        _merge_cells(ws, col_idx, merge_start, group_end)


def _apply_alignment(ws, start_row, end_row, header_map):
    guide_col = header_map.get("Guide")

    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        for cell in row:
            if guide_col and cell.column == guide_col:
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="left",
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center",
                    wrap_text=False,
                )


def _apply_body_style(ws, header_map, start_row, end_row):
    thin = Side(style="thin", color="000000")
    all_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    light_gray_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")

    product_col = header_map.get("상품명")

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=cell.font.name, size=7)
            cell.border = all_border

            if col_idx == product_col and cell.value not in (None, ""):
                cell.fill = light_gray_fill


def build_product_intro_from_template(
    df: pd.DataFrame,
    include_retail_price: bool = True,
    include_supply_price: bool = True,
) -> io.BytesIO:
    wb = _load_product_intro_template_workbook()
    ws = wb[wb.sheetnames[0]]

    header_row = 1
    start_row = 2

    removable_headers = []
    if not include_retail_price:
        removable_headers.append("소비자가(KRW)")
    if not include_supply_price:
        removable_headers.extend(["공급가(KRW)", "공급가합계(KRW)"])

    if removable_headers:
        current_header_map = _get_header_map(ws, header_row=header_row)
        delete_cols = []

        for header in removable_headers:
            col_idx = current_header_map.get(header)
            if col_idx:
                delete_cols.append(col_idx)

        for col_idx in sorted(delete_cols, reverse=True):
            ws.delete_cols(col_idx, 1)

    header_map = _get_header_map(ws, header_row=header_row)

    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

    price_headers = {"소비자가(KRW)", "공급가(KRW)", "공급가합계(KRW)"}

    records = df.to_dict("records")
    for row_idx, record in enumerate(records, start=start_row):
        for excel_header, df_col in HEADER_MAPPING.items():
            col_idx = header_map.get(excel_header)
            if not col_idx:
                continue

            value = _safe_value(record.get(df_col, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if excel_header in price_headers and value not in ("", None):
                try:
                    cell.value = float(value)
                    cell.number_format = '₩#,##0'
                except Exception:
                    pass

    end_row = start_row + len(records) - 1
    if records:
        product_groups = _get_product_groups(ws, header_map, start_row, end_row)
        product_col = header_map.get("상품명")

        for group_start, group_end in product_groups:
            _merge_cells(ws, product_col, group_start, group_end)
            _merge_within_group_same_value(ws, header_map, "Flavor", group_start, group_end)
            _merge_within_group_same_value(ws, header_map, "Guide", group_start, group_end)

        _apply_alignment(ws, start_row, end_row, header_map)
        _apply_body_style(ws, header_map, start_row, end_row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def render_product_intro_export():
    st.subheader("상품소개서 엑셀 다운로드")
    st.caption("템플릿 파일을 그대로 사용하고, 상품 데이터를 채워 넣어 다운로드합니다.")
    
    c1, c2, c3, c4 = st.columns([1.2, 1, 1, 1])

    with c1:
        keyword = st.text_input("상품명 검색", "")

    with c2:
        use_yn = st.selectbox("사용여부", ["전체", "Y", "N"], index=1)

    with c3:
        include_retail_price = st.checkbox("소비자가 포함", value=True)

    with c4:
        include_supply_price = st.checkbox("공급가 포함", value=True)

    try:
        df = get_product_intro_export_data(
            brand_keyword=keyword.strip(),
            use_yn=use_yn,
        )
    except Exception as e:
        st.error(f"데이터 조회 중 오류 발생: {e}")
        return

    if df.empty:
        st.warning("조회된 상품이 없습니다.")
        return

    export_df = df.copy()

    drop_cols = []

    if not include_retail_price and "proposal_retail_price_krw" in export_df.columns:
        drop_cols.append("proposal_retail_price_krw")

    if not include_supply_price:
        if "supply_price_krw" in export_df.columns:
            drop_cols.append("supply_price_krw")
        if "supply_total_krw" in export_df.columns:
            drop_cols.append("supply_total_krw")

    if drop_cols:
        export_df = export_df.drop(columns=drop_cols, errors="ignore")

    preview_df = export_df.rename(columns=DISPLAY_RENAME).copy()

    desired_order = [
        "상품명", "사이즈", "Flavor", "Strength", "Length", "RG", "Time",
        "Guide", "소비자가(KRW)", "공급가(KRW)", "공급가합계(KRW)"
    ]
    preview_df = preview_df[[c for c in desired_order if c in preview_df.columns]]

    for col in ["소비자가(KRW)", "공급가(KRW)", "공급가합계(KRW)"]:
        if col in preview_df.columns:
            preview_df[col] = preview_df[col].apply(
                lambda x: f"₩{int(float(x)):,}" if str(x).strip() not in ("", "nan", "None") else ""
            )

    st.markdown("### 미리보기")
    st.dataframe(preview_df, use_container_width=True, hide_index=True)

    try:
        excel_bytes = build_product_intro_from_template(
            export_df,
            include_retail_price=include_retail_price,
            include_supply_price=include_supply_price,
        )
    except Exception as e:
        st.error(f"엑셀 생성 중 오류 발생: {e}")
        return

    today = datetime.now().strftime("%Y%m%d")
    file_name = f"상품소개서_발송용_{today}.xlsx"

    st.download_button(
        "상품소개서 다운로드",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# =========================================================
# 견적서 기능
# =========================================================
def _load_estimate_template_workbook(template_type: str = "A4"):
    path = TEMPLATE_PATH_A3 if template_type == "A3" else TEMPLATE_PATH
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"템플릿 파일이 없습니다: {path}\n"
            "프로젝트 폴더의 templates 아래에 파일을 넣어주세요."
        )
    return load_workbook(path)


def _clear_estimate_detail_rows(ws, layout: dict):
    """데이터 영역의 셀 값을 모두 비운다 (병합셀 좌상단 대표셀만 초기화)."""
    start_row = layout["data_start_row"]
    end_row = layout["data_end_row"]
    cols = [
        layout["col_display"],   # 품명/사이즈 병합영역 대표셀
        layout["col_qty"],
        layout["col_retail"],
        layout["col_supply"],
        layout["col_vat"],
        layout["col_total"],
    ]
    for row_idx in range(start_row, end_row + 1):
        for col in cols:
            ws[f"{col}{row_idx}"] = None


def _apply_estimate_row_style(ws, row_idx, layout: dict):
    # 품명/사이즈, 수량 컬럼: 가운데 정렬
    for addr in [f"{layout['col_display']}{row_idx}", f"{layout['col_qty']}{row_idx}"]:
        cell = ws[addr]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(name="맑은 고딕", size=10)

    # 소비자가 / 공급단가 / 부가세: 오른쪽 정렬
    for col in [layout["col_retail"], layout["col_supply"], layout["col_vat"]]:
        cell = ws[f"{col}{row_idx}"]
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
        cell.font = Font(name="맑은 고딕", size=10)

    # 합계 컬럼: 오른쪽 정렬 + 굵게
    cell_total = ws[f"{layout['col_total']}{row_idx}"]
    cell_total.alignment = Alignment(horizontal="right", vertical="center", wrap_text=False)
    cell_total.font = Font(name="맑은 고딕", size=10, bold=True)


def _apply_group_outer_border(ws, group_ranges, layout: dict):
    """같은 product_name 그룹의 외곽에 굵은 테두리를 적용한다.

    그룹 영역은 layout["border_columns"] 범위(견적 행의 표시 영역). 내부 행 사이는 기존 테두리 유지.
    """
    thick = Side(style="medium", color="000000")
    columns = layout["border_columns"]

    for group_start, group_end in group_ranges:
        for row_idx in range(group_start, group_end + 1):
            for col_letter in columns:
                cell = ws[f"{col_letter}{row_idx}"]
                existing = cell.border
                # 기존 테두리(내부 얇은 선) 보존
                left = existing.left if existing and existing.left else Side()
                right = existing.right if existing and existing.right else Side()
                top = existing.top if existing and existing.top else Side()
                bottom = existing.bottom if existing and existing.bottom else Side()

                # 외곽 위치에만 굵은 테두리 적용
                if col_letter == columns[0]:
                    left = thick
                if col_letter == columns[-1]:
                    right = thick
                if row_idx == group_start:
                    top = thick
                if row_idx == group_end:
                    bottom = thick

                cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _write_estimate_rows(ws, df: pd.DataFrame, layout: dict, use_proposal_retail_price=False):
    start_row = layout["data_start_row"]
    end_row = layout["data_end_row"]
    col_display = layout["col_display"]
    col_qty = layout["col_qty"]
    col_retail = layout["col_retail"]
    col_supply = layout["col_supply"]
    col_vat = layout["col_vat"]
    col_total = layout["col_total"]

    records = df.to_dict("records")
    max_count = end_row - start_row + 1

    if len(records) > max_count:
        records = records[:max_count]

    # 그룹 묶음 판단: source_row_no의 가장 앞자리(맨 앞 숫자)가 같은 행끼리 한 그룹
    group_ranges = []  # [(group_start_row, group_end_row), ...]
    prev_group_key = object()  # 첫 행과 반드시 달라지도록 sentinel
    group_start = None

    def _first_digit(value):
        """source_row_no의 가장 앞 두 자리 숫자를 반환. 추출 불가 시 None.

        예: 1881 -> '18', 1234 -> '12', 9 -> '9' (한 자리만 있으면 한 자리 그대로),
            'A12' 같은 경우 숫자만 모아서 앞 두 자리 -> '12'
        """
        if value is None:
            return None
        try:
            if isinstance(value, float) and pd.isna(value):
                return None
        except Exception:
            pass
        s = str(value).strip()
        if not s:
            return None
        digits = "".join(ch for ch in s if ch.isdigit())
        if not digits:
            return None
        return digits[:2]

    for row_idx, row in enumerate(records, start=start_row):
        product_name = _safe_value(row.get("product_name", ""))
        size_name = _safe_value(row.get("size_name", ""))

        # 수량: None / NaN 만 0으로 변환 (실제 0은 0으로 유지)
        qty_raw = row.get("qty", 0)
        try:
            qty = int(qty_raw) if qty_raw not in (None, "") and not (isinstance(qty_raw, float) and pd.isna(qty_raw)) else 0
        except (ValueError, TypeError):
            qty = 0

        retail_price = row.get("retail_price_krw", 0) or 0
        if use_proposal_retail_price:
            proposal_price = row.get("proposal_retail_price_krw", None)
            if proposal_price not in (None, "", 0):
                retail_price = proposal_price
        supply_price = row.get("estimate_supply_price_krw", None)
        if supply_price in (None, "", 0):
            supply_price = row.get("supply_price_krw", 0) or 0

        display_name = f"{product_name} / {size_name}" if str(size_name).strip() else str(product_name)

        ws[f"{col_display}{row_idx}"] = display_name
        ws[f"{col_qty}{row_idx}"] = qty
        ws[f"{col_retail}{row_idx}"] = retail_price
        ws[f"{col_supply}{row_idx}"] = supply_price
        ws[f"{col_vat}{row_idx}"] = f"=ROUND({col_qty}{row_idx}*{col_supply}{row_idx}*0.1,0)"
        ws[f"{col_total}{row_idx}"] = f"={col_qty}{row_idx}*{col_supply}{row_idx}+{col_vat}{row_idx}"

        for col in [col_retail, col_supply, col_vat, col_total]:
            ws[f"{col}{row_idx}"].number_format = '₩#,##0'

        _apply_estimate_row_style(ws, row_idx, layout)

        # 그룹 추적: source_row_no의 가장 앞자리가 바뀌면 이전 그룹 종료
        group_key = _first_digit(row.get("source_row_no"))
        if group_key != prev_group_key:
            if group_start is not None:
                group_ranges.append((group_start, row_idx - 1))
            group_start = row_idx
            prev_group_key = group_key

    # 마지막 그룹 종료 처리
    if group_start is not None and records:
        group_ranges.append((group_start, start_row + len(records) - 1))

    # 굵은 외곽 테두리는 그룹 키가 None이 아닌(=유효한 source_row_no가 있는) 그룹에만 적용
    valid_group_ranges = []
    for gs, ge in group_ranges:
        # 그룹 시작 행에 해당하는 record를 찾아 group_key 재확인
        rec_idx = gs - start_row
        if 0 <= rec_idx < len(records):
            key = _first_digit(records[rec_idx].get("source_row_no"))
            if key is not None:
                valid_group_ranges.append((gs, ge))

    if valid_group_ranges:
        _apply_group_outer_border(ws, valid_group_ranges, layout)


def _set_estimate_total_formulas(ws, layout: dict):
    start_row = layout["data_start_row"]
    end_row = layout["data_end_row"]
    total_row = layout["total_row"]
    col_qty = layout["col_qty"]
    col_supply = layout["col_supply"]
    col_vat = layout["col_vat"]
    col_total = layout["col_total"]

    ws[f"{col_supply}{total_row}"] = (
        f"=SUMPRODUCT({col_qty}{start_row}:{col_qty}{end_row},"
        f"{col_supply}{start_row}:{col_supply}{end_row})"
    )
    ws[f"{col_vat}{total_row}"] = f"=SUM({col_vat}{start_row}:{col_vat}{end_row})"
    ws[f"{col_total}{total_row}"] = f"=SUM({col_total}{start_row}:{col_total}{end_row})"

    for col in (col_supply, col_vat, col_total):
        cell = ws[f"{col}{total_row}"]
        cell.number_format = '₩#,##0'
        # 합계 굵게 (기존 폰트 속성 유지하면서 bold만 추가)
        existing_font = cell.font
        cell.font = Font(
            name=existing_font.name or "맑은 고딕",
            size=existing_font.size or 10,
            bold=True,
            color=existing_font.color,
        )


def _update_estimate_amount_text(ws, df: pd.DataFrame, layout: dict):
    total_supply = 0
    if not df.empty:
        qty = pd.to_numeric(df["qty"], errors="coerce").fillna(0)

        if "estimate_supply_price_krw" in df.columns:
            supply = pd.to_numeric(df["estimate_supply_price_krw"], errors="coerce").fillna(0)
        else:
            supply = pd.to_numeric(df["supply_price_krw"], errors="coerce").fillna(0)

        total_supply = int((qty * supply).sum())

    ws[layout["amount_text_cell"]] = (
        f" 금 액 (견적금액) : {total_supply:,.0f} 원 (V.A.T. 제외)"
    )


def _fill_estimate_header(ws, partner_info: dict, layout: dict, show_grade_discount: bool = False):
    today_str = datetime.now().strftime("%Y.%m.%d")
    ws[layout["date_cell"]] = today_str

    partner_name = str(partner_info.get("partner_name", "") or "").strip()
    address = str(partner_info.get("address", "") or "").strip()
    owner_name = str(partner_info.get("owner_name", "") or "").strip()
    contact_name = str(partner_info.get("contact_name", "") or "").strip()
    phone = str(partner_info.get("phone", "") or "").strip()

    ws[layout["partner_name_cell_1"]] = partner_name
    ws[layout["partner_name_cell_2"]] = partner_name
    ws[layout["address_cell"]] = address

    if owner_name:
        ws[layout["contact_cell"]] = f"{owner_name} 대표님"
    elif contact_name:
        ws[layout["contact_cell"]] = contact_name
    else:
        ws[layout["contact_cell"]] = f"{partner_name} 담당자"

    ws[layout["phone_cell"]] = phone

    # 등급/할인율 옵션 표기
    if show_grade_discount:
        grade_code = str(partner_info.get("grade_code", "") or "-").strip()
        discount_rate = float(partner_info.get("estimate_discount_rate", 0) or 0)
        discount_pct = int(round(discount_rate * 100))
        ws[layout["grade_discount_cell"]] = f"Grade: {grade_code}(할인율: {discount_pct}%)"
    else:
        ws[layout["grade_discount_cell"]] = None

def build_estimate_workbook(
    cigar_df: pd.DataFrame,
    non_cigar_df: pd.DataFrame,
    partner_info: dict,
    use_proposal_retail_price: bool = False,
    show_grade_discount: bool = False,
    template_type: str = "A4",
) -> io.BytesIO:
    if template_type not in ESTIMATE_LAYOUTS:
        raise ValueError(f"지원하지 않는 견적서 템플릿 종류: {template_type}")

    layout = ESTIMATE_LAYOUTS[template_type]
    wb = _load_estimate_template_workbook(template_type=template_type)

    sheet_cigar_name = layout["sheet_cigar"]
    sheet_non_cigar_name = layout["sheet_non_cigar"]

    if sheet_cigar_name not in wb.sheetnames:
        raise ValueError(
            f"템플릿에 '{sheet_cigar_name}' 시트가 없습니다. (template_type={template_type})"
        )
    ws_cigar = wb[sheet_cigar_name]

    if sheet_non_cigar_name in wb.sheetnames:
        ws_non_cigar = wb[sheet_non_cigar_name]
    else:
        # 시가 외 시트가 없는 경우에만 시가 시트를 복사해서 만들고, 라벨을 OTHERS로 설정
        ws_non_cigar = wb.copy_worksheet(ws_cigar)
        ws_non_cigar.title = sheet_non_cigar_name
        ws_non_cigar[layout["label_cell"]] = "OTHERS"

    # 시가 시트 라벨은 항상 CIGAR로 통일 (템플릿 원본도 CIGAR)
    ws_cigar[layout["label_cell"]] = "CIGAR"
    # 시가 외 시트 라벨은 템플릿 원본('OTHERS')을 그대로 유지하므로 덮어쓰지 않음

    for ws, df in [(ws_cigar, cigar_df), (ws_non_cigar, non_cigar_df)]:
        _fill_estimate_header(ws, partner_info, layout, show_grade_discount=show_grade_discount)
        _clear_estimate_detail_rows(ws, layout)
        _write_estimate_rows(
            ws,
            df,
            layout,
            use_proposal_retail_price=use_proposal_retail_price,
        )
        _set_estimate_total_formulas(ws, layout)
        _update_estimate_amount_text(ws, df, layout)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def _build_sample_estimate_data():
    cigar_df = pd.DataFrame([
        {
            "product_name": "TABACALERA",
            "size_name": "Robusto",
            "qty": 3,
            "retail_price_krw": 18000,
            "supply_price_krw": 12000,
        },
        {
            "product_name": "1881 Perique",
            "size_name": "Corona",
            "qty": 2,
            "retail_price_krw": 22000,
            "supply_price_krw": 15000,
        },
    ])

    non_cigar_df = pd.DataFrame([
        {
            "product_name": "시가 커터",
            "size_name": "",
            "qty": 2,
            "retail_price_krw": 25000,
            "supply_price_krw": 18000,
        },
        {
            "product_name": "애쉬트레이",
            "size_name": "",
            "qty": 1,
            "retail_price_krw": 45000,
            "supply_price_krw": 32000,
        },
    ])

    partner_info = {
        "partner_name": "데일리시가 파트너 샘플",
        "address": "경기도 수원시 영통구 예시로 123",
        "owner_name": "홍길동",
        "phone": "010-1234-5678",
    }
    return cigar_df, non_cigar_df, partner_info



def _sort_estimate_editor_df(df: pd.DataFrame, is_non_cigar: bool = False) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()

    # 시가 외: source_row_no 기준 정렬
    if is_non_cigar:
        if "source_row_no" in work.columns:
            work["source_row_no"] = pd.to_numeric(work["source_row_no"], errors="coerce")
            work = work.sort_values(
                by=["source_row_no"],
                ascending=[True],
                kind="stable",
                na_position="last",
            )
        return work

    # 시가: source_row_no 기준 정렬 (없으면 product_name → size_name → product_code 순)
    if "source_row_no" in work.columns:
        work["source_row_no"] = pd.to_numeric(work["source_row_no"], errors="coerce")
        work = work.sort_values(
            by=["source_row_no"],
            ascending=[True],
            kind="stable",
            na_position="last",
        )
    else:
        sort_cols = [c for c in ["product_name", "size_name", "product_code"] if c in work.columns]
        if sort_cols:
            work = work.sort_values(
                by=sort_cols,
                ascending=[True] * len(sort_cols),
                kind="stable",
                na_position="last",
            )
    return work


def _get_estimate_editor_column_config(is_non_cigar: bool = False):
    config = {
        "product_code": st.column_config.TextColumn("상품코드", disabled=True),
        "product_name": st.column_config.TextColumn("상품명", disabled=True),
        "size_name": st.column_config.TextColumn("사이즈", disabled=True),
        "retail_price_krw": st.column_config.NumberColumn("소비자가", format="₩%.0f", disabled=True),
        "proposal_retail_price_krw": st.column_config.NumberColumn("제안소비자가", format="₩%.0f", disabled=True),
        "supply_price_krw": st.column_config.NumberColumn("원공급가", format="₩%.0f", disabled=True),
        "qty": st.column_config.NumberColumn("수량", min_value=0, step=1),
    }

    if not is_non_cigar:
        config["estimate_discount_rate_text"] = st.column_config.TextColumn("할인율", disabled=True)
        config["estimate_supply_price_krw"] = st.column_config.NumberColumn("견적공급가", format="₩%.0f", disabled=True)

    return config

def _apply_partner_grade_discount(df: pd.DataFrame, discount_rate: float) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()
    work["estimate_discount_rate"] = float(discount_rate or 0)

    base_supply = pd.to_numeric(work.get("supply_price_krw", 0), errors="coerce").fillna(0)
    work["base_supply_price_krw"] = base_supply
    work["estimate_supply_price_krw"] = (
        base_supply * (1 - work["estimate_discount_rate"])
    ).round(0).astype(int)

    return work

def render_estimate_export():
    st.subheader("견적서 엑셀 다운로드")
    st.caption("견적서 템플릿에 시가 / 시가 외 품목을 나누어 채워 다운로드합니다.")

    partner_df = get_all_partner_for_select()
    if partner_df.empty:
        st.warning("등록된 거래처가 없습니다.")
        return

    partner_options = {
        f"{row['partner_name']}": row["partner_id"]
        for _, row in partner_df.iterrows()
    }

    selected_partner_name = st.selectbox("거래처 선택", list(partner_options.keys()))
    selected_partner_id = partner_options[selected_partner_name]

    partner_info = get_partner_detail_by_id(selected_partner_id)
    estimate_discount_rate = float(partner_info.get("estimate_discount_rate", 0) or 0)

    st.caption(
        f"현재 등급: {partner_info.get('grade_code', '-')}"
        f" / 시가 견적 할인율: {int(estimate_discount_rate * 100)}%"
    )

    use_proposal_retail_price = st.checkbox(
        "견적서 소비자가를 제안소비자가로 대체",
        value=True,
        help="체크 시 견적서의 소비자가(F열)에 기본 소비자가 대신 제안소비자가를 표시합니다."
    )

    show_grade_discount = st.checkbox(
        "견적서에 등급/할인율 표기",
        value=False,
        help="체크 시 엑셀 H8 셀에 Grade: 등급(할인율: 5%) 형식으로 표시합니다."
    )

    only_in_stock = st.checkbox(
        "재고 있는 시가만 표시",
        value=False,
        help="체크 시 현재고가 1개 이상인 품목만 목록에 표시합니다. 미래 수입 예정 재고는 제외됩니다.",
    )

    cigar_master_df = get_estimate_cigar_items(only_in_stock=only_in_stock)
    non_cigar_master_df = get_estimate_non_cigar_items()

    # 시가 외: non_cigar_product_mst.wholesale_price 가 있는(>0) 항목만 표시
    if not non_cigar_master_df.empty:
        if "wholesale_price" in non_cigar_master_df.columns:
            wholesale_series = pd.to_numeric(
                non_cigar_master_df["wholesale_price"], errors="coerce"
            ).fillna(0)
            non_cigar_master_df = non_cigar_master_df[wholesale_series > 0].copy()
        else:
            st.warning(
                "시가 외 데이터에 wholesale_price 컬럼이 없습니다. "
                "db.get_estimate_non_cigar_items()가 wholesale_price를 반환하도록 수정이 필요합니다."
            )

    st.markdown("### 시가")
    if cigar_master_df.empty:
        st.info("조회 가능한 시가 품목이 없습니다.")
        cigar_df = pd.DataFrame(columns=[
            "product_code", "product_name", "size_name", "qty",
            "retail_price_krw", "proposal_retail_price_krw",
            "supply_price_krw", "estimate_supply_price_krw"
        ])
    else:
        cigar_edit_df = cigar_master_df.copy()
        cigar_edit_df = _sort_estimate_editor_df(cigar_edit_df, is_non_cigar=False)
        cigar_edit_df = _apply_partner_grade_discount(cigar_edit_df, estimate_discount_rate)

        # 디폴트 수량: 현재고가 1개 이상이면 1, 그렇지 않으면 0
        if "current_stock" in cigar_edit_df.columns:
            stock_series = pd.to_numeric(cigar_edit_df["current_stock"], errors="coerce").fillna(0)
            cigar_edit_df["qty"] = (stock_series > 0).astype(int)
        else:
            cigar_edit_df["qty"] = 0

        cigar_edit_df["estimate_discount_rate_text"] = (
            pd.to_numeric(cigar_edit_df["estimate_discount_rate"], errors="coerce")
            .fillna(0)
            .mul(100)
            .round(0)
            .astype(int)
            .astype(str) + "%"
        )

        cigar_column_order = [
            c for c in [
                "source_row_no",
                "product_code",
                "product_name",
                "size_name",
                "current_stock",
                "retail_price_krw",
                "proposal_retail_price_krw",
                "supply_price_krw",
                "estimate_discount_rate_text",
                "estimate_supply_price_krw",
                "qty",
            ]
            if c in cigar_edit_df.columns
        ]

        # 디버그: source_row_no가 보이지 않거나 정렬이 이상해 보일 때 켜서 확인
        debug_show_row_no = st.checkbox(
            "원본 행번호(source_row_no) 표시 [디버그]",
            value=False,
            help="정렬 기준이 되는 source_row_no 값을 표시합니다. NaN으로 보이는 항목이 맨 뒤로 밀립니다.",
            key="cigar_debug_show_row_no",
        )
        if not debug_show_row_no and "source_row_no" in cigar_column_order:
            cigar_column_order = [c for c in cigar_column_order if c != "source_row_no"]

        cigar_column_config = _get_estimate_editor_column_config(is_non_cigar=False)
        cigar_column_config["current_stock"] = st.column_config.NumberColumn(
            "현재고", disabled=True, format="%d개"
        )
        cigar_column_config["source_row_no"] = st.column_config.NumberColumn(
            "원본행번호", disabled=True, format="%d"
        )

        edited_cigar_df = st.data_editor(
            cigar_edit_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=cigar_column_config,
            column_order=cigar_column_order,
        )
        # 수량 0인 항목도 엑셀에 포함하여 export
        cigar_df = edited_cigar_df.copy()

    st.markdown("### 시가 외")
    if non_cigar_master_df.empty:
        st.info("조회 가능한 시가 외 품목이 없습니다.")
        non_cigar_df = pd.DataFrame(columns=[
            "product_code", "product_name", "size_name", "qty",
            "retail_price_krw", "proposal_retail_price_krw",
            "supply_price_krw"
        ])
    else:
        non_cigar_edit_df = non_cigar_master_df.copy()
        non_cigar_edit_df = _sort_estimate_editor_df(non_cigar_edit_df, is_non_cigar=True)
        non_cigar_edit_df["qty"] = 0

        non_cigar_column_order = [
            c for c in [
                "source_row_no",
                "product_code",
                "product_name",
                "size_name",
                "retail_price_krw",
                "proposal_retail_price_krw",
                "supply_price_krw",
                "qty",
            ]
            if c in non_cigar_edit_df.columns
        ]

        # 디버그: 시가 외도 source_row_no 확인용
        debug_show_row_no_nc = st.checkbox(
            "원본 행번호(source_row_no) 표시 [디버그-시가 외]",
            value=False,
            help="시가 외 데이터의 source_row_no 값을 표시합니다.",
            key="non_cigar_debug_show_row_no",
        )
        if not debug_show_row_no_nc and "source_row_no" in non_cigar_column_order:
            non_cigar_column_order = [c for c in non_cigar_column_order if c != "source_row_no"]

        non_cigar_column_config = _get_estimate_editor_column_config(is_non_cigar=True)
        non_cigar_column_config["source_row_no"] = st.column_config.NumberColumn(
            "원본행번호", disabled=True, format="%d"
        )

        edited_non_cigar_df = st.data_editor(
            non_cigar_edit_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=non_cigar_column_config,
            column_order=non_cigar_column_order,
        )
        # 수량 0인 항목도 엑셀에 포함하여 export
        non_cigar_df = edited_non_cigar_df.copy()

    if cigar_df.empty and non_cigar_df.empty:
        st.info("수량을 입력한 품목이 없습니다.")
        return

    download_cigar_df = cigar_df.copy()
    download_non_cigar_df = non_cigar_df.copy()

    if use_proposal_retail_price:
        if "proposal_retail_price_krw" in download_cigar_df.columns:
            download_cigar_df["retail_price_krw"] = (
                pd.to_numeric(download_cigar_df["proposal_retail_price_krw"], errors="coerce")
                .fillna(pd.to_numeric(download_cigar_df["retail_price_krw"], errors="coerce"))
            )

        if "proposal_retail_price_krw" in download_non_cigar_df.columns:
            download_non_cigar_df["retail_price_krw"] = (
                pd.to_numeric(download_non_cigar_df["proposal_retail_price_krw"], errors="coerce")
                .fillna(pd.to_numeric(download_non_cigar_df["retail_price_krw"], errors="coerce"))
            )

    # 파일명: 견적서_상호명_날짜4자리(MMDD).xlsx
    today4 = datetime.now().strftime("%m%d")

    raw_partner_name = str(partner_info.get("partner_name") or selected_partner_name or "").strip()
    # & 제거 + 파일명 금지문자(/ \ : * ? " < > |) 제거 + 공백 정리
    safe_partner_name = raw_partner_name.replace("&", "")
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        safe_partner_name = safe_partner_name.replace(ch, "")
    safe_partner_name = " ".join(safe_partner_name.split()).strip()

    def _file_name_for(template_type: str) -> str:
        suffix = "_A3" if template_type == "A3" else ""
        if safe_partner_name:
            return f"견적서_{safe_partner_name}_{today4}{suffix}.xlsx"
        return f"견적서_{today4}{suffix}.xlsx"

    def _build_bytes_for(template_type: str) -> io.BytesIO:
        return build_estimate_workbook(
            download_cigar_df,
            download_non_cigar_df,
            partner_info,
            use_proposal_retail_price=use_proposal_retail_price,
            show_grade_discount=show_grade_discount,
            template_type=template_type,
        )

    col_a4, col_a3 = st.columns(2)

    mime_xlsx = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    with col_a4:
        try:
            a4_bytes = _build_bytes_for("A4")
            st.download_button(
                "견적서 다운로드 (A4)",
                data=a4_bytes,
                file_name=_file_name_for("A4"),
                mime=mime_xlsx,
                use_container_width=True,
                key="estimate_download_a4",
            )
        except FileNotFoundError as e:
            st.error(f"A4 템플릿 오류: {e}")
        except Exception as e:
            st.error(f"A4 견적서 생성 중 오류: {e}")

    with col_a3:
        try:
            a3_bytes = _build_bytes_for("A3")
            st.download_button(
                "견적서 다운로드 (A3)",
                data=a3_bytes,
                file_name=_file_name_for("A3"),
                mime=mime_xlsx,
                use_container_width=True,
                key="estimate_download_a3",
            )
        except FileNotFoundError as e:
            st.error(f"A3 템플릿 오류: {e}")
        except Exception as e:
            st.error(f"A3 견적서 생성 중 오류: {e}")

def render():
    st.title("문서출력")
    tab1, tab2, tab3 = st.tabs(["상품소개서", "견적서", "기타"])

    with tab1:
        render_product_intro_export()

    with tab2:
        render_estimate_export()

    with tab3:
        st.info("향후 거래명세서 / 가격표 / 파트너 제안서 등을 추가할 수 있습니다.")