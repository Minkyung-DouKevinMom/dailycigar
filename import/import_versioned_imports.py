import sqlite3
import openpyxl
import json
import os
import re
from collections import defaultdict

DB_PATH = r"C:\DAILYCIGAR_DB\cigar.db"
EXCEL_PATH = r"C:\DAILYCIGAR_DB\버전별상세수입내용.xlsx"
SHEET_NAME = "가격분석"

print("DB 파일 존재:", os.path.exists(DB_PATH))
print("엑셀 파일 존재:", os.path.exists(EXCEL_PATH))

# --------------------------------------------------
# 유틸
# --------------------------------------------------
def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None

def to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except Exception:
        return None

def to_int(v):
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except Exception:
        return None

def json_safe(v):
    if v is None:
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)

def infer_usd_to_krw_rate(import_unit_cost_krw, export_unit_price_usd):
    if import_unit_cost_krw and export_unit_price_usd and export_unit_price_usd != 0:
        return import_unit_cost_krw / export_unit_price_usd
    return None

def infer_php_to_krw_rate(local_unit_price_krw, local_unit_price_php):
    if local_unit_price_krw and local_unit_price_php and local_unit_price_php != 0:
        return local_unit_price_krw / local_unit_price_php
    return None

# --------------------------------------------------
# 워크북 로드
# data_only=False : 수식 원본
# data_only=True  : 계산 결과
# --------------------------------------------------
wb_formula = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
wb_value = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

wsf = wb_formula[SHEET_NAME]
wsv = wb_value[SHEET_NAME]

# 실제 헤더는 2행
headers = [wsv.cell(2, col).value for col in range(1, 33)]
headers = [str(h).strip() if h is not None else f"COL_{i+1}" for i, h in enumerate(headers)]

# --------------------------------------------------
# DB 연결
# --------------------------------------------------
conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()
cur.execute("PRAGMA foreign_keys = ON;")

# --------------------------------------------------
# 1. tax_rule 등록/재사용
# 시트 수식 기준:
# 개별소비세 = 무게 * 61
# 담배소비세 = 무게 * 103
# 지방교육세 = 담배소비세 * 0.4399
# 국민건강 = 무게 * 85.8
# 부가세 = (수입원가 + 개별소비세) * 0.1  <-- 시트 기준 그대로 저장
# --------------------------------------------------
RULE_NAME = "가격분석 시트 기준 기본세율"
EFFECTIVE_FROM = "2025-01-01"

cur.execute("""
    SELECT id
    FROM tax_rule
    WHERE rule_name = ?
      AND effective_from = ?
""", (RULE_NAME, EFFECTIVE_FROM))
row = cur.fetchone()

if row:
    tax_rule_id = row[0]
else:
    cur.execute("""
        INSERT INTO tax_rule (
            rule_name,
            effective_from,
            effective_to,
            individual_tax_per_g,
            tobacco_tax_per_g,
            local_education_rate,
            health_charge_per_g,
            import_vat_rate,
            notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        RULE_NAME,
        EFFECTIVE_FROM,
        None,
        61.0,
        103.0,
        0.4399,
        85.8,
        0.10,
        "버전별상세수입내용.xlsx 가격분석 시트 수식 기준"
    ))
    tax_rule_id = cur.lastrowid

# --------------------------------------------------
# 2. 행 읽기
# --------------------------------------------------
items_by_version = defaultdict(list)

for row_no in range(3, wsv.max_row + 1):
    version_name = clean_text(wsv.cell(row_no, 1).value)
    product_name = clean_text(wsv.cell(row_no, 2).value)
    size_name = clean_text(wsv.cell(row_no, 3).value)

    # 완전 빈 행 스킵
    if not version_name and not product_name and not size_name:
        continue

    product_code = clean_text(wsv.cell(row_no, 4).value)

    export_box_price_usd = to_float(wsv.cell(row_no, 5).value)
    discounted_box_price_usd = to_float(wsv.cell(row_no, 6).value)
    discount_rate = to_float(wsv.cell(row_no, 7).value)
    import_unit_qty = to_int(wsv.cell(row_no, 8).value)
    export_unit_price_usd = to_float(wsv.cell(row_no, 9).value)
    import_unit_cost_krw = to_float(wsv.cell(row_no, 10).value)
    import_total_cost_krw = to_float(wsv.cell(row_no, 11).value)

    unit_weight_g = to_float(wsv.cell(row_no, 12).value)
    total_weight_g = to_float(wsv.cell(row_no, 13).value)

    individual_tax_krw = to_float(wsv.cell(row_no, 14).value)
    tobacco_tax_krw = to_float(wsv.cell(row_no, 15).value)
    local_education_tax_krw = to_float(wsv.cell(row_no, 16).value)
    health_charge_krw = to_float(wsv.cell(row_no, 17).value)
    import_vat_krw = to_float(wsv.cell(row_no, 18).value)
    tax_total_krw = to_float(wsv.cell(row_no, 19).value)
    tax_total_all_krw = to_float(wsv.cell(row_no, 20).value)
    korea_cost_krw = to_float(wsv.cell(row_no, 21).value)

    local_box_price_php = to_float(wsv.cell(row_no, 22).value)
    local_unit_price_php = to_float(wsv.cell(row_no, 23).value)
    local_unit_price_krw = to_float(wsv.cell(row_no, 24).value)

    retail_price_krw = to_float(wsv.cell(row_no, 25).value)
    supply_price_krw = to_float(wsv.cell(row_no, 26).value)
    supply_vat_krw = to_float(wsv.cell(row_no, 27).value)
    supply_total_krw = to_float(wsv.cell(row_no, 28).value)
    margin_krw = to_float(wsv.cell(row_no, 29).value)
    retail_margin_rate = to_float(wsv.cell(row_no, 30).value)
    wholesale_margin_rate = to_float(wsv.cell(row_no, 31).value)
    store_retail_price_krw = to_float(wsv.cell(row_no, 32).value)

    # 원본값 JSON
    raw_row = {}
    raw_formula = {}
    for col_idx, header in enumerate(headers, start=1):
        raw_row[header] = json_safe(wsv.cell(row_no, col_idx).value)
        raw_formula[header] = json_safe(wsf.cell(row_no, col_idx).value)

    item = {
        "version_name": version_name,
        "product_name": product_name,
        "size_name": size_name,
        "product_code": product_code,

        "export_box_price_usd": export_box_price_usd,
        "discounted_box_price_usd": discounted_box_price_usd,
        "discount_rate": discount_rate,
        "import_unit_qty": import_unit_qty,
        "export_unit_price_usd": export_unit_price_usd,
        "import_unit_cost_krw": import_unit_cost_krw,
        "import_total_cost_krw": import_total_cost_krw,

        "unit_weight_g": unit_weight_g,
        "total_weight_g": total_weight_g,

        "individual_tax_krw": individual_tax_krw,
        "tobacco_tax_krw": tobacco_tax_krw,
        "local_education_tax_krw": local_education_tax_krw,
        "health_charge_krw": health_charge_krw,
        "import_vat_krw": import_vat_krw,
        "tax_total_krw": tax_total_krw,
        "tax_total_all_krw": tax_total_all_krw,
        "korea_cost_krw": korea_cost_krw,

        "local_box_price_php": local_box_price_php,
        "local_unit_price_php": local_unit_price_php,
        "local_unit_price_krw": local_unit_price_krw,

        "retail_price_krw": retail_price_krw,
        "supply_price_krw": supply_price_krw,
        "supply_vat_krw": supply_vat_krw,
        "supply_total_krw": supply_total_krw,
        "margin_krw": margin_krw,
        "retail_margin_rate": retail_margin_rate,
        "wholesale_margin_rate": wholesale_margin_rate,
        "store_retail_price_krw": store_retail_price_krw,

        "source_row_no": row_no,
        "raw_row_json": json.dumps(raw_row, ensure_ascii=False),
        "raw_formula_json": json.dumps(raw_formula, ensure_ascii=False),
    }

    items_by_version[version_name].append(item)

# --------------------------------------------------
# 3. 버전별 batch 생성/갱신
# --------------------------------------------------
for version_name, items in items_by_version.items():
    total_item_count = len(items)
    total_unit_qty = sum((x["import_unit_qty"] or 0) for x in items)
    total_weight_g = sum((x["total_weight_g"] or 0) for x in items)
    total_amount_usd = sum((x["export_unit_price_usd"] or 0) * (x["import_unit_qty"] or 0) for x in items)
    total_amount_krw = sum((x["import_total_cost_krw"] or 0) for x in items)

    # 버전 내 첫 번째 유효값 기준
    usd_to_krw_rate = None
    php_to_krw_rate = None
    local_markup_rate = None

    for x in items:
        usd_to_krw_rate = infer_usd_to_krw_rate(x["import_unit_cost_krw"], x["export_unit_price_usd"])
        if usd_to_krw_rate:
            break

    for x in items:
        php_to_krw_rate = infer_php_to_krw_rate(x["local_unit_price_krw"], x["local_unit_price_php"])
        if php_to_krw_rate:
            break

    # local_unit_price_krw = local_unit_price_php * 25 * 1.12 형태가 보임
    # php_to_krw_rate / 25 로 대략 마크업 추정 가능
    if php_to_krw_rate:
        local_markup_rate = php_to_krw_rate / 25.0

    cur.execute("""
        INSERT INTO import_batch (
            version_name,
            import_date,
            supplier_name,
            tax_rule_id,
            usd_to_krw_rate,
            php_to_krw_rate,
            local_markup_rate,
            total_item_count,
            total_unit_qty,
            total_weight_g,
            total_amount_usd,
            total_amount_krw,
            notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(version_name) DO UPDATE SET
            tax_rule_id = excluded.tax_rule_id,
            usd_to_krw_rate = excluded.usd_to_krw_rate,
            php_to_krw_rate = excluded.php_to_krw_rate,
            local_markup_rate = excluded.local_markup_rate,
            total_item_count = excluded.total_item_count,
            total_unit_qty = excluded.total_unit_qty,
            total_weight_g = excluded.total_weight_g,
            total_amount_usd = excluded.total_amount_usd,
            total_amount_krw = excluded.total_amount_krw,
            notes = excluded.notes
    """, (
        version_name,
        None,
        None,
        tax_rule_id,
        usd_to_krw_rate,
        php_to_krw_rate,
        local_markup_rate,
        total_item_count,
        total_unit_qty,
        total_weight_g,
        total_amount_usd,
        total_amount_krw,
        "버전별상세수입내용.xlsx / 가격분석 시트 기준 집계"
    ))

    cur.execute("SELECT id FROM import_batch WHERE version_name = ?", (version_name,))
    batch_id = cur.fetchone()[0]

    # 기존 동일 version 데이터 삭제 후 재적재
    cur.execute("DELETE FROM import_item WHERE batch_id = ?", (batch_id,))

    # 상세 입력
    for x in items:
        cur.execute("""
    INSERT INTO import_item (
        batch_id,
        product_name,
        size_name,
        product_code,

        export_box_price_usd,
        discounted_box_price_usd,
        discount_rate,
        import_unit_qty,
        export_unit_price_usd,
        import_unit_cost_krw,
        import_total_cost_krw,

        unit_weight_g,
        total_weight_g,

        individual_tax_krw,
        tobacco_tax_krw,
        local_education_tax_krw,
        health_charge_krw,
        import_vat_krw,
        tax_total_krw,
        tax_total_all_krw,
        korea_cost_krw,

        local_box_price_php,
        local_unit_price_php,
        local_unit_price_krw,

        retail_price_krw,
        supply_price_krw,
        supply_vat_krw,
        supply_total_krw,
        margin_krw,
        retail_margin_rate,
        wholesale_margin_rate,
        store_retail_price_krw,

        source_row_no,
        raw_row_json,
        raw_formula_json
    ) VALUES (
        ?, ?, ?, ?,
        ?, ?, ?, ?, ?, ?, ?,
        ?, ?,
        ?, ?, ?, ?, ?, ?, ?, ?,
        ?, ?, ?,
        ?, ?, ?, ?, ?, ?, ?, ?,
        ?, ?, ?
    )
""", (
    batch_id,
    x["product_name"],
    x["size_name"],
    x["product_code"],

    x["export_box_price_usd"],
    x["discounted_box_price_usd"],
    x["discount_rate"],
    x["import_unit_qty"],
    x["export_unit_price_usd"],
    x["import_unit_cost_krw"],
    x["import_total_cost_krw"],

    x["unit_weight_g"],
    x["total_weight_g"],

    x["individual_tax_krw"],
    x["tobacco_tax_krw"],
    x["local_education_tax_krw"],
    x["health_charge_krw"],
    x["import_vat_krw"],
    x["tax_total_krw"],
    x["tax_total_all_krw"],
    x["korea_cost_krw"],

    x["local_box_price_php"],
    x["local_unit_price_php"],
    x["local_unit_price_krw"],

    x["retail_price_krw"],
    x["supply_price_krw"],
    x["supply_vat_krw"],
    x["supply_total_krw"],
    x["margin_krw"],
    x["retail_margin_rate"],
    x["wholesale_margin_rate"],
    x["store_retail_price_krw"],

    x["source_row_no"],
    x["raw_row_json"],
    x["raw_formula_json"]

        ))

conn.commit()
conn.close()

print("완료: tax_rule / import_batch / import_item 입력이 끝났습니다.")